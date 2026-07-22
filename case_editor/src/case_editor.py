from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
import re
import json
from PyQt6.QtCore import Qt, pyqtSignal, QStringListModel, QSignalBlocker, QTimer, QPointF, QRectF
from PyQt6.QtGui import QFont, QKeyEvent, QPen, QBrush, QColor, QPainter, QPainterPath
from PyQt6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QCompleter,
    QDialog,
    QDialogButtonBox,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSpinBox,
    QSplitter,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QFormLayout,
    QGraphicsView,
    QGraphicsScene,
    QGraphicsItem,
    QGraphicsEllipseItem,
    QGraphicsRectItem,
    QGraphicsTextItem,
    QGraphicsPathItem,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QSizePolicy,
    QStyledItemDelegate,
    QAbstractItemDelegate
)


class ReadOnlyLineEdit(QLineEdit):
    """只读文本输入框，不支持粘贴操作"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
    
    def keyPressEvent(self, event: QKeyEvent) -> None:
        """按键事件处理，阻止粘贴操作"""
        # 检查是否是 Ctrl+V 粘贴操作
        if event.key() == Qt.Key.Key_V and event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            # 阻止粘贴操作
            return
        # 其他按键正常处理
        super().keyPressEvent(event)


# ----------------------------
# 布局对齐参数
# ----------------------------
FORM_LABEL_WIDTH = 80
FORM_HSPACING = 6
FORM_VSPACING = 6


def _setup_groupbox_style(group: QGroupBox) -> None:
    group.setStyleSheet(
        """
        QGroupBox {
            font-weight: bold;
            font-size: 14px;
            border: 2px solid #cccccc;
            border-radius: 5px;
            margin-top: 10px;
            padding-top: 10px;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px;
        }
        """
    )


def _align_form_layout(form_layout: QFormLayout) -> None:
    form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
    form_layout.setFormAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
    form_layout.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.AllNonFixedFieldsGrow)
    form_layout.setHorizontalSpacing(FORM_HSPACING)
    form_layout.setVerticalSpacing(FORM_VSPACING)
    form_layout.setContentsMargins(0, 0, 0, 0)


def _fix_label_for_field(form_layout: QFormLayout, field_widget: QWidget) -> None:
    lbl = form_layout.labelForField(field_widget)
    if lbl is None:
        return
    lbl.setFixedWidth(FORM_LABEL_WIDTH)
    lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)


def _setup_scroll(scroll: QScrollArea) -> None:
    scroll.setWidgetResizable(True)
    scroll.setFrameShape(QFrame.Shape.NoFrame)
    scroll.setContentsMargins(0, 0, 0, 0)
    scroll.setViewportMargins(0, 0, 0, 0)
    scroll.setStyleSheet("QScrollArea { border: none; }")


# ----------------------------
# 时间解析/格式化
# ----------------------------
_TIME_RE = re.compile(r"^\s*(\d+)\s*(ms|s)?\s*$", re.IGNORECASE)


def parse_time_to_ms(s: str, default_ms: Optional[int] = None) -> Optional[int]:
    if s is None:
        return default_ms
    s = str(s).strip()
    if not s:
        return default_ms
    m = _TIME_RE.match(s)
    if not m:
        return default_ms
    n = int(m.group(1))
    unit = (m.group(2) or "ms").lower()
    return n * 1000 if unit == "s" else n


def fmt_ms(ms: Optional[int]) -> str:
    return f"{int(ms)}ms" if ms is not None else ""


# ----------------------------
# DSL 数据模型
# ----------------------------
@dataclass
class SetSignalModel:
    kind: str  # "env" | "sys"
    name: str
    value: str


@dataclass
class SetStepModel:
    signals: List[SetSignalModel] = field(default_factory=list)
    wait_ms: int = 0
    next_checks: List[str] = field(default_factory=list)  # e.g. ["C1", "C2"]
    comment: str = ""  # 注释说明


@dataclass
class CheckItemModel:
    kind: str  # "sig" | "env" | "sys"
    name: str

    mode: str  # "single" | "list" | "range"
    op: str = "=="  # for single: "==", ">", "<", ">=", "<=", "="

    single_value: str = ""
    list_values: List[str] = field(default_factory=list)
    range_a: str = ""
    range_b: str = ""

    wait_ms: int = 0
    timeout_ms: int = 1000
    duration_ms: int = 0
    async_: bool = False


@dataclass
class CheckStepModel:
    items: List[CheckItemModel] = field(default_factory=list)
    comment: str = ""  # 注释说明


# ----------------------------
# SET DSL 解析/生成
# ----------------------------
_RE_SET_WAIT = re.compile(r"\bwait\s+(\d+)\s*(ms|s)\b", re.IGNORECASE)
# 修复：then check 只匹配到 comment 之前的内容
_RE_SET_THEN = re.compile(r"\bthen\s+check\s+([A-Za-z0-9_,\s]+?)(?:\s+\bcomment\b|$)", re.IGNORECASE)
# 修复：comment 匹配时包含前面的空格，以便正确删除
_RE_SET_COMMENT = re.compile(r"\s+\bcomment\s+\"([^\"]*)\"\s*$", re.IGNORECASE)


def parse_set_step(text: str) -> Tuple[SetStepModel, bool, str]:
    raw = (text or "").strip()
    if not raw:
        return SetStepModel(signals=[SetSignalModel(kind="sys", name="", value="")]), False, "空文本"

    wait_ms = 0
    next_checks: List[str] = []
    comment = ""

    # 先提取并移除 comment（必须在最后位置）
    comment_m = _RE_SET_COMMENT.search(raw)
    if comment_m:
        comment = comment_m.group(1)
        raw = _RE_SET_COMMENT.sub("", raw).strip()

    # 提取并移除 wait
    wait_m = _RE_SET_WAIT.search(raw)
    if wait_m:
        wait_ms = parse_time_to_ms(wait_m.group(1) + wait_m.group(2), default_ms=0) or 0
        raw = _RE_SET_WAIT.sub("", raw).strip()

    # 提取并移除 then check
    then_m = _RE_SET_THEN.search(raw)
    if then_m:
        part = then_m.group(1)
        ids = [x.strip() for x in part.split(",") if x.strip()]
        next_checks = [x if x.upper().startswith("C") else x for x in ids]
        raw = _RE_SET_THEN.sub("", raw).strip()

    # 剩余部分解析 signals
    parts = [p.strip() for p in raw.split(" && ") if p.strip()]
    signals: List[SetSignalModel] = []

    # 支持：set sys::A::B=1 或 sys::A::B=1
    for p in parts:
        p2 = p.strip()
        if p2.lower().startswith("set "):
            p2 = p2[4:].strip()

        if "=" not in p2:
            # 允许用户输入不完整
            name = p2.strip()
            val = ""
        else:
            name, val = p2.split("=", 1)
            name = name.strip()
            val = val.strip()

        kind = "sys"
        if name.startswith("env::"):
            kind = "env"
        elif name.startswith("sys::"):
            kind = "sys"
        signals.append(SetSignalModel(kind=kind, name=name, value=val))

    if not signals:
        return SetStepModel(signals=[SetSignalModel(kind="sys", name="", value="")]), False, "未解析到信号"

    return SetStepModel(signals=signals, wait_ms=wait_ms, next_checks=next_checks, comment=comment), True, ""


def render_set_step(model: SetStepModel) -> str:
    chunks: List[str] = []
    for i, s in enumerate(model.signals):
        name = (s.name or "").strip()
        value = (s.value or "").strip()

        # 若用户只填了 FunctionSwitch::X，也允许；但生成时尽量带前缀
        if s.kind == "sys" and name and not name.startswith("sys::"):
            name = "sys::" + name if "::" in name else name
        if s.kind == "env" and name and not name.startswith("env::"):
            name = "env::" + name if name.startswith("CAN ") or "::" in name else name

        if i == 0:
            if name and value:
                chunks.append(f"set {name}={value}")
            elif name and not value:
                chunks.append(f"set {name}=")
            else:
                chunks.append("set ")
        else:
            if name and value:
                chunks.append(f"{name}={value}")
            elif name and not value:
                chunks.append(f"{name}=")
            else:
                chunks.append("")

    out = " && ".join(chunks).strip()

    if int(model.wait_ms or 0) > 0:
        out += f" wait {int(model.wait_ms)}ms"

    if model.next_checks:
        out += " then check " + ",".join([c.strip() for c in model.next_checks if c.strip()])

    if model.comment:
        out += f' comment "{model.comment}"'

    return out.strip()

def _build_hier_index_by_kind(completions_by_kind: Dict[str, List[str]]) -> Dict[str, Dict[Tuple[str, ...], List[str]]]:
    """
    把类似：
      sys::FunctionSwitch::CSW_Enable_S
      sys::simulink::Ego_PosX
    构造成分层索引：
      index["sys"][()] -> ["FunctionSwitch", "simulink", ...]
      index["sys"][("simulink",)] -> ["Ego_PosX", "Ego_PosY", ...]
    """
    out: Dict[str, Dict[Tuple[str, ...], List[str]]] = {}
 
    for kind, paths in (completions_by_kind or {}).items():
        idx: Dict[Tuple[str, ...], set[str]] = {}
        for p in paths or []:
            if not isinstance(p, str):
                continue
            parts = [x for x in p.split("::") if x]  # 去掉空段（含末尾 ::）
            if not parts:
                continue
 
            # paths 通常已经按 kind 分组了；这里仍做一次兼容判断
            if parts[0] in ("sys", "sig", "env"):
                real_kind = parts[0]
                segs = parts[1:]
            else:
                real_kind = kind
                segs = parts
 
            if real_kind != kind:
                continue
 
            for i in range(len(segs)):
                prefix = tuple(segs[:i])
                idx.setdefault(prefix, set()).add(segs[i])
 
        out[kind] = {k: sorted(v, key=lambda s: s.lower()) for k, v in idx.items()}
 
    return out
 
 
class _HierLineEditCompleter:
    """
    QLineEdit 的逐级补全控制器：
    - model 里放"下一层 segment"（不放完整路径）
    - 使用 UnfilteredPopupCompletion，避免 QCompleter 用整行做二次过滤
    - 显式连接 activated[str]，避免选中后 QLineEdit 默认覆盖导致前缀丢失
    - 支持sys变量的struct成员补全（使用.作为分隔符）
    - 不显示前缀（sys::/env::/sig::），因为已通过下拉框选择类型
    - 支持补全完成回调（on_completed），用于通知外部更新
    """

    def __init__(
        self,
        edit: QLineEdit,
        kind_getter,
        index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        allowed_kinds: List[str],
        dbc_parser=None,
        on_completed=None,  # 新增：补全完成时的回调函数
    ) -> None:
        self._edit = edit
        self._kind_getter = kind_getter
        self._index_by_kind = index_by_kind or {}
        self._allowed_kinds = [k for k in (allowed_kinds or []) if k in self._index_by_kind]
        self._dbc_parser = dbc_parser
        self._on_completed = on_completed  # 新增

        self._model = QStringListModel(self._edit)
        self._completer = QCompleter(self._model, self._edit)
 
        # 关键点：不要让 completer 用"整行文本"二次过滤 segment 候选
        try:
            self._completer.setCompletionMode(QCompleter.CompletionMode.UnfilteredPopupCompletion)
        except AttributeError:
            # 兼容部分 PyQt6 版本
            self._completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
 
        self._completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self._edit.setCompleter(self._completer)
 
        # 状态：用于 activated 时判断是否需要补 "::" 或 "."
        self._last_base = ""
        self._last_kind = ""
        self._last_prefix_segs: Tuple[str, ...] = tuple()
        self._last_has_sep = False
        self._last_has_dot = False  # 标记是否包含.（用于struct成员）
 
        self._edit.textEdited.connect(self._on_text_edited)
 
        # 关键点：显式连接 str 重载，避免前缀被默认插入行为覆盖
        self._completer.activated[str].connect(self._on_activated)
 
        self._update_candidates(self._edit.text(), force_popup=False)
 
    def refresh(self) -> None:
        self._update_candidates(self._edit.text(), force_popup=False)
 
    def _detect_kind_and_rest(self, text: str) -> Tuple[str, str]:
        """检测类型和剩余文本（不包含前缀，因为已通过下拉框选择）"""
        t = text or ""
        # 直接使用下拉框选择的类型，不从文本中检测前缀
        default_kind = (self._kind_getter() or "").strip()
        if default_kind not in self._allowed_kinds and self._allowed_kinds:
            default_kind = self._allowed_kinds[0]
        return default_kind, t
 
    def _set_candidates(self, items: List[str], force_popup: bool) -> None:
        self._model.setStringList(items or [])
 
        # 兜底：再禁一次 prefix 过滤
        self._completer.setCompletionPrefix("")
 
        if force_popup and items and self._edit.hasFocus():
            self._completer.complete()
 
    def _on_text_edited(self, text: str) -> None:
        self._update_candidates(text, force_popup=True)
 
    def _update_candidates(self, text: str, force_popup: bool) -> None:
        t = text or ""
        has_sep = "::" in t
        has_dot = "." in t
        self._last_has_sep = has_sep
        self._last_has_dot = has_dot
 
        kind, rest = self._detect_kind_and_rest(t)
        self._last_kind = kind
        idx = self._index_by_kind.get(kind, {})
 
        # 特殊处理：sys类型的struct成员补全
        if kind == "sys" and has_dot and self._dbc_parser:
            # 分离变量路径和成员前缀
            var_part, member_prefix = t.rsplit(".", 1)
            # 构造完整的变量路径（包含sys::前缀）
            full_var_path = f"sys::{var_part}"
            
            # 从DBCParser获取struct成员补全
            if self._dbc_parser:
                suggestions = self._dbc_parser.get_signal_completion(full_var_path + ".", "sys")
                if suggestions:
                    # 提取成员名称（去掉前缀）
                    members = [s.rsplit(".", 1)[-1] for s in suggestions]
                    if member_prefix:
                        members = [m for m in members if m.lower().startswith(member_prefix.lower())]
                    self._last_base = var_part + "."
                    self._last_prefix_segs = tuple()
                    self._set_candidates(members, force_popup)
                    return
        
        # base / partial：用于"替换当前 segment"
        if has_sep:
            if t.endswith("::"):
                self._last_base = t
                partial = ""
            else:
                pos = t.rfind("::")
                self._last_base = t[: pos + 2]
                partial = t[pos + 2 :]
        else:
            self._last_base = ""
            partial = t
 
        # 1) 没有 ::：提示当前 kind 的第一层（不提示kind本身）
        if not has_sep:
            p = (partial or "").strip()
            roots = idx.get((), [])
            if p:
                roots = [x for x in roots if x.lower().startswith(p.lower())]
            self._last_prefix_segs = tuple()
            self._set_candidates(roots, force_popup)
            return

        # 2) 有 ::：按层级提示下一段 children
        parts = rest.split("::")
        ends_with_sep = t.endswith("::")

        if ends_with_sep:
            prefix_segs = tuple([x for x in parts[:-1] if x])
            seg_partial = ""
        else:
            prefix_segs = tuple([x for x in parts[:-1] if x])
            seg_partial = parts[-1] if parts else ""

        self._last_prefix_segs = prefix_segs

        children = idx.get(prefix_segs, [])
        if seg_partial:
            children = [c for c in children if c.lower().startswith(seg_partial.lower())]

        self._set_candidates(children, force_popup)

    def _on_activated(self, chosen: str) -> None:
        chosen = (chosen or "").strip()
        if not chosen:
            return

        chosen = chosen.rstrip(":")
        idx = self._index_by_kind.get(self._last_kind, {})

        # A) 特殊处理：sys类型的struct成员补全
        if self._last_kind == "sys" and self._last_has_dot:
            # 替换当前 segment：base + chosen
            new_text = (self._last_base + chosen) if self._last_base else chosen
            # 使用 QTimer 延迟执行，避免 QCompleter 默认行为覆盖
            def apply_text():
                blocker = QSignalBlocker(self._edit)
                self._edit.setText(new_text)
                self._edit.setCursorPosition(len(new_text))
                del blocker
                self._update_candidates(new_text, force_popup=False)
                # 补全完成，触发回调（struct成员补全后可能已完成）
                if self._on_completed and not new_text.endswith("."):
                    self._on_completed()
            QTimer.singleShot(0, apply_text)
            return

        # B) 替换当前 segment：base + chosen
        new_text = (self._last_base + chosen) if self._last_base else chosen

        # 如果 chosen 下面还有子节点，自动补 "::"
        has_more_children = idx.get(self._last_prefix_segs + (chosen,), [])
        if has_more_children:
            if not new_text.endswith("::"):
                new_text += "::"

        # 使用 QTimer 延迟执行，避免 QCompleter 默认行为覆盖
        def apply_text():
            blocker = QSignalBlocker(self._edit)
            self._edit.setText(new_text)
            self._edit.setCursorPosition(len(new_text))
            del blocker
            self._update_candidates(new_text, force_popup=new_text.endswith("::"))
            # 补全完成（没有更多子节点），触发回调
            if self._on_completed and not has_more_children:
                self._on_completed()
        QTimer.singleShot(0, apply_text)


class _SignalValueHelper:
    """
    信号值提示辅助类

    功能：
    1. 根据信号定义动态切换 QLineEdit/QComboBox
    2. 显示范围或枚举值含义的标签
    3. 获取/设置值
    """

    def __init__(
        self,
        kind_getter,           # 获取当前信号类型的函数
        name_getter,           # 获取当前信号名的函数
        layout: QHBoxLayout,   # 值输入控件所在的布局
        insert_index: int,     # 值控件在布局中的索引位置
        dbc_parser,            # DBCParser实例
        value_placeholder: str = "值",  # 默认placeholder
        value_width: int = 0,  # 值控件宽度（0表示不固定）
        info_width: int = 120, # 信息标签最小宽度
    ):
        self._kind_getter = kind_getter
        self._name_getter = name_getter
        self._layout = layout
        self._insert_index = insert_index
        self._dbc_parser = dbc_parser
        self._value_placeholder = value_placeholder
        self._value_width = value_width
        self._info_width = info_width

        # 当前控件
        self._value_edit: Optional[QLineEdit] = None
        self._value_combo: Optional[QComboBox] = None
        self._info_label: Optional[QLineEdit] = None  # 使用只读QLineEdit保持视觉一致性

        # 当前信号信息
        self._current_choices: Optional[Dict] = None
        self._current_min: Optional[float] = None
        self._current_max: Optional[float] = None

        # 创建初始控件
        self._ensure_line_edit()
        self._ensure_info_label()

    def _ensure_line_edit(self) -> QLineEdit:
        """确保 QLineEdit 存在"""
        if self._value_edit is None:
            self._value_edit = QLineEdit()
            self._value_edit.setPlaceholderText(self._value_placeholder)
            if self._value_width > 0:
                self._value_edit.setFixedWidth(self._value_width)
            self._layout.insertWidget(self._insert_index, self._value_edit)
        return self._value_edit

    def _ensure_combo(self) -> QComboBox:
        """确保 QComboBox 存在（不可编辑，只能选择枚举值）"""
        if self._value_combo is None:
            self._value_combo = QComboBox()
            self._value_combo.setEditable(False)  # 不可编辑，只能选择choice内的值
            if self._value_width > 0:
                self._value_combo.setFixedWidth(self._value_width)
            self._layout.insertWidget(self._insert_index, self._value_combo)
            self._value_combo.currentIndexChanged.connect(self._on_combo_index_changed)
        return self._value_combo

    def _ensure_info_label(self) -> QLineEdit:
        """确保信息标签存在（使用只读QLineEdit保持视觉一致性）"""
        if self._info_label is None:
            self._info_label = QLineEdit()
            self._info_label.setReadOnly(True)
            self._info_label.setStyleSheet("background-color: #f0f0f0; color: #666666;")
            self._info_label.setMinimumWidth(self._info_width)
            self._layout.insertWidget(self._insert_index + 1, self._info_label)
            # 默认显示 unknown
            self._info_label.setText("unknown")
        return self._info_label

    def _on_combo_index_changed(self, index: int) -> None:
        """QComboBox 选中变化时更新信息标签"""
        if self._value_combo is None or self._info_label is None or self._current_choices is None:
            return

        if index >= 0:
            # 从 choices 中获取描述
            current_data = self._value_combo.currentData()
            if current_data is not None and current_data in self._current_choices:
                self._info_label.setText(str(self._current_choices[current_data]))

    def update_from_signal_name(self) -> None:
        """
        根据当前信号名更新值控件

        调用时机：
        - name_edit.textChanged
        - kind_combo.currentTextChanged
        - 补全选中后
        """
        kind = self._kind_getter()
        name = self._name_getter()

        # 只处理 sig 和 env 类型
        if kind not in ("sig", "env"):
            self._switch_to_line_edit()
            self._info_label.setText("unknown")
            return

        # 解析信号路径（name 可能包含或不包含 sig::/env:: 前缀）
        info = None
        if self._dbc_parser and name:
            # 构建完整路径用于查询
            full_path = name
            if not name.startswith("CAN "):
                # name 可能是已包含类型前缀的完整路径
                if name.startswith("sig::") or name.startswith("env::"):
                    full_path = name
                else:
                    full_path = f"CAN {name}"  # 尝试补全格式
            info = self._dbc_parser.get_signal_info(full_path, kind)

        if info is None:
            self._switch_to_line_edit()
            self._info_label.setText("unknown")
            return

        choices = info.get("choices")
        minimum = info.get("minimum")
        maximum = info.get("maximum")

        self._current_choices = choices
        self._current_min = minimum
        self._current_max = maximum

        if choices:
            # 有枚举值：切换为 QComboBox
            self._switch_to_combo(choices)
            self._info_label.show()
            # 初始时显示第一个枚举值的含义
            if self._value_combo and self._value_combo.count() > 0:
                first_val = list(choices.keys())[0]
                self._info_label.setText(str(choices[first_val]))
        else:
            # 无枚举值：保持 QLineEdit，显示范围标签
            self._switch_to_line_edit()
            self._info_label.show()
            if minimum is not None and maximum is not None:
                self._info_label.setText(f"[{minimum} ~ {maximum}]")
            else:
                self._info_label.setText("unknown")

    def _switch_to_line_edit(self) -> None:
        """切换到 QLineEdit"""
        combo = self._value_combo
        edit = self._ensure_line_edit()

        if combo is not None and combo.isVisible():
            # 保存当前值
            current_value = combo.currentText()
            combo.hide()
            edit.show()
            edit.setText(current_value)
            # 不强制设置焦点，保持用户当前编辑的控件

    def _switch_to_combo(self, choices: Dict) -> None:
        """切换到 QComboBox 并填充枚举值（不可编辑，只能选择枚举值）"""
        edit = self._value_edit
        combo = self._ensure_combo()

        # 保存当前值（从 edit 或 combo 获取）
        current_value = ""
        if edit is not None and edit.isVisible():
            current_value = edit.text()
            edit.hide()
        if combo is not None and combo.isVisible():
            # 从 combo 获取当前选中项的实际值
            data = combo.currentData()
            if data is not None:
                current_value = str(data)

        # 清空并重新填充枚举值
        combo.clear()
        for value, description in choices.items():
            combo.addItem(f"{value}", userData=value)

        # 尝试恢复之前的值（如果值在枚举列表中）
        found = False
        if current_value:
            for i in range(combo.count()):
                if str(combo.itemData(i)) == current_value.strip():
                    combo.setCurrentIndex(i)
                    found = True
                    break

        # 如果没找到匹配项，默认选择第一个
        if not found and combo.count() > 0:
            combo.setCurrentIndex(0)

        combo.show()
        # 更新信息标签显示当前选中项的含义
        if self._info_label and choices and combo.currentIndex() >= 0:
            current_data = combo.currentData()
            if current_data is not None and current_data in choices:
                self._info_label.setText(str(choices[current_data]))

    def get_value(self) -> str:
        """获取当前值"""
        combo = self._value_combo
        edit = self._value_edit

        if combo is not None and combo.isVisible():
            # 返回实际数值（userData），而非显示文本
            data = combo.currentData()
            if data is not None:
                return str(data)
            return combo.currentText().strip()
        elif edit is not None:
            return edit.text().strip()
        return ""

    def set_value(self, value: str) -> None:
        """设置当前值"""
        combo = self._value_combo
        edit = self._value_edit

        if combo is not None and combo.isVisible():
            # 尝试在 combo 中找到匹配项
            for i in range(combo.count()):
                if str(combo.itemData(i)) == str(value).strip():
                    combo.setCurrentIndex(i)
                    return
            # 没找到匹配项（combo不可编辑），默认选择第一个
            if combo.count() > 0:
                combo.setCurrentIndex(0)
        elif edit is not None:
            edit.setText(value)

    def show_widgets(self) -> None:
        """显示值控件和标签"""
        if self._value_combo is not None and self._value_combo.count() > 0:
            self._value_combo.show()
        elif self._value_edit is not None:
            self._value_edit.show()
        if self._info_label is not None and self._current_choices or (self._current_min is not None and self._current_max is not None):
            self._info_label.show()

    def hide_widgets(self) -> None:
        """隐藏值控件和标签（用于CHECK的list/range模式）"""
        if self._value_combo is not None:
            self._value_combo.hide()
        if self._value_edit is not None:
            self._value_edit.hide()
        if self._info_label is not None:
            self._info_label.hide()

    def get_info_label(self) -> Optional[QLineEdit]:
        """获取信息标签（用于外部更新，如CHECK的list/range模式）"""
        return self._info_label

    def get_current_range(self) -> Tuple[Optional[float], Optional[float]]:
        """获取当前信号的范围（min, max）"""
        return self._current_min, self._current_max

    def get_current_choices(self) -> Optional[Dict]:
        """获取当前信号的枚举值"""
        return self._current_choices

    def get_signal_name(self) -> str:
        """获取当前信号名"""
        return self._name_getter()

    def validate_value(self) -> Optional[str]:
        """
        验证当前值是否有效

        返回：
        - None: 值有效或无法验证（无范围信息）
        - str: 错误消息，描述值不在范围内
        """
        value_str = self.get_value()
        if not value_str:
            return None  # 空值不验证

        # choice 模式：combo 不可编辑，值一定在枚举列表中，无需验证

        # 范围模式：检查值是否在 min-max 范围内
        if self._current_min is not None and self._current_max is not None:
            try:
                # 尝试解析数值（支持十进制、十六进制）
                if value_str.startswith("0x") or value_str.startswith("0X"):
                    value = float(int(value_str, 16))
                else:
                    value = float(value_str)

                if value < self._current_min or value > self._current_max:
                    signal_name = self.get_signal_name()
                    return f"信号 {signal_name} 的值 {value_str} 不在范围 [{self._current_min} ~ {self._current_max}] 内"
            except ValueError:
                # 无法解析为数值，不验证
                pass

        return None


# ----------------------------
# CHECK DSL 解析/生成
# ----------------------------
_RE_CHECK_ASYNC = re.compile(r"\basync\s+(true|false)\b", re.IGNORECASE)
_RE_CHECK_WAIT = re.compile(r"\bwait\s+(\d+)\s*(ms|s)\b", re.IGNORECASE)
_RE_CHECK_TIMEOUT = re.compile(r"\btimeout(?:OfCheck)?\s+(\d+)\s*(ms|s)\b", re.IGNORECASE)
_RE_CHECK_DURATION = re.compile(r"\b(duration|checkInTime)\s+(\d+)\s*(ms|s)\b", re.IGNORECASE)
_RE_CHECK_COMMENT = re.compile(r"\bcomment\s+\"([^\"]*)\"", re.IGNORECASE)

_RE_CHECK_IN = re.compile(r"\bin\s*\[(.*?)\]\s*$", re.IGNORECASE)
_RE_CHECK_RANGE = re.compile(r"=\s*([^\s]+)\s*\.\.\s*([^\s]+)\s*$")
# 匹配单值：操作符后跟值，注意 != 要放在 = 前面避免误匹配
_RE_CHECK_SINGLE = re.compile(r"(==|!=|>=|<=|>|<|=)\s*(\S+?)(?=\s+(?:async|timeout|duration|wait|comment\b)|\s*$)", re.IGNORECASE)


def _strip_params_from_check_expr(expr: str) -> Tuple[str, Dict[str, Any]]:
    s = expr.strip()
    info: Dict[str, Any] = {}

    m = _RE_CHECK_ASYNC.search(s)
    if m:
        info["async_"] = m.group(1).lower() == "true"
        s = _RE_CHECK_ASYNC.sub("", s).strip()
    else:
        info["async_"] = False

    m = _RE_CHECK_TIMEOUT.search(s)
    if m:
        info["timeout_ms"] = parse_time_to_ms(m.group(1) + m.group(2), default_ms=1000) or 1000
        s = _RE_CHECK_TIMEOUT.sub("", s).strip()
    else:
        info["timeout_ms"] = 1000

    m = _RE_CHECK_DURATION.search(s)
    if m:
        info["duration_ms"] = parse_time_to_ms(m.group(2) + m.group(3), default_ms=0) or 0
        s = _RE_CHECK_DURATION.sub("", s).strip()
    else:
        info["duration_ms"] = 0

    m = _RE_CHECK_WAIT.search(s)
    if m:
        info["wait_ms"] = parse_time_to_ms(m.group(1) + m.group(2), default_ms=0) or 0
        s = _RE_CHECK_WAIT.sub("", s).strip()
    else:
        info["wait_ms"] = 0

    # 移除 comment（comment 是全局属性，不属于单个检查项）
    s = _RE_CHECK_COMMENT.sub("", s).strip()

    # async=true 时 wait 不生效
    if info.get("async_"):
        info["wait_ms"] = 0

    return s.strip(), info


def parse_check_step(text: str) -> Tuple[CheckStepModel, bool, str]:
    raw = (text or "").strip()
    if not raw:
        return CheckStepModel(items=[CheckItemModel(kind="sig", name="", mode="single")]), False, "空文本"

    # 提取 comment（全局属性，不属于单个检查项）
    comment = ""
    comment_m = _RE_CHECK_COMMENT.search(raw)
    if comment_m:
        comment = comment_m.group(1)

    # 先移除 comment，再分割 parts，避免 comment 被当作表达式的一部分
    tmp = _RE_CHECK_COMMENT.sub("", raw).strip()

    # 允许：第一段有 "check "，后续段可能省略 "check "
    parts = [p.strip() for p in re.split(r"\s*&&\s*", tmp) if p.strip()]
    items: List[CheckItemModel] = []

    for p in parts:
        p2 = p.strip()
        if p2.lower().startswith("check "):
            p2 = p2[6:].strip()

        # 先提取 async/timeout/duration/wait/comment
        expr, info = _strip_params_from_check_expr(p2)

        # 判断值模式
        mode = "single"
        op = "=="
        single_value = ""
        list_values: List[str] = []
        range_a = ""
        range_b = ""

        # list: "... in [1,2,3]"
        m_in = _RE_CHECK_IN.search(expr)
        if m_in:
            mode = "list"
            inside = m_in.group(1).strip()
            list_values = [x.strip() for x in inside.split(",") if x.strip()]
            name = expr[: m_in.start()].strip()
        else:
            # range: "... = a .. b"
            m_range = _RE_CHECK_RANGE.search(expr)
            if m_range:
                mode = "range"
                range_a = m_range.group(1).strip()
                range_b = m_range.group(2).strip()
                name = expr[: m_range.start()].strip()
            else:
                m_single = _RE_CHECK_SINGLE.search(expr)
                if m_single:
                    mode = "single"
                    op = m_single.group(1)
                    single_value = m_single.group(2).strip()
                    name = expr[: m_single.start()].strip()
                else:
                    # 解析失败时：尽量把整个 expr 当 name
                    name = expr.strip()

        kind = "sig"
        if name.startswith("env::"):
            kind = "env"
        elif name.startswith("sys::"):
            kind = "sys"
        elif name.startswith("sig::"):
            kind = "sig"

        items.append(
            CheckItemModel(
                kind=kind,
                name=name,
                mode=mode,
                op=op,
                single_value=single_value,
                list_values=list_values,
                range_a=range_a,
                range_b=range_b,
                wait_ms=int(info.get("wait_ms") or 0),
                timeout_ms=int(info.get("timeout_ms") or 1000),
                duration_ms=int(info.get("duration_ms") or 0),
                async_=bool(info.get("async_") or False),
            )
        )

    if not items:
        return CheckStepModel(items=[CheckItemModel(kind="sig", name="", mode="single")]), False, "未解析到检查项"

    return CheckStepModel(items=items, comment=comment), True, ""


def render_check_step(model: CheckStepModel) -> str:
    chunks: List[str] = []
    for i, it in enumerate(model.items):
        name = (it.name or "").strip()

        # 尽量带前缀
        if it.kind in ("sig", "env", "sys"):
            if it.kind == "sig" and name and not name.startswith("sig::"):
                name = "sig::" + name if "::" in name else name
            if it.kind == "env" and name and not name.startswith("env::"):
                name = "env::" + name if "::" in name else name
            if it.kind == "sys" and name and not name.startswith("sys::"):
                name = "sys::" + name if "::" in name else name

        if i==0:
            expr = f"check {name}".rstrip()
        else:
            expr = f"{name}".rstrip()

        if it.mode == "list":
            expr += " in [" + ",".join([x.strip() for x in it.list_values if x.strip()]) + "]"
        elif it.mode == "range":
            a = (it.range_a or "").strip()
            b = (it.range_b or "").strip()
            expr += f"={a}..{b}"
        else:
            op = (it.op or "==").strip()
            v = (it.single_value or "").strip()
            expr += f"{op}{v}"

        # 参数：timeout 默认 1000ms，但生成时仍显式输出，便于可读/可控
        timeout_ms = int(it.timeout_ms or 1000)
        duration_ms = int(it.duration_ms or 0)
        wait_ms = int(it.wait_ms or 0)
        async_ = bool(it.async_)

        if not async_ and wait_ms > 0:
            expr += f" wait {wait_ms}ms"

        expr += f" timeout {timeout_ms}ms"

        if duration_ms > 0:
            expr += f" duration {duration_ms}ms"

        expr += f" async {'true' if async_ else 'false'}"

        chunks.append(expr.strip())

    out = " && ".join(chunks).strip()

    # 添加 comment（全局属性）
    if model.comment:
        out += f' comment "{model.comment}"'

    return out


# ----------------------------
# Step 编辑弹窗：SET
# ----------------------------
class SetSignalRow(QWidget):
    removed = pyqtSignal(object)

    def __init__(
        self,
        completions_by_kind: Dict[str, List[str]],
        hier_index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        parent=None,
        dbc_parser=None,
    ):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind
        self._hier_index_by_kind = hier_index_by_kind
        self._dbc_parser = dbc_parser
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        # 6列布局：信号类型 | 信号名 | 信号值 | 范围/描述 | 删除

        # 第1列：信号类型选择
        self.kind_combo = QComboBox(self)
        self.kind_combo.addItems(["env", "sys"])
        self.kind_combo.setFixedWidth(60)
        layout.addWidget(self.kind_combo)

        # 第2列：信号名称输入（stretch=2，信号名通常较长）
        self.name_edit = QLineEdit(self)
        self.name_edit.setPlaceholderText("信号名称（逐级补全：输入 sys:: 后提示 namespace）")
        layout.addWidget(self.name_edit, 2)

        # 第3、4列：信号值 + 范围/描述（使用 _SignalValueHelper）
        self._value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=layout,
            insert_index=2,  # 在 name_edit 之后插入
            dbc_parser=self._dbc_parser,
            value_placeholder="值（如 0x1 / 1 / true）",
            value_width=100,
            info_width=120,
        )

        # 第5列：删除按钮
        btn_del = QPushButton("删除", self)
        btn_del.clicked.connect(lambda: self.removed.emit(self))
        layout.addWidget(btn_del)

        # 逐级补全（只对 name_edit），补全完成时触发值控件更新
        self._hier = _HierLineEditCompleter(
            edit=self.name_edit,
            kind_getter=lambda: self.kind_combo.currentText(),
            index_by_kind=self._hier_index_by_kind,
            allowed_kinds=["env", "sys"],
            dbc_parser=self._dbc_parser,
            on_completed=self._value_helper.update_from_signal_name,  # 补全完成时更新值控件
        )
        self.kind_combo.currentTextChanged.connect(self._on_kind_changed)
        self._on_kind_changed(self.kind_combo.currentText())  # 初始化placeholderText

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._value_helper.update_from_signal_name)
        self.kind_combo.currentTextChanged.connect(
            lambda _: self._value_helper.update_from_signal_name()
        )

    def _on_kind_changed(self, kind: str) -> None:
        """根据类型更新placeholderText和刷新补全"""
        self._hier.refresh()
        if kind == "sys":
            self.name_edit.setPlaceholderText("信号名称（逐级补全：namespace::variable 或 namespace::struct::variable）")
        elif kind == "env":
            self.name_edit.setPlaceholderText("信号名称（逐级补全：CAN X::Message::Signal）")
        else:
            self.name_edit.setPlaceholderText("信号名称")

    def set_data(self, s: SetSignalModel) -> None:
        self.kind_combo.setCurrentText(s.kind if s.kind in ("env", "sys") else "sys")
        # 去掉前缀（sys::/env::），因为已通过下拉框选择类型
        name = s.name or ""
        if name.startswith("sys::"):
            name = name[5:]
        elif name.startswith("env::"):
            name = name[5:]
        self.name_edit.setText(name)
        self._value_helper.set_value(s.value or "")
        self._on_kind_changed(self.kind_combo.currentText())  # 更新placeholderText
        # 触发更新值控件
        self._value_helper.update_from_signal_name()

    def get_data(self) -> SetSignalModel:
        # 添加前缀（sys::/env::），因为输入框中没有前缀
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()
        if name and not name.startswith("sys::") and not name.startswith("env::"):
            name = f"{kind}::{name}"
        return SetSignalModel(
            kind=kind,
            name=name,
            value=self._value_helper.get_value(),
        )

    def validate(self) -> Optional[str]:
        """验证数据，返回错误消息（None表示有效）"""
        return self._value_helper.validate_value()

class SetStepDialog(QDialog):
    def __init__(
        self,
        raw_text: str,
        completions_by_kind: Dict[str, List[str]],
        available_check_ids: List[str],
        parent=None,
        dbc_parser=None,
    ):
        super().__init__(parent)
        self.setWindowTitle("编辑 SET Step")
        self.resize(900, 600)

        self._raw_text = raw_text or ""
        self._completions_by_kind = completions_by_kind
        self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)
        self._available_check_ids = available_check_ids
        self._dbc_parser = dbc_parser

        self._model = SetStepModel(signals=[SetSignalModel(kind="sys", name="", value="")], wait_ms=0, next_checks=[])
        self._parsed_ok = False
        self._parse_msg = ""

        self._build_ui()
        self._try_parse_and_fill()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        self.tip = QLabel("", self)
        self.tip.setStyleSheet("color: #b00020;")
        self.tip.setWordWrap(True)
        layout.addWidget(self.tip)

        step_box = QGroupBox("Step 参数", self)
        _setup_groupbox_style(step_box)
        step_l = QFormLayout(step_box)
        _align_form_layout(step_l)

        # wait(ms)
        self.wait_edit = QLineEdit(step_box)
        self.wait_edit.setPlaceholderText("0")
        step_l.addRow("wait(ms):", self.wait_edit)
        _fix_label_for_field(step_l, self.wait_edit)

        row_next = QWidget(step_box)
        row_next_l = QHBoxLayout(row_next)
        row_next_l.setContentsMargins(0, 0, 0, 0)
        row_next_l.setSpacing(6)

        self.next_check_enable = QCheckBox("启用 then check", row_next)
        self.next_checks_edit = QLineEdit(step_box)
        self.next_checks_edit.setPlaceholderText("输入 C1,C2...（可选）")
        self.next_checks_edit.setEnabled(False)

        self.next_check_enable.toggled.connect(self.next_checks_edit.setEnabled)

        row_next_l.addWidget(self.next_check_enable)
        row_next_l.addWidget(self.next_checks_edit, 1)

        step_l.addRow("next check:", row_next)
        _fix_label_for_field(step_l, row_next)

        if self._available_check_ids:
            hint = QLabel(f"可用 CHECK：{', '.join(self._available_check_ids)}", step_box)
            hint.setStyleSheet("color: #666666;")
            hint.setWordWrap(True)
            step_l.addRow("", hint)

        # 添加 comment 输入框
        self.comment_edit = QLineEdit(step_box)
        self.comment_edit.setPlaceholderText("输入注释说明（可选）")
        step_l.addRow("comment:", self.comment_edit)
        _fix_label_for_field(step_l, self.comment_edit)

        layout.addWidget(step_box)

        sig_box = QGroupBox("SET 信号列表", self)
        _setup_groupbox_style(sig_box)
        sig_l = QVBoxLayout(sig_box)

        sig_header = QWidget(sig_box)
        sig_header_l = QHBoxLayout(sig_header)
        sig_header_l.setContentsMargins(0, 0, 0, 0)
        sig_header_l.setSpacing(6)
        sig_hint = QLabel("可添加多条；wait/then check 对整个 step 生效", sig_box)
        sig_hint.setStyleSheet("color: #666666;")
        sig_hint.setWordWrap(True)
        btn_add = QPushButton("添加 signal", sig_box)
        btn_add.clicked.connect(self._add_row)
        sig_header_l.addWidget(sig_hint, 1)
        sig_header_l.addWidget(btn_add)
        sig_l.addWidget(sig_header)

        self.sig_scroll = QScrollArea(sig_box)
        _setup_scroll(self.sig_scroll)

        self.sig_container = QWidget(self.sig_scroll)
        self.sig_container.setContentsMargins(0, 0, 0, 0)

        self.sig_layout = QVBoxLayout(self.sig_container)
        self.sig_layout.setContentsMargins(0, 0, 0, 0)
        self.sig_layout.setSpacing(6)
        self.sig_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.sig_scroll.setWidget(self.sig_container)
        sig_l.addWidget(self.sig_scroll, 1)

        layout.addWidget(sig_box, 1)

        self.buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)
        self.buttons.accepted.connect(self._on_ok)
        self.buttons.rejected.connect(self.reject)
        layout.addWidget(self.buttons)

    def _set_tip(self, msg: str) -> None:
        self.tip.setText(msg or "")

    def _clear_rows(self) -> None:
        while self.sig_layout.count():
            item = self.sig_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

    def _add_row(self, data: Optional[SetSignalModel] = None) -> None:
        row = SetSignalRow(self._completions_by_kind, self._hier_index_by_kind, self.sig_container, self._dbc_parser)
        row.removed.connect(self._remove_row)
    
        if isinstance(data, SetSignalModel):
            row.set_data(data)
    
        self.sig_layout.addWidget(row)

    def _remove_row(self, row: SetSignalRow) -> None:
        if self.sig_layout.count() <= 1:
            row.set_data(SetSignalModel(kind="sys", name="", value=""))
            return
        self.sig_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()

    def _try_parse_and_fill(self) -> None:
        model, ok, msg = parse_set_step(self._raw_text)
        self._model = model
        self._parsed_ok = ok
        self._parse_msg = msg
 
        if not ok:
            self._set_tip("当前文本无法完全解析，将以表单内容重新生成并覆盖主界面文本。")
        else:
            self._set_tip("")
 
        self.wait_edit.setText(str(int(self._model.wait_ms or 0)))
        if self._model.next_checks:
            self.next_check_enable.setChecked(True)
            self.next_checks_edit.setEnabled(True)
            self.next_checks_edit.setText(",".join(self._model.next_checks))
        else:
            self.next_check_enable.setChecked(False)
            self.next_checks_edit.setEnabled(False)
            self.next_checks_edit.setText("")
        
        # 设置 comment
        self.comment_edit.setText(self._model.comment or "")
 
        self._clear_rows()
        if not self._model.signals:
            self._model.signals = [SetSignalModel(kind="sys", name="", value="")]
 
        # 关键修复点：这里必须传 s（不要传 ok / self._parsed_ok）
        for s in self._model.signals:
            self._add_row(s)

    def _collect_model(self) -> SetStepModel:
        signals: List[SetSignalModel] = []
        for i in range(self.sig_layout.count()):
            w = self.sig_layout.itemAt(i).widget()
            if isinstance(w, SetSignalRow):
                signals.append(w.get_data())

        wait_ms = int(self.wait_edit.text().strip() or "0")

        next_checks: List[str] = []
        if self.next_check_enable.isChecked():
            part = self.next_checks_edit.text().strip()
            if part:
                ids = [x.strip() for x in part.split(",") if x.strip()]
                next_checks = ids

        # 至少保留 1 条 signal，避免生成空
        if not signals:
            signals = [SetSignalModel(kind="sys", name="", value="")]

        comment = self.comment_edit.text().strip()
        return SetStepModel(signals=signals, wait_ms=wait_ms, next_checks=next_checks, comment=comment)

    def _on_ok(self) -> None:
        # 验证所有信号的值是否在范围内
        for i in range(self.sig_layout.count()):
            w = self.sig_layout.itemAt(i).widget()
            if isinstance(w, SetSignalRow):
                error_msg = w.validate()
                if error_msg:
                    QMessageBox.warning(self, "警告", error_msg)
                    return

        # 基础校验（尽量不阻塞）
        model = self._collect_model()
        # then check 的引用：如果提供了 available_check_ids，则做提示
        if model.next_checks and self._available_check_ids:
            bad = [c for c in model.next_checks if c not in self._available_check_ids]
            if bad:
                ret = QMessageBox.warning(
                    self,
                    "提示",
                    f"以下 CHECK 步骤在当前 CHECK 模块中不存在：{', '.join(bad)}\n仍然要保存吗？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if ret != QMessageBox.StandardButton.Yes:
                    return

        self._model = model
        self.accept()

    def to_dsl(self) -> str:
        return render_set_step(self._model)


# ----------------------------
# Step 编辑弹窗：CHECK
# ----------------------------
class CheckItemRow(QWidget):
    removed = pyqtSignal(object)

    def __init__(
        self,
        completions_by_kind: Dict[str, List[str]],
        hier_index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        parent=None,
        dbc_parser=None,
    ):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind
        self._hier_index_by_kind = hier_index_by_kind
        self._dbc_parser = dbc_parser
        self._build_ui()

    def _build_ui(self) -> None:
        # 启用样式背景以支持自定义绘制
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        # 设置边界框样式
        self.setStyleSheet("""
            CheckItemRow {
                border: 1px solid #c0c0c0;
                border-radius: 4px;
                background-color: #fafafa;
            }
        """)
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # 第一行：类型 + 名称 + 删除按钮
        row1 = QWidget(self)
        l1 = QHBoxLayout(row1)
        l1.setContentsMargins(0, 0, 0, 0)
        l1.setSpacing(6)

        kind_label = QLabel("类型:", row1)
        kind_label.setFixedWidth(45)
        kind_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        l1.addWidget(kind_label)

        self.kind_combo = QComboBox(row1)
        self.kind_combo.addItems(["sig", "env", "sys"])
        self.kind_combo.setFixedWidth(70)
        l1.addWidget(self.kind_combo)

        name_label = QLabel("名称:", row1)
        name_label.setFixedWidth(45)
        name_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        l1.addWidget(name_label)

        self.name_edit = QLineEdit(row1)
        self.name_edit.setPlaceholderText("信号名称（逐级补全：输入 sys:: / sig:: / env:: 后逐级提示）")
        l1.addWidget(self.name_edit, 1)

        btn_del = QPushButton("删除", row1)
        btn_del.setMaximumWidth(60)
        btn_del.clicked.connect(lambda: self.removed.emit(self))
        l1.addWidget(btn_del)

        root.addWidget(row1)

        # 第二行：值模式 + 比较符/值
        row2 = QWidget(self)
        self._row2_layout = QHBoxLayout(row2)
        self._row2_layout.setContentsMargins(0, 0, 0, 0)
        self._row2_layout.setSpacing(6)

        mode_label = QLabel("模式:", row2)
        mode_label.setFixedWidth(45)
        mode_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self._row2_layout.addWidget(mode_label)

        self.mode_combo = QComboBox(row2)
        self.mode_combo.addItems(["single", "list", "range"])
        self.mode_combo.setFixedWidth(70)
        self._row2_layout.addWidget(self.mode_combo)

        self.op_label = QLabel("比较符:", row2)
        self.op_label.setFixedWidth(45)
        self.op_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self._row2_layout.addWidget(self.op_label)

        self.op_combo = QComboBox(row2)
        self.op_combo.addItems(["==", ">", "<", ">=", "<=", "!="])
        self.op_combo.setFixedWidth(70)
        self._row2_layout.addWidget(self.op_combo)

        # 使用 _SignalValueHelper 替代 single_edit
        # insert_index 是在 op_combo 之后的位置
        self._single_value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=self._row2_layout,
            insert_index=self._row2_layout.indexOf(self.op_combo) + 1,
            dbc_parser=self._dbc_parser,
            value_placeholder="单值，例如 3 / 0x1 / true",
            info_width=100,
        )

        self.list_edit = QLineEdit(row2)
        self.list_edit.setPlaceholderText("列表：用逗号分隔，例如 1,2,3")
        self._row2_layout.addWidget(self.list_edit, 1)
        self.list_edit.hide()

        self.range_a_edit = QLineEdit(row2)
        self.range_a_edit.setPlaceholderText("a")
        self._row2_layout.addWidget(self.range_a_edit, 1)
        self.range_b_edit = QLineEdit(row2)
        self.range_b_edit.setPlaceholderText("b")
        self._row2_layout.addWidget(self.range_b_edit, 1)
        self.range_sep = QLabel("..", row2)
        self._row2_layout.insertWidget(self._row2_layout.indexOf(self.range_b_edit), self.range_sep)

        # 为 list/range 模式添加范围信息标签
        self._list_range_info_label = QLabel()
        self._list_range_info_label.setStyleSheet("color: #666666; font-size: 11px;")
        self._list_range_info_label.setMinimumWidth(100)
        self._row2_layout.addWidget(self._list_range_info_label)
        self._list_range_info_label.hide()

        self.range_a_edit.hide()
        self.range_sep.hide()
        self.range_b_edit.hide()

        root.addWidget(row2)

        # 第三行：参数（与第一、二行左对齐）
        row3 = QWidget(self)
        l3 = QHBoxLayout(row3)
        l3.setContentsMargins(0, 0, 0, 0)
        l3.setSpacing(6)

        # wait(ms) - 第一列，与"类型"、"模式"对齐
        wait_label = QLabel("wait(ms):", row3)
        wait_label.setFixedWidth(70)
        wait_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        l3.addWidget(wait_label)
        self.wait_edit = QLineEdit(row3)
        self.wait_edit.setPlaceholderText("0")
        self.wait_edit.setFixedWidth(70)
        l3.addWidget(self.wait_edit)

        # timeout(ms)
        timeout_label = QLabel("timeout(ms):", row3)
        timeout_label.setFixedWidth(80)
        timeout_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        l3.addWidget(timeout_label)
        self.timeout_edit = QLineEdit(row3)
        self.timeout_edit.setPlaceholderText("1000")
        self.timeout_edit.setFixedWidth(70)
        l3.addWidget(self.timeout_edit)

        # duration(ms)
        duration_label = QLabel("duration(ms):", row3)
        duration_label.setFixedWidth(80)
        duration_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        l3.addWidget(duration_label)
        self.duration_edit = QLineEdit(row3)
        self.duration_edit.setPlaceholderText("0")
        self.duration_edit.setFixedWidth(70)
        l3.addWidget(self.duration_edit)

        # async
        async_label = QLabel("async:", row3)
        async_label.setFixedWidth(55)
        async_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        l3.addWidget(async_label)
        self.async_combo = QComboBox(row3)
        self.async_combo.addItems(["false", "true"])
        self.async_combo.setFixedWidth(70)
        l3.addWidget(self.async_combo)

        l3.addStretch(1)

        root.addWidget(row3)

        # 逐级补全（只对 name_edit），补全完成时触发值控件更新
        self._hier = _HierLineEditCompleter(
            edit=self.name_edit,
            kind_getter=lambda: self.kind_combo.currentText(),
            index_by_kind=self._hier_index_by_kind,
            allowed_kinds=["sig", "env", "sys"],
            dbc_parser=self._dbc_parser,
            on_completed=self._update_value_controls,  # 补全完成时更新值控件
        )
        self.kind_combo.currentTextChanged.connect(self._on_kind_changed)
        self._on_kind_changed(self.kind_combo.currentText())  # 初始化placeholderText

        # mode/async 联动
        self.mode_combo.currentTextChanged.connect(self._on_mode_changed)
        self.async_combo.currentTextChanged.connect(self._on_async_changed)
        self._on_mode_changed(self.mode_combo.currentText())
        self._on_async_changed(self.async_combo.currentText())

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._update_value_controls)
        self.kind_combo.currentTextChanged.connect(lambda _: self._update_value_controls())

    def _update_value_controls(self) -> None:
        """根据信号名和模式更新值控件"""
        mode = self.mode_combo.currentText()

        # 先更新 single 模式的 helper
        self._single_value_helper.update_from_signal_name()

        # 对于 list/range 模式，更新范围信息标签
        if mode in ("list", "range"):
            self._update_list_range_info()

    def _update_list_range_info(self) -> None:
        """为 list/range 模式更新范围信息标签"""
        kind = self.kind_combo.currentText()
        name = self.name_edit.text()

        if kind not in ("sig", "env") or not self._dbc_parser or not name:
            self._list_range_info_label.setText("")
            self._list_range_info_label.hide()
            return

        # 构建完整路径
        full_path = name
        if not name.startswith("CAN "):
            if name.startswith("sig::") or name.startswith("env::"):
                full_path = name
            else:
                full_path = f"CAN {name}"

        info = self._dbc_parser.get_signal_info(full_path, kind)
        if info is None:
            self._list_range_info_label.setText("")
            self._list_range_info_label.hide()
            return

        choices = info.get("choices")
        minimum = info.get("minimum")
        maximum = info.get("maximum")

        if choices:
            # 有枚举值，显示可选值列表（截取前几个）
            vals = list(choices.keys())[:5]
            self._list_range_info_label.setText(f"可选: {', '.join(str(v) for v in vals)}...")
            self._list_range_info_label.show()
        elif minimum is not None and maximum is not None:
            self._list_range_info_label.setText(f"[{minimum} ~ {maximum}]")
            self._list_range_info_label.show()
        else:
            self._list_range_info_label.hide()

    def _on_kind_changed(self, kind: str) -> None:
        """根据类型更新placeholderText和刷新补全"""
        self._hier.refresh()
        if kind == "sys":
            self.name_edit.setPlaceholderText("信号名称（逐级补全：namespace::variable 或 namespace::struct::variable）")
        elif kind == "env":
            self.name_edit.setPlaceholderText("信号名称（逐级补全：CAN X::Message::Signal）")
        elif kind == "sig":
            self.name_edit.setPlaceholderText("信号名称（逐级补全：CAN X::Message::Signal）")
        else:
            self.name_edit.setPlaceholderText("信号名称")

    def _on_mode_changed(self, mode: str) -> None:
        mode = (mode or "single").lower()
        if mode == "single":
            # single模式：显示比较符和单值输入框，隐藏列表和范围相关控件
            self.op_combo.show()
            self.op_label.show()
            self._single_value_helper.show_widgets()
            self.list_edit.hide()
            self.range_a_edit.hide()
            self.range_b_edit.hide()
            self.range_sep.hide()
            self._list_range_info_label.hide()
        elif mode == "list":
            # list模式：只显示列表输入框，隐藏比较符、单值和范围相关控件
            self.op_combo.hide()
            self.op_label.hide()
            self._single_value_helper.hide_widgets()
            self.list_edit.show()
            self.range_a_edit.hide()
            self.range_b_edit.hide()
            self.range_sep.hide()
            self._update_list_range_info()
        else:  # range
            # range模式：只显示范围输入框和分隔符，隐藏比较符、单值和列表相关控件
            self.op_combo.hide()
            self.op_label.hide()
            self._single_value_helper.hide_widgets()
            self.list_edit.hide()
            self.range_a_edit.show()
            self.range_b_edit.show()
            self.range_sep.show()
            self._update_list_range_info()

    def _on_async_changed(self, v: str) -> None:
        async_true = (v or "").strip().lower() == "true"
        if async_true:
            self.wait_edit.setText("0")
        self.wait_edit.setEnabled(not async_true)

    def set_data(self, it: CheckItemModel) -> None:
        self.kind_combo.setCurrentText(it.kind if it.kind in ("sig", "env", "sys") else "sig")
        # 去掉前缀（sig::/env::/sys::），因为已通过下拉框选择类型
        name = it.name or ""
        if name.startswith("sig::"):
            name = name[5:]
        elif name.startswith("env::"):
            name = name[5:]
        elif name.startswith("sys::"):
            name = name[5:]
        self.name_edit.setText(name)

        self.mode_combo.setCurrentText(it.mode if it.mode in ("single", "list", "range") else "single")
        self.op_combo.setCurrentText(it.op if it.op in ("==", ">", "<", ">=", "<=", "!=") else "==")

        self._single_value_helper.set_value(it.single_value or "")
        self.list_edit.setText(",".join(it.list_values or []))
        self.range_a_edit.setText(it.range_a or "")
        self.range_b_edit.setText(it.range_b or "")

        self.wait_edit.setText(str(int(it.wait_ms or 0)))
        self.timeout_edit.setText(str(int(it.timeout_ms or 1000)))
        self.duration_edit.setText(str(int(it.duration_ms or 0)))
        self.async_combo.setCurrentText("true" if it.async_ else "false")
        self._on_async_changed(self.async_combo.currentText())
        self._on_mode_changed(self.mode_combo.currentText())
        self._hier.refresh()
        # 触发更新值控件
        self._update_value_controls()

    def get_data(self) -> CheckItemModel:
        kind = self.kind_combo.currentText()
        # 添加前缀（sig::/env::/sys::），因为输入框中没有前缀
        name = self.name_edit.text().strip()
        if name and not name.startswith("sig::") and not name.startswith("env::") and not name.startswith("sys::"):
            name = f"{kind}::{name}"
        mode = self.mode_combo.currentText()

        op = self.op_combo.currentText()
        single = self._single_value_helper.get_value()
        lst = [x.strip() for x in self.list_edit.text().split(",") if x.strip()]
        ra = self.range_a_edit.text().strip()
        rb = self.range_b_edit.text().strip()

        async_ = self.async_combo.currentText().strip().lower() == "true"
        wait_ms = int(self.wait_edit.text().strip() or "0") if not async_ else 0

        return CheckItemModel(
            kind=kind,
            name=name,
            mode=mode,
            op=op,
            single_value=single,
            list_values=lst,
            range_a=ra,
            range_b=rb,
            wait_ms=wait_ms,
            timeout_ms=int(self.timeout_edit.text().strip() or "1000"),
            duration_ms=int(self.duration_edit.text().strip() or "0"),
            async_=async_,
        )

    def validate(self) -> Optional[str]:
        """验证数据，返回错误消息（None表示有效）"""
        mode = self.mode_combo.currentText()
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()

        # 获取信号范围信息
        min_val, max_val = self._single_value_helper.get_current_range()

        # single 模式：验证单个值是否在范围内
        if mode == "single":
            return self._single_value_helper.validate_value()

        # list 模式：验证每个列表值是否在范围内
        if mode == "list":
            list_text = self.list_edit.text().strip()
            if list_text and min_val is not None and max_val is not None:
                values = [x.strip() for x in list_text.split(",") if x.strip()]
                for v in values:
                    try:
                        if v.startswith("0x") or v.startswith("0X"):
                            num_val = float(int(v, 16))
                        else:
                            num_val = float(v)
                        if num_val < min_val or num_val > max_val:
                            return f"信号 {name} 的列表值 {v} 不在范围 [{min_val} ~ {max_val}] 内"
                    except ValueError:
                        pass  # 无法解析为数值，不验证

        # range 模式：验证 a 和 b 是否在范围内，且 a <= b
        if mode == "range":
            a_text = self.range_a_edit.text().strip()
            b_text = self.range_b_edit.text().strip()

            if a_text and b_text:
                try:
                    if a_text.startswith("0x") or a_text.startswith("0X"):
                        a_val = float(int(a_text, 16))
                    else:
                        a_val = float(a_text)

                    if b_text.startswith("0x") or b_text.startswith("0X"):
                        b_val = float(int(b_text, 16))
                    else:
                        b_val = float(b_text)

                    # 验证 a <= b
                    if a_val > b_val:
                        return f"信号 {name} 的范围值 {a_text}..{b_text} 不合法，起始值应小于等于结束值"

                    # 验证是否在信号范围内
                    if min_val is not None and max_val is not None:
                        if a_val < min_val or a_val > max_val:
                            return f"信号 {name} 的范围起始值 {a_text} 不在范围 [{min_val} ~ {max_val}] 内"
                        if b_val < min_val or b_val > max_val:
                            return f"信号 {name} 的范围结束值 {b_text} 不在范围 [{min_val} ~ {max_val}] 内"
                except ValueError:
                    pass  # 无法解析为数值，不验证

        return None

class CheckStepDialog(QDialog):
    def __init__(self, raw_text: str, completions_by_kind: Dict[str, List[str]], parent=None, dbc_parser=None):
        super().__init__(parent)
        self.setWindowTitle("编辑 CHECK Step")
        self.resize(980, 650)

        self._raw_text = raw_text or ""
        self._completions_by_kind = completions_by_kind
        self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)
        self._dbc_parser = dbc_parser

        self._model = CheckStepModel(items=[CheckItemModel(kind="sig", name="", mode="single")])
        self._parsed_ok = False
        self._parse_msg = ""

        self._build_ui()
        self._try_parse_and_fill()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        self.tip = QLabel("", self)
        self.tip.setStyleSheet("color: #b00020;")
        self.tip.setWordWrap(True)
        layout.addWidget(self.tip)

        items_box = QGroupBox("CHECK 信号列表", self)
        _setup_groupbox_style(items_box)
        items_l = QVBoxLayout(items_box)

        header = QWidget(items_box)
        header_l = QHBoxLayout(header)
        header_l.setContentsMargins(0, 0, 0, 0)
        header_l.setSpacing(6)
        info = QLabel("每条独立 timeout/duration/async/wait；async=true 时 wait 置灰", items_box)
        info.setStyleSheet("color: #666666;")
        info.setWordWrap(True)
        btn_add = QPushButton("添加检查项", items_box)
        btn_add.clicked.connect(self._add_row)
        header_l.addWidget(info, 1)
        header_l.addWidget(btn_add)
        items_l.addWidget(header)

        self.items_scroll = QScrollArea(items_box)
        _setup_scroll(self.items_scroll)

        self.items_container = QWidget(self.items_scroll)
        self.items_container.setContentsMargins(0, 0, 0, 0)

        self.items_layout = QVBoxLayout(self.items_container)
        self.items_layout.setContentsMargins(0, 0, 0, 0)
        self.items_layout.setSpacing(10)
        self.items_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.items_scroll.setWidget(self.items_container)
        items_l.addWidget(self.items_scroll, 1)

        layout.addWidget(items_box, 1)

        # 添加 comment 输入框
        comment_box = QGroupBox("Step 参数", self)
        _setup_groupbox_style(comment_box)
        comment_l = QFormLayout(comment_box)
        _align_form_layout(comment_l)

        self.comment_edit = QLineEdit(comment_box)
        self.comment_edit.setPlaceholderText("输入注释说明（可选）")
        comment_l.addRow("comment:", self.comment_edit)
        _fix_label_for_field(comment_l, self.comment_edit)

        layout.addWidget(comment_box)

        self.buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, self)
        self.buttons.accepted.connect(self._on_ok)
        self.buttons.rejected.connect(self.reject)
        layout.addWidget(self.buttons)

    def _set_tip(self, msg: str) -> None:
        self.tip.setText(msg or "")

    def _clear_rows(self) -> None:
        while self.items_layout.count():
            item = self.items_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

    def _add_row(self, data: Optional[CheckItemModel] = None) -> None:
        row = CheckItemRow(self._completions_by_kind, self._hier_index_by_kind, self.items_container, self._dbc_parser)
        row.removed.connect(self._remove_row)
    
        if isinstance(data, CheckItemModel):
            row.set_data(data)
    
        self.items_layout.addWidget(row)

    def _remove_row(self, row: CheckItemRow) -> None:
        if self.items_layout.count() <= 1:
            row.set_data(CheckItemModel(kind="sig", name="", mode="single"))
            return
        self.items_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()

    def _try_parse_and_fill(self) -> None:
        model, ok, msg = parse_check_step(self._raw_text)
        self._model = model
        self._parsed_ok = ok
        self._parse_msg = msg
 
        if not ok:
            self._set_tip("当前文本无法完全解析，将以表单内容重新生成并覆盖主界面文本。")
        else:
            self._set_tip("")

        self._clear_rows()
        if not self._model.items:
            self._model.items = [CheckItemModel(kind="sig", name="", mode="single")]

        # 关键修复点：这里必须传 it（不要传 ok / self._parsed_ok）
        for it in self._model.items:
            self._add_row(it)
        
        # 设置 comment
        self.comment_edit.setText(self._model.comment or "")

    def _collect_model(self) -> CheckStepModel:
        items: List[CheckItemModel] = []
        for i in range(self.items_layout.count()):
            w = self.items_layout.itemAt(i).widget()
            if isinstance(w, CheckItemRow):
                items.append(w.get_data())

        if not items:
            items = [CheckItemModel(kind="sig", name="", mode="single")]

        # timeout 默认 1000ms：若用户填 0，则自动补 1000
        for it in items:
            if int(it.timeout_ms or 0) <= 0:
                it.timeout_ms = 1000
            if it.async_:
                it.wait_ms = 0

        comment = self.comment_edit.text().strip()
        return CheckStepModel(items=items, comment=comment)

    def _on_ok(self) -> None:
        # 验证所有 CHECK 项的值是否在范围内
        for i in range(self.items_layout.count()):
            w = self.items_layout.itemAt(i).widget()
            if isinstance(w, CheckItemRow):
                error_msg = w.validate()
                if error_msg:
                    QMessageBox.warning(self, "警告", error_msg)
                    return

        self._model = self._collect_model()
        self.accept()

    def to_dsl(self) -> str:
        return render_check_step(self._model)


# ----------------------------
# 主界面模块：CASE / META
# ----------------------------
class CaseInfoWidget(QWidget):
    """CASE信息模块"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        group = QGroupBox("CASE信息", self)
        _setup_groupbox_style(group)

        form_layout = QFormLayout(group)
        _align_form_layout(form_layout)

        self.case_name_edit = QLineEdit(group)
        self.case_name_edit.setPlaceholderText("输入CASE名称，如：OrinN_MR25_CSW_027_001")
        self.case_name_edit.setMinimumWidth(400)
        form_layout.addRow("CASE名称*:", self.case_name_edit)
        _fix_label_for_field(form_layout, self.case_name_edit)

        self.case_id_edit = QLineEdit(group)
        self.case_id_edit.setPlaceholderText("输入CASE ID")
        self.case_id_edit.setMinimumWidth(400)
        form_layout.addRow("CASE ID:", self.case_id_edit)
        _fix_label_for_field(form_layout, self.case_id_edit)

        layout.addWidget(group)
        layout.addStretch()

    def get_case_info(self) -> Dict[str, str]:
        return {"name": self.case_name_edit.text().strip(), "id": self.case_id_edit.text().strip()}

    def set_case_info(self, name: str, case_id: str = "") -> None:
        self.case_name_edit.setText(name)
        self.case_id_edit.setText(case_id)

    def clear(self) -> None:
        self.case_name_edit.clear()
        self.case_id_edit.clear()


class MetaInfoWidget(QWidget):
    """META信息模块"""

    # 状态更新信号，用于通知主窗口更新状态栏
    status_updated = pyqtSignal(str)
    # 预设选择变化信号
    preset_selection_changed = pyqtSignal()

    def __init__(self, parent=None, project_manager=None):
        super().__init__(parent)
        self.project_manager = project_manager
        self.scene_mapping_data = {}  # 存储场景映射表数据 {mapping_name: {scene_name: {"id": scene_id, "desc": scene_desc}}}
        self._selected_preset_signals = []  # 选中的预设信号ID列表，如 ["P1", "P2"]
        self._selected_preset_scene = ""  # 选中的预设场景ID
        self._selected_preset_scene_runtime = ""  # 选中的预设场景运行时间
        self.setup_ui()
        
    @staticmethod
    def _normalize_combo_text(value: Any) -> str:
        """
        兼容历史脏值：
        - 去尾逗号: xxx,
        - 去包裹引号: "xxx" / 'xxx'
        """
        s = str(value if value is not None else "").strip()
        s = s.rstrip(",").strip()
        if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
            s = s[1:-1].strip()
        return s

    def setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        group = QGroupBox("META信息", self)
        _setup_groupbox_style(group)

        form_layout = QFormLayout(group)
        _align_form_layout(form_layout)

        self.test_point_edit = QLineEdit(group)
        self.test_point_edit.setPlaceholderText("输入测试点描述")
        form_layout.addRow("测试点*:", self.test_point_edit)
        _fix_label_for_field(form_layout, self.test_point_edit)

        self.priority_combo = QComboBox(group)
        self.priority_combo.addItems(["P0-高", "P1-中", "P2-低"])
        form_layout.addRow("优先级:", self.priority_combo)
        _fix_label_for_field(form_layout, self.priority_combo)

        self.owner_edit = QLineEdit(group)
        self.owner_edit.setPlaceholderText("输入负责人")
        form_layout.addRow("负责人:", self.owner_edit)
        _fix_label_for_field(form_layout, self.owner_edit)

        # 场景映射表选择
        self.scene_mapping_combo = QComboBox(group)
        self.scene_mapping_combo.addItem("请选择场景映射表")
        self.scene_mapping_combo.currentTextChanged.connect(self._on_scene_mapping_changed)
        form_layout.addRow("场景映射表:", self.scene_mapping_combo)
        _fix_label_for_field(form_layout, self.scene_mapping_combo)

        # 场景名称选择
        scene_name_row = QWidget(group)
        scene_name_layout = QHBoxLayout(scene_name_row)
        scene_name_layout.setContentsMargins(0, 0, 0, 0)
        scene_name_layout.setSpacing(6)

        self.scene_name_combo = QComboBox(group)
        self.scene_name_combo.addItem("请先选择场景映射表")
        self.scene_name_combo.setEnabled(False)
        self.scene_name_combo.currentTextChanged.connect(self._on_scene_name_changed)
        scene_name_layout.addWidget(self.scene_name_combo, 1)

        self.view_scene_desc_btn = QPushButton("查看描述", group)
        self.view_scene_desc_btn.setEnabled(False)
        self.view_scene_desc_btn.clicked.connect(self._on_view_scene_description)
        scene_name_layout.addWidget(self.view_scene_desc_btn)

        self.edit_scene_btn = QPushButton("编辑场景", group)
        self.edit_scene_btn.setEnabled(False)
        self.edit_scene_btn.clicked.connect(self._on_edit_scene)
        scene_name_layout.addWidget(self.edit_scene_btn)

        form_layout.addRow("场景名称:", scene_name_row)
        _fix_label_for_field(form_layout, self.scene_name_combo)

        self.scenario_id_edit = QLineEdit(group)
        self.scenario_id_edit.setPlaceholderText("场景ID（根据场景名称自动填充）")
        self.scenario_id_edit.setReadOnly(True)
        form_layout.addRow("场景ID:", self.scenario_id_edit)
        _fix_label_for_field(form_layout, self.scenario_id_edit)

        self.ai_analysis_checkbox = QCheckBox("启用AI分析", group)
        form_layout.addRow("", self.ai_analysis_checkbox)
        _fix_label_for_field(form_layout, self.ai_analysis_checkbox)

        # 启用预设行：复选框 + 选择预设按钮
        preset_row = QWidget(group)
        preset_row_layout = QHBoxLayout(preset_row)
        preset_row_layout.setContentsMargins(0, 0, 0, 0)
        preset_row_layout.setSpacing(6)

        self.use_preset_checkbox = QCheckBox("启用预设", preset_row)
        preset_row_layout.addWidget(self.use_preset_checkbox)

        self.select_preset_btn = QPushButton("选择预设", preset_row)
        self.select_preset_btn.setEnabled(False)
        self.select_preset_btn.clicked.connect(self._on_select_preset)
        preset_row_layout.addWidget(self.select_preset_btn)
        preset_row_layout.addStretch()

        form_layout.addRow("", preset_row)
        _fix_label_for_field(form_layout, preset_row)

        # 连接启用预设复选框信号
        self.use_preset_checkbox.toggled.connect(self._on_use_preset_changed)

        self.record_checkbox = QCheckBox("启用记录", group)
        form_layout.addRow("", self.record_checkbox)
        _fix_label_for_field(form_layout, self.record_checkbox)

        layout.addWidget(group)
        layout.addStretch()
        
        # 加载场景映射表数据
        self._load_scene_mappings()

    def _load_scene_mappings(self) -> None:
        """加载场景映射表数据"""
        if not self.project_manager:
            return
        
        try:
            # 获取所有场景映射表
            scene_mappings = self.project_manager.get_scene_mappings()
            
            # 清空当前数据
            self.scene_mapping_data = {}
            
            # 加载每个场景映射表的数据
            for mapping_info in scene_mappings:
                mapping_name = mapping_info.get("name", "")
                if not mapping_name:
                    continue
                
                # 获取映射表文件路径
                file_path = self.project_manager.load_scene_mapping(mapping_name)
                if not file_path or not file_path.exists():
                    continue
                
                # 读取Excel文件
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    # 查找列索引
                    header_row = None
                    for row in ws.iter_rows(min_row=1, max_row=10):
                        row_values = [cell.value for cell in row]
                        if "场景描述" in row_values and "场景文件名" in row_values and "场景编号" in row_values:
                            header_row = row
                            break

                    if not header_row:
                        continue

                    # 获取列索引
                    desc_col = None
                    name_col = None
                    id_col = None
                    dir_col = None  # 场景目录列
                    for idx, cell in enumerate(header_row):
                        if cell.value == "场景描述":
                            desc_col = idx
                        elif cell.value == "场景文件名":
                            name_col = idx
                        elif cell.value == "场景编号":
                            id_col = idx
                        elif cell.value == "场景目录":
                            dir_col = idx

                    if desc_col is None or name_col is None or id_col is None:
                        continue

                    # 读取数据行
                    scene_data = {}
                    common_scene_dir = ""  # 公共场景目录（第一行的场景目录）
                    for row in ws.iter_rows(min_row=header_row[0].row + 1):
                        scene_name = row[name_col].value
                        scene_id = row[id_col].value
                        scene_desc = row[desc_col].value if desc_col is not None else ""
                        scene_dir = row[dir_col].value if dir_col is not None else ""

                        # 如果第一行有场景目录，保存为公共场景目录
                        if not common_scene_dir and scene_dir:
                            common_scene_dir = str(scene_dir)

                        if scene_name and scene_id is not None:
                            # 确保场景ID是整数
                            try:
                                scene_id = int(scene_id)
                                # 如果当前行场景目录为空，使用公共场景目录
                                final_dir = str(scene_dir) if scene_dir else common_scene_dir
                                scene_data[str(scene_name)] = {
                                    "id": scene_id,
                                    "desc": str(scene_desc) if scene_desc else "",
                                    "dir": final_dir
                                }
                            except (ValueError, TypeError):
                                continue

                    self.scene_mapping_data[mapping_name] = scene_data
                    
                except Exception as e:
                    print(f"加载场景映射表 '{mapping_name}' 失败: {e}")
                    continue
            
            # 更新场景映射表下拉框
            self.scene_mapping_combo.clear()
            self.scene_mapping_combo.addItem("请选择场景映射表")
            for mapping_name in self.scene_mapping_data.keys():
                self.scene_mapping_combo.addItem(mapping_name)
            
        except Exception as e:
            print(f"加载场景映射表数据失败: {e}")
    
    def _on_scene_mapping_changed(self, mapping_name: str) -> None:
        """场景映射表改变时的处理"""
        # 清空场景名称下拉框
        self.scene_name_combo.clear()
        self.scene_name_combo.addItem("请先选择场景映射表")
        self.scene_name_combo.setEnabled(False)
        
        # 清空场景ID
        self.scenario_id_edit.clear()
        
        if not mapping_name or mapping_name == "请选择场景映射表":
            return
        
        # 获取该映射表的数据
        scene_data = self.scene_mapping_data.get(mapping_name, {})
        if not scene_data:
            return
        
        # 更新场景名称下拉框
        self.scene_name_combo.clear()
        self.scene_name_combo.addItem("请选择场景名称")
        for scene_name in sorted(scene_data.keys()):
            self.scene_name_combo.addItem(scene_name)
        
        self.scene_name_combo.setEnabled(True)
    
    def _on_scene_name_changed(self, scene_name: str) -> None:
        """场景名称改变时的处理"""
        # 清空场景ID
        self.scenario_id_edit.clear()

        # 禁用查看描述按钮和编辑场景按钮
        self.view_scene_desc_btn.setEnabled(False)
        self.edit_scene_btn.setEnabled(False)

        if not scene_name or scene_name == "请选择场景名称":
            return

        # 获取当前选择的场景映射表
        mapping_name = self.scene_mapping_combo.currentText()
        if not mapping_name or mapping_name == "请选择场景映射表":
            return

        # 获取场景数据
        scene_data = self.scene_mapping_data.get(mapping_name, {})
        scene_info = scene_data.get(scene_name)

        if scene_info:
            self.scenario_id_edit.setText(str(scene_info["id"]))
            # 启用查看描述按钮
            self.view_scene_desc_btn.setEnabled(True)
            # 如果有场景目录，启用编辑场景按钮
            if scene_info.get("dir"):
                self.edit_scene_btn.setEnabled(True)
    
    def _on_view_scene_description(self) -> None:
        """查看场景描述"""
        # 获取当前选择的场景名称
        scene_name = self.scene_name_combo.currentText()
        if not scene_name or scene_name == "请选择场景名称":
            return

        # 获取当前选择的场景映射表
        mapping_name = self.scene_mapping_combo.currentText()
        if not mapping_name or mapping_name == "请选择场景映射表":
            return

        # 获取场景描述
        scene_data = self.scene_mapping_data.get(mapping_name, {})
        scene_info = scene_data.get(scene_name)

        if scene_info:
            scene_desc = scene_info.get("desc", "")
            # 显示场景描述对话框
            QMessageBox.information(
                self,
                f"场景描述 - {scene_name}",
                f"场景名称: {scene_name}\n\n场景描述:\n{scene_desc}"
            )

    def _on_edit_scene(self) -> None:
        """编辑场景 - 打开MATLAB Driving Scenario Designer并加载场景文件"""
        import os
        from .matlab_engine import get_dsd_matlab_engine

        # 获取当前选择的场景名称
        scene_name = self.scene_name_combo.currentText()
        if not scene_name or scene_name == "请选择场景名称":
            return

        # 获取当前选择的场景映射表
        mapping_name = self.scene_mapping_combo.currentText()
        if not mapping_name or mapping_name == "请选择场景映射表":
            return

        # 获取场景数据
        scene_data = self.scene_mapping_data.get(mapping_name, {})
        scene_info = scene_data.get(scene_name)

        if not scene_info:
            return

        scene_dir = scene_info.get("dir", "")
        if not scene_dir:
            QMessageBox.warning(self, "警告", "该场景没有配置场景目录")
            return

        # 构建完整的mat文件路径
        mat_file_name = scene_name if scene_name.endswith(".mat") else f"{scene_name}.mat"
        mat_file_path = os.path.join(scene_dir, mat_file_name)

        # 检查文件是否存在
        file_exists = os.path.exists(mat_file_path)
        if not file_exists:
            # 文件不存在，询问用户是否要创建新场景
            reply = QMessageBox.question(
                self,
                "场景不存在",
                f"场景文件不存在:\n{mat_file_path}\n\n"
                f"是否要创建新场景？\n\n"
                f"请在 Driving Scenario Designer 中创建场景后，\n"
                f"保存到以下位置：\n"
                f"目录: {scene_dir}\n"
                f"文件名: {mat_file_name}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # 检查是否已有 MATLAB 引擎在运行
        matlab_engine = get_dsd_matlab_engine()
        if matlab_engine.is_running():
            QMessageBox.information(self, "提示", "MATLAB Driving Scenario Designer 已在运行中")
            return

        # 启动 MATLAB Driving Scenario Designer
        def on_status_update(message: str):
            """状态更新回调，通过信号通知主窗口"""
            self.status_updated.emit(message)

        def on_dsd_closed():
            """DSD 关闭后的回调，检查场景文件是否已创建"""
            if not file_exists:
                # 原本文件不存在，检查是否已创建
                if os.path.exists(mat_file_path):
                    self.status_updated.emit(f"场景文件已创建: {mat_file_path}")
                else:
                    self.status_updated.emit(f"警告: 场景文件未创建，请保存到 {mat_file_path}")

        # 在后台线程中启动
        import threading
        # 如果文件存在则加载，否则创建新场景
        scenario_file = mat_file_path if file_exists else None
        thread = threading.Thread(
            target=matlab_engine.start_driving_scenario_designer,
            args=(scenario_file, on_status_update, on_dsd_closed),
            daemon=True
        )
        thread.start()
    
    def get_meta_info(self) -> Dict[str, Any]:
        return {
            "test_point": self.test_point_edit.text().strip(),
            "priority": self.priority_combo.currentText(),
            "owner": self.owner_edit.text().strip(),
            "scenario_id": self.scenario_id_edit.text().strip(),
            "scenario_name": self.scene_name_combo.currentText() if self.scene_name_combo.currentText() != "请选择场景名称" else "",
            "scene_mapping": self.scene_mapping_combo.currentText() if self.scene_mapping_combo.currentText() != "请选择场景映射表" else "",
            "ai_analysis": self.ai_analysis_checkbox.isChecked(),
            "use_preset": self.use_preset_checkbox.isChecked(),
            "preset_signals": "".join(self._selected_preset_signals) if self._selected_preset_signals else "",
            "preset_scene": self._selected_preset_scene,
            "preset_scene_runtime": self._selected_preset_scene_runtime,
            "record": self.record_checkbox.isChecked(),
        }

    def set_meta_info(self, meta_info: Dict[str, Any]) -> None:
        self.test_point_edit.setText(meta_info.get("test_point", ""))
        self.priority_combo.setCurrentText(meta_info.get("priority", "P1-中"))
        self.owner_edit.setText(meta_info.get("owner", ""))
 
        # 设置场景映射表（先清洗）
        scene_mapping = self._normalize_combo_text(meta_info.get("scene_mapping", ""))
        if scene_mapping:
            idx = self.scene_mapping_combo.findText(scene_mapping, Qt.MatchFlag.MatchExactly)
            if idx >= 0:
                self.scene_mapping_combo.blockSignals(True)
                self.scene_mapping_combo.setCurrentIndex(idx)
                # 手动触发，确保场景名称下拉框已加载
                self._on_scene_mapping_changed(scene_mapping)
                self.scene_mapping_combo.blockSignals(False)
            else:
                self.scene_mapping_combo.setCurrentIndex(0)
        else:
            self.scene_mapping_combo.setCurrentIndex(0)
 
        # 设置场景名称（再清洗）
        scenario_name = self._normalize_combo_text(meta_info.get("scenario_name", ""))
        if scenario_name:
            idx = self.scene_name_combo.findText(scenario_name, Qt.MatchFlag.MatchExactly)
            if idx >= 0:
                self.scene_name_combo.blockSignals(True)
                self.scene_name_combo.setCurrentIndex(idx)
                self._on_scene_name_changed(scenario_name)
                self.scene_name_combo.blockSignals(False)
            else:
                self.scene_name_combo.setCurrentIndex(0)
        else:
            self.scene_name_combo.setCurrentIndex(0)
 
        # scenario_id 直接覆盖（清理尾逗号）
        scenario_id = str(meta_info.get("scenario_id", "")).strip().rstrip(",")
        if scenario_id:
            self.scenario_id_edit.setText(scenario_id)
 
        self.ai_analysis_checkbox.setChecked(bool(meta_info.get("ai_analysis", False)))
        self.use_preset_checkbox.setChecked(bool(meta_info.get("use_preset", False)))

        # 解析 preset_signals（如 "P1P2P3" 转为 ["P1", "P2", "P3"]）
        preset_signals_str = meta_info.get("preset_signals", "")
        if preset_signals_str:
            import re
            self._selected_preset_signals = re.findall(r'P\d+', preset_signals_str)
        else:
            self._selected_preset_signals = []

        self._selected_preset_scene = meta_info.get("preset_scene", "")
        self._selected_preset_scene_runtime = meta_info.get("preset_scene_runtime", "")

        self.record_checkbox.setChecked(bool(meta_info.get("record", False)))

    def refresh_scene_mappings(self) -> None:
        """刷新场景映射表，并尽量保留当前选择"""
        prev_mapping = self.scene_mapping_combo.currentText()
        prev_scene = self.scene_name_combo.currentText()
    
        # 刷新过程中阻断信号，避免误触发“内容已修改”
        blocker_mapping = QSignalBlocker(self.scene_mapping_combo)
        blocker_scene = QSignalBlocker(self.scene_name_combo)
        try:
            self._load_scene_mappings()
    
            # 恢复映射表选择（如果还存在）
            mapping_idx = self.scene_mapping_combo.findText(prev_mapping, Qt.MatchFlag.MatchExactly)
            if mapping_idx >= 0 and prev_mapping and prev_mapping != "请选择场景映射表":
                self.scene_mapping_combo.setCurrentIndex(mapping_idx)
                self._on_scene_mapping_changed(prev_mapping)
    
                # 恢复场景名称选择（如果还存在）
                scene_idx = self.scene_name_combo.findText(prev_scene, Qt.MatchFlag.MatchExactly)
                if scene_idx >= 0 and prev_scene and prev_scene != "请选择场景名称":
                    self.scene_name_combo.setCurrentIndex(scene_idx)
                    self._on_scene_name_changed(prev_scene)
                else:
                    self.scene_name_combo.setCurrentIndex(0)
                    self._on_scene_name_changed(self.scene_name_combo.currentText())
            else:
                # 原映射表被删除或无效，回到默认
                self.scene_mapping_combo.setCurrentIndex(0)
                self._on_scene_mapping_changed(self.scene_mapping_combo.currentText())
        finally:
            del blocker_scene
            del blocker_mapping

    def _on_use_preset_changed(self, checked: bool) -> None:
        """启用预设复选框状态变化"""
        self.select_preset_btn.setEnabled(checked)
        if not checked:
            self._selected_preset_signals = []
            self._selected_preset_scene = ""
            self._selected_preset_scene_runtime = ""

    def _on_select_preset(self) -> None:
        """打开预设选择对话框"""
        if not self.project_manager or not self.project_manager.is_project_open():
            return

        # 获取预设配置
        config = self.project_manager.project_config
        automation = config.get("automation", {})
        set_preset = automation.get("set_preset", {})
        preset_signals = set_preset.get("preset_signals", [])
        preset_scene = set_preset.get("preset_scene", {})

        if not preset_signals and not preset_scene:
            QMessageBox.information(self, "提示", "请先在项目菜单 -> Automation -> 设置预设 中配置预设信息")
            return

        # 创建预设选择对话框 - 简洁风格设计
        from PyQt6.QtWidgets import (
            QDialog, QVBoxLayout, QHBoxLayout, QCheckBox, QDialogButtonBox,
            QGroupBox, QTableWidget, QTableWidgetItem, QHeaderView, QLabel,
            QFrame, QScrollArea
        )
        dlg = QDialog(self)
        dlg.setWindowTitle("选择预设")
        dlg.resize(700, 450)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QTableWidget {
                border: none;
                background-color: transparent;
                alternate-background-color: transparent;
                gridline-color: transparent;
                color: #333333;
                font-size: 12px;
                outline: none;
                padding: 0px;
                margin: 0px;
            }
            QTableWidget::item {
                padding-left: 8px;
                padding-right: 5px;
                padding-top: 8px;
                padding-bottom: 8px;
                margin: 0px;
                border: none;
                outline: none;
                background-color: transparent;
                border-bottom: 1px solid #e0e0e0;
            }
            QTableWidget::item:selected {
                background-color: transparent;
                color: #000000;
            }
            QTableWidget::item:hover {
                background-color: #f0f0f0;
            }
            QHeaderView::section {
                background-color: transparent;
                padding-left: 8px;
                padding-right: 5px;
                padding-top: 8px;
                padding-bottom: 8px;
                margin: 0px;
                border: none;
                border-bottom: 2px solid #007acc;
                font-weight: bold;
                font-size: 12px;
                color: #007acc;
            }
            QCheckBox {
                spacing: 8px;
                color: #333333;
                background-color: transparent;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border-radius: 3px;
                border: 2px solid #007acc;
                background-color: #ffffff;
            }
            QCheckBox::indicator:checked {
                background-color: #007acc;
                border-color: #007acc;
                image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxMiIgaGVpZ2h0PSIxMiIgdmlld0JveD0iMCAwIDI0IDI0IiBmaWxsPSJub25lIiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjMiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCI+PHBvbHlsaW5lIHBvaW50cz0iMjAgNiA5IDE3IDQgMTIiPjwvcG9seWxpbmU+PC9zdmc+);
            }
            QCheckBox::indicator:hover {
                border-color: #005a9e;
            }
            QPushButton {
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
                border: none;
            }
            QPushButton#btnOk {
                background-color: #007acc;
                color: white;
            }
            QPushButton#btnOk:hover {
                background-color: #0098ff;
            }
            QPushButton#btnOk:pressed {
                background-color: #005a9e;
            }
            QPushButton#btnCancel {
                background-color: #e0e0e0;
                color: #333333;
            }
            QPushButton#btnCancel:hover {
                background-color: #d0d0d0;
            }
            QLabel {
                color: #333333;
                background-color: transparent;
            }
            QLabel#sectionTitle {
                font-weight: bold;
                font-size: 14px;
                color: #007acc;
                padding: 5px 0px;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setSpacing(12)
        layout.setContentsMargins(25, 20, 25, 20)

        # 预设信号选择 - 使用表格
        if preset_signals:
            # 标题
            signal_title = QLabel("预设信号 (Preset Signals)", dlg)
            signal_title.setObjectName("sectionTitle")
            layout.addWidget(signal_title)

            # 创建表格
            self._signal_table = QTableWidget(dlg)
            self._signal_table.setColumnCount(4)
            self._signal_table.setHorizontalHeaderLabels(["选择", "ID", "信号名称", "注释"])
            self._signal_table.setRowCount(len(preset_signals))
            self._signal_table.verticalHeader().setVisible(False)
            self._signal_table.setAlternatingRowColors(False)
            self._signal_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            self._signal_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            self._signal_table.setShowGrid(False)
            self._signal_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            self._signal_table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)

            # 设置列宽
            header = self._signal_table.horizontalHeader()
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
            header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
            header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
            header.resizeSection(0, 60)
            header.resizeSection(1, 80)
            header.setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)

            # 填充数据
            for row, signal in enumerate(preset_signals):
                signal_id = signal.get("id", "")
                signal_name = signal.get("signal_name", "")
                signal_value = signal.get("signal_value", "")
                comment = signal.get("comment", "")

                # 复选框 - 放在容器中实现左对齐
                cb_widget = QWidget(self._signal_table)
                cb_layout = QHBoxLayout(cb_widget)
                cb_layout.setContentsMargins(8, 0, 5, 0)
                cb_layout.setAlignment(Qt.AlignmentFlag.AlignLeft)
                cb = QCheckBox(cb_widget)
                cb.setChecked(signal_id in self._selected_preset_signals)
                cb_layout.addWidget(cb)
                self._signal_table.setCellWidget(row, 0, cb_widget)

                # ID
                id_item = QTableWidgetItem(signal_id)
                id_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self._signal_table.setItem(row, 1, id_item)

                # 信号名称
                name_item = QTableWidgetItem(signal_name)
                name_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                name_item.setToolTip(f"值: {signal_value}")
                self._signal_table.setItem(row, 2, name_item)

                # 注释
                comment_item = QTableWidgetItem(comment if comment else "-")
                comment_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self._signal_table.setItem(row, 3, comment_item)

            self._signal_table.resizeRowsToContents()
            layout.addWidget(self._signal_table)
            layout.addSpacing(15)

        # 预设场景选择
        if preset_scene and preset_scene.get("scene_name"):
            # 标题
            scene_title = QLabel("预设场景 (Preset Scene)", dlg)
            scene_title.setObjectName("sectionTitle")
            layout.addWidget(scene_title)

            # 场景信息容器
            scene_widget = QWidget(dlg)
            scene_widget.setStyleSheet("background-color: transparent;")
            scene_widget_layout = QHBoxLayout(scene_widget)
            scene_widget_layout.setContentsMargins(10, 8, 10, 8)
            scene_widget_layout.setSpacing(15)

            # 场景选择复选框
            self._scene_checkbox = QCheckBox(scene_widget)
            self._scene_checkbox.setChecked(bool(self._selected_preset_scene))
            scene_widget_layout.addWidget(self._scene_checkbox)

            # 场景信息
            info_widget = QWidget(scene_widget)
            info_widget.setStyleSheet("background-color: transparent;")
            info_layout = QVBoxLayout(info_widget)
            info_layout.setContentsMargins(0, 0, 0, 0)
            info_layout.setSpacing(4)

            scene_name = preset_scene.get("scene_name", "")
            scene_id = preset_scene.get("scene_id", "")
            runtime = preset_scene.get("runtime", "")

            name_label = QLabel(scene_name, info_widget)
            name_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #007acc;")
            info_layout.addWidget(name_label)

            detail_label = QLabel(f"ID: {scene_id}  |  运行时间: {runtime}ms" if runtime else f"ID: {scene_id}", info_widget)
            detail_label.setStyleSheet("color: #666666; font-size: 11px;")
            info_layout.addWidget(detail_label)

            scene_widget_layout.addWidget(info_widget, 1)
            layout.addWidget(scene_widget)
        else:
            self._scene_checkbox = None

        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_cancel = QPushButton("取 消", dlg)
        btn_cancel.setObjectName("btnCancel")
        btn_cancel.setMinimumWidth(90)
        btn_cancel.clicked.connect(dlg.reject)
        btn_layout.addWidget(btn_cancel)

        btn_ok = QPushButton("确 定", dlg)
        btn_ok.setObjectName("btnOk")
        btn_ok.setMinimumWidth(90)
        btn_ok.setDefault(True)
        btn_ok.clicked.connect(dlg.accept)
        btn_layout.addWidget(btn_ok)

        layout.addLayout(btn_layout)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            # 收集选中的预设信号
            self._selected_preset_signals = []
            if hasattr(self, '_signal_table'):
                for row in range(self._signal_table.rowCount()):
                    cb_widget = self._signal_table.cellWidget(row, 0)
                    if cb_widget:
                        # 从容器widget中找到checkbox
                        cb = cb_widget.findChild(QCheckBox)
                        if cb and cb.isChecked():
                            id_item = self._signal_table.item(row, 1)
                            if id_item:
                                self._selected_preset_signals.append(id_item.text())

            # 收集选中的预设场景
            if self._scene_checkbox and self._scene_checkbox.isChecked():
                self._selected_preset_scene = preset_scene.get("scene_id", "")
                self._selected_preset_scene_runtime = preset_scene.get("runtime", "")
            else:
                self._selected_preset_scene = ""
                self._selected_preset_scene_runtime = ""

            # 发出信号通知内容已变化
            self.preset_selection_changed.emit()

    def get_selected_presets(self) -> Dict[str, Any]:
        """获取选中的预设信息"""
        return {
            "preset_signals": self._selected_preset_signals.copy(),
            "preset_scene": self._selected_preset_scene
        }

    def set_selected_presets(self, preset_signals: List[str], preset_scene: str) -> None:
        """设置选中的预设信息"""
        self._selected_preset_signals = preset_signals.copy() if preset_signals else []
        self._selected_preset_scene = preset_scene or ""

    def clear(self) -> None:
        self.test_point_edit.clear()
        self.priority_combo.setCurrentIndex(1)
        self.owner_edit.clear()
        self.scenario_id_edit.clear()
        self.scene_mapping_combo.setCurrentIndex(0)
        self.scene_name_combo.setCurrentIndex(0)
        self.ai_analysis_checkbox.setChecked(False)
        self.use_preset_checkbox.setChecked(False)
        self._selected_preset_signals = []
        self._selected_preset_scene = ""
        self._selected_preset_scene_runtime = ""
        self.record_checkbox.setChecked(False)


# ----------------------------
# 主界面 Step 行（只读）
# ----------------------------
class StepWidget(QWidget):
    add_requested = pyqtSignal(object)     # self
    delete_requested = pyqtSignal(object)  # self
    edit_requested = pyqtSignal(object)    # self
    template_selected = pyqtSignal(object, dict)  # self, template_data

    def __init__(self, step_id: str, step_type: str = "SET", parent=None):
        super().__init__(parent)
        self.step_id = step_id
        self.step_type = step_type
        self._build_ui()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        row = QWidget(self)
        row_layout = QHBoxLayout(row)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(FORM_HSPACING)

        self.id_label = QLabel(f"{self.step_id}:", row)
        self.id_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.id_label.setStyleSheet("color: #569CD6;")
        self.id_label.setFixedWidth(FORM_LABEL_WIDTH)
        self.id_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        content_container = QWidget(row)
        content_layout = QHBoxLayout(content_container)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(6)

        self.content_edit = ReadOnlyLineEdit(content_container)
        if self.step_type == "SET":
            self.content_edit.setPlaceholderText("set sys::FunctionSwitch::CSW_Enable_S=0x1 wait 500ms then check C1")
        else:
            self.content_edit.setPlaceholderText(
                "check sig::CAN 1::ADC_0x29C::CSW_Stats_S==3 timeout 1000ms async false"
            )
        self.content_edit.setMinimumWidth(480)
        self.content_edit.setStyleSheet(
            """
            QLineEdit {
                border: 1px solid #cccccc;
                border-radius: 3px;
                padding: 5px;
                background-color: #f7f7f7;
            }
            QLineEdit:read-only {
                color: #333333;
            }
            """
        )
        content_layout.addWidget(self.content_edit, 1)

        # 模板选择按钮（点击弹出菜单）
        self.template_btn = QPushButton("选择模板", content_container)
        self.template_btn.setFixedWidth(75)
        self.template_btn.setStyleSheet(
            """
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 5px;
            }
            QPushButton:hover { background-color: #1976D2; }
            """
        )
        self.template_menu = QMenu(self.template_btn)
        self.template_menu.setStyleSheet(
            """
            QMenu {
                background-color: white;
                color: #333333;
            }
            QMenu::item:selected {
                background-color: #e3f2fd;
                color: #1976D2;
            }
            """
        )
        self.template_btn.clicked.connect(self._show_template_menu)
        content_layout.addWidget(self.template_btn)

        # 编辑按钮（蓝色，高度与选择模板一致）
        self.btn_edit = QPushButton("编辑", content_container)
        self.btn_edit.setFixedWidth(68)
        self.btn_edit.setFixedHeight(self.template_btn.sizeHint().height())
        self.btn_edit.setStyleSheet(
            """
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover { background-color: #1976D2; }
            """
        )
        self.btn_edit.clicked.connect(lambda: self.edit_requested.emit(self))
        content_layout.addWidget(self.btn_edit)

        self.btn_add = QPushButton("+", content_container)
        self.btn_add.setFixedWidth(68)
        self.btn_add.setFixedHeight(self.template_btn.sizeHint().height())
        self.btn_add.setStyleSheet(
            """
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 3px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover { background-color: #45a049; }
            """
        )
        self.btn_add.clicked.connect(lambda: self.add_requested.emit(self))
        content_layout.addWidget(self.btn_add)

        self.btn_del = QPushButton("-", content_container)
        self.btn_del.setFixedWidth(68)
        self.btn_del.setFixedHeight(self.template_btn.sizeHint().height())
        self.btn_del.setStyleSheet(
            """
            QPushButton {
                background-color: #ff6b6b;
                color: white;
                border: none;
                border-radius: 3px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover { background-color: #ff5252; }
            """
        )
        self.btn_del.clicked.connect(lambda: self.delete_requested.emit(self))
        content_layout.addWidget(self.btn_del)

        row_layout.addWidget(self.id_label)
        row_layout.addWidget(content_container, 1)
        root.addWidget(row)

    def _show_template_menu(self) -> None:
        """显示模板选择菜单"""
        self.template_menu.exec(self.template_btn.mapToGlobal(self.template_btn.rect().bottomLeft()))

    def _on_template_selected(self, template_data: Dict[str, Any]) -> None:
        """模板选择处理"""
        self.template_selected.emit(self, template_data)

    def update_templates(self, templates: List[Dict[str, Any]]) -> None:
        """更新模板列表"""
        self.template_menu.clear()
        for template in templates:
            comment = template.get("comment", "")
            action = self.template_menu.addAction(comment)
            action.triggered.connect(lambda checked, t=template: self._on_template_selected(t))

    def set_step_id(self, new_id: str) -> None:
        self.step_id = new_id
        self.id_label.setText(f"{new_id}:")

    def get_step_content(self) -> str:
        return self.content_edit.text().strip()

    def set_step_content(self, content: Any) -> None:
        if content is None or content is False:
            content = ""
        elif not isinstance(content, str):
            content = str(content)
        self.content_edit.setText(content)


# ----------------------------
# SET 模块（顺序/删除稳定 + 弹窗编辑）
# ----------------------------
class SetModuleWidget(QWidget):
    steps_changed = pyqtSignal()

    def __init__(self, parent=None, dbc_parser=None, project_manager=None):
        super().__init__(parent)
        self.steps: List[StepWidget] = []
        self.completions: List[str] = []
        self.get_check_ids_provider = None  # type: ignore[assignment]
        self._dbc_parser = dbc_parser
        self._project_manager = project_manager
        self._build_ui()
        self.add_step(after_step=None)

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        group = QGroupBox("SET模块", self)
        _setup_groupbox_style(group)
        # 设置最小高度，使其可以显示3个step
        group.setMinimumHeight(200)

        group_layout = QVBoxLayout(group)
        group_layout.setContentsMargins(0, 0, 0, 0)
        group_layout.setSpacing(0)

        scroll = QScrollArea(group)
        _setup_scroll(scroll)

        self.steps_container = QWidget(scroll)
        self.steps_container.setContentsMargins(0, 0, 0, 0)

        self.steps_layout = QVBoxLayout(self.steps_container)
        self.steps_layout.setContentsMargins(0, 0, 0, 0)
        self.steps_layout.setSpacing(6)
        self.steps_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        scroll.setWidget(self.steps_container)
        group_layout.addWidget(scroll)
        layout.addWidget(group)

    def set_completions(self, completions: List[str]) -> None:
        self.completions = completions

    def _completions_by_kind(self) -> Dict[str, List[str]]:
        sys_list = [c for c in self.completions if c.startswith("sys::")]
        env_list = [c for c in self.completions if c.startswith("env::")]
        sig_list = [c for c in self.completions if c.startswith("sig::")]
        return {"sys": sys_list, "env": env_list, "sig": sig_list}

    def _get_set_templates(self) -> List[Dict[str, Any]]:
        """获取SET模板列表"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return []
        config = self._project_manager.project_config
        return config.get("automation", {}).get("set_template", {}).get("templates", [])

    def add_step(self, content: Any = "", after_step: Optional[StepWidget] = None) -> None:
        step = StepWidget("S?", "SET", self.steps_container)
        step.set_step_content(content)

        step.add_requested.connect(self._on_add_requested)
        step.delete_requested.connect(self._on_delete_requested)
        step.edit_requested.connect(self._on_edit_requested)
        step.template_selected.connect(self._on_template_selected)

        # 更新模板列表
        templates = self._get_set_templates()
        step.update_templates(templates)

        if after_step is not None and after_step in self.steps:
            idx = self.steps.index(after_step) + 1
            self.steps.insert(idx, step)
        else:
            self.steps.append(step)

        self._rebuild_layout()
        self._renumber_steps()
        self.steps_changed.emit()

    def _on_add_requested(self, step: StepWidget) -> None:
        self.add_step(after_step=step)

    def _on_delete_requested(self, step: StepWidget) -> None:
        self.remove_step(step)

    def _on_template_selected(self, step: StepWidget, template: Dict[str, Any]) -> None:
        """处理模板选择"""
        self._apply_template(step, template)

    def _apply_template(self, step: StepWidget, template: Dict[str, Any]) -> None:
        """应用模板到SET步骤"""
        # 获取当前内容，用于保留wait和then check
        raw = step.get_step_content()

        # 解析当前内容
        model, ok, _ = parse_set_step(raw)

        # 应用模板：添加信号到signals列表
        signal_name = template.get("signal_name", "")
        signal_value = template.get("signal_value", "")
        comment = template.get("comment", "")

        # 创建新的信号
        kind = "sys"
        if signal_name.startswith("env::"):
            kind = "env"
        new_signal = SetSignalModel(kind=kind, name=signal_name, value=signal_value)

        # 如果模板有信号，添加到现有signals中（或替换空的信号）
        if signal_name:
            # 检查是否需要替换空的信号
            if model.signals and model.signals[0].name == "":
                model.signals[0] = new_signal
            else:
                model.signals.append(new_signal)

        # 设置comment
        if comment:
            model.comment = comment

        # 生成新的DSL并设置
        new_dsl = render_set_step(model)
        step.set_step_content(new_dsl)
        self.steps_changed.emit()

    def _on_edit_requested(self, step: StepWidget) -> None:
        raw = step.get_step_content()
        check_ids: List[str] = []
        if callable(self.get_check_ids_provider):
            try:
                check_ids = list(self.get_check_ids_provider())
            except Exception:
                check_ids = []

        dlg = SetStepDialog(
            raw_text=raw,
            completions_by_kind={"sys": self._completions_by_kind()["sys"], "env": self._completions_by_kind()["env"]},
            available_check_ids=check_ids,
            parent=self,
            dbc_parser=self._dbc_parser,
        )
        # 使用 show() 代替 exec()，使对话框变为非模态，不阻塞其他窗口
        dlg.finished.connect(lambda result: self._on_set_dialog_finished(result, dlg, step))
        dlg.show()

    def _on_set_dialog_finished(self, result, dlg, step):
        """SET对话框关闭后的回调"""
        if result == QDialog.DialogCode.Accepted:
            step.set_step_content(dlg.to_dsl())
            self.steps_changed.emit()

    def remove_step(self, step: StepWidget) -> None:
        if step not in self.steps:
            return
        if len(self.steps) <= 1:
            step.set_step_content("")
            return

        self.steps.remove(step)
        self.steps_layout.removeWidget(step)
        step.setParent(None)
        step.deleteLater()

        self._rebuild_layout()
        self._renumber_steps()
        self.steps_changed.emit()

    def _rebuild_layout(self) -> None:
        while self.steps_layout.count():
            item = self.steps_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                self.steps_layout.removeWidget(w)
        for s in self.steps:
            self.steps_layout.addWidget(s)

    def _renumber_steps(self) -> None:
        for i, s in enumerate(self.steps, start=1):
            s.set_step_id(f"S{i}")

    def apply_check_id_mapping(self, mapping: Dict[str, str]) -> None:
        if not mapping:
            return

        # 支持 then check C1 / then check C1,C2 / then check C1, C2
        pat = re.compile(r"(then\s+check\s+)(C\d+(?:\s*,\s*C\d+)*)", re.IGNORECASE)

        for step in self.steps:
            text = step.get_step_content()
            if not text:
                continue

            def repl(m: re.Match) -> str:
                head = m.group(1)
                ids_part = m.group(2)
                ids = [x.strip() for x in ids_part.split(",") if x.strip()]
                ids2 = [mapping.get(x, x) for x in ids]
                return head + ",".join(ids2)

            new_text = pat.sub(repl, text)
            if new_text != text:
                step.set_step_content(new_text)

    def get_steps(self) -> List[Dict[str, str]]:
        return [{"id": s.step_id, "content": s.get_step_content()} for s in self.steps]

    def set_steps(self, steps: List[Dict[str, Any]]) -> None:
        for s in self.steps:
            self.steps_layout.removeWidget(s)
            s.setParent(None)
            s.deleteLater()
        self.steps.clear()

        for step_data in steps:
            self.add_step(content=step_data.get("content", ""), after_step=self.steps[-1] if self.steps else None)

        if not self.steps:
            self.add_step(after_step=None)

    def clear(self) -> None:
        for s in self.steps[1:]:
            self.steps_layout.removeWidget(s)
            s.setParent(None)
            s.deleteLater()

        if self.steps:
            self.steps[0].set_step_content("")
            self.steps = [self.steps[0]]
            self._rebuild_layout()
            self._renumber_steps()
            self.steps_changed.emit()
        else:
            self.add_step(after_step=None)

    def refresh_templates(self) -> None:
        """刷新所有步骤的模板列表"""
        templates = self._get_set_templates()
        for step in self.steps:
            step.update_templates(templates)


# ----------------------------
# CHECK 模块（顺序/删除稳定 + 弹窗编辑）
# ----------------------------
class CheckModuleWidget(QWidget):
    steps_changed = pyqtSignal()
    check_id_mapping_emitted = pyqtSignal(dict)  # old->new

    def __init__(self, parent=None, dbc_parser=None, project_manager=None):
        super().__init__(parent)
        self.steps: List[StepWidget] = []
        self.completions: List[str] = []
        self._dbc_parser = dbc_parser
        self._project_manager = project_manager
        self._build_ui()
        self.add_step(after_step=None)

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        group = QGroupBox("CHECK模块", self)
        _setup_groupbox_style(group)
        # 设置最小高度，使其可以显示3个step
        group.setMinimumHeight(200)

        group_layout = QVBoxLayout(group)
        group_layout.setContentsMargins(0, 0, 0, 0)
        group_layout.setSpacing(0)

        scroll = QScrollArea(group)
        _setup_scroll(scroll)

        self.steps_container = QWidget(scroll)
        self.steps_container.setContentsMargins(0, 0, 0, 0)

        self.steps_layout = QVBoxLayout(self.steps_container)
        self.steps_layout.setContentsMargins(0, 0, 0, 0)
        self.steps_layout.setSpacing(6)
        self.steps_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        scroll.setWidget(self.steps_container)
        group_layout.addWidget(scroll)
        layout.addWidget(group)

    def set_completions(self, completions: List[str]) -> None:
        self.completions = completions

    def _completions_by_kind(self) -> Dict[str, List[str]]:
        sys_list = [c for c in self.completions if c.startswith("sys::")]
        env_list = [c for c in self.completions if c.startswith("env::")]
        sig_list = [c for c in self.completions if c.startswith("sig::")]
        return {"sys": sys_list, "env": env_list, "sig": sig_list}

    def _get_check_templates(self) -> List[Dict[str, Any]]:
        """获取CHECK模板列表"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return []
        config = self._project_manager.project_config
        return config.get("automation", {}).get("check_template", {}).get("templates", [])

    def add_step(self, content: Any = "", after_step: Optional[StepWidget] = None) -> None:
        step = StepWidget("C?", "CHECK", self.steps_container)
        step.set_step_content(content)

        step.add_requested.connect(self._on_add_requested)
        step.delete_requested.connect(self._on_delete_requested)
        step.edit_requested.connect(self._on_edit_requested)
        step.template_selected.connect(self._on_template_selected)

        # 更新模板列表
        templates = self._get_check_templates()
        step.update_templates(templates)

        if after_step is not None and after_step in self.steps:
            idx = self.steps.index(after_step) + 1
            self.steps.insert(idx, step)
        else:
            self.steps.append(step)

        self._rebuild_layout()
        mapping = self._renumber_steps_and_get_mapping()
        self.steps_changed.emit()
        if mapping:
            self.check_id_mapping_emitted.emit(mapping)

    def _on_add_requested(self, step: StepWidget) -> None:
        self.add_step(after_step=step)

    def _on_delete_requested(self, step: StepWidget) -> None:
        self.remove_step(step)

    def _on_template_selected(self, step: StepWidget, template: Dict[str, Any]) -> None:
        """处理模板选择"""
        self._apply_template(step, template)

    def _apply_template(self, step: StepWidget, template: Dict[str, Any]) -> None:
        """应用模板到CHECK步骤"""
        # 获取模板数据
        signal_name = template.get("signal_name", "")
        value_mode = template.get("value_mode", "single")
        operator = template.get("operator", "==")
        signal_value = template.get("signal_value", "")
        comment = template.get("comment", "")

        # 确定信号类型
        kind = "sig"
        if signal_name.startswith("sys::"):
            kind = "sys"
        elif signal_name.startswith("env::"):
            kind = "env"

        # 创建CheckItemModel
        item = CheckItemModel(
            kind=kind,
            name=signal_name,
            mode=value_mode,
            op=operator,
            single_value=signal_value if value_mode == "single" else "",
            list_values=[x.strip() for x in signal_value.split(",")] if value_mode == "list" else [],
            range_a=signal_value.split("..")[0].strip() if value_mode == "range" and ".." in signal_value else "",
            range_b=signal_value.split("..")[1].strip() if value_mode == "range" and ".." in signal_value else "",
        )

        # 创建CheckStepModel
        model = CheckStepModel(items=[item], comment=comment)

        # 生成DSL并设置
        new_dsl = render_check_step(model)
        step.set_step_content(new_dsl)
        self.steps_changed.emit()

    def _on_edit_requested(self, step: StepWidget) -> None:
        raw = step.get_step_content()
        dlg = CheckStepDialog(raw_text=raw, completions_by_kind=self._completions_by_kind(), parent=self, dbc_parser=self._dbc_parser)
        # 使用 show() 代替 exec()，使对话框变为非模态，不阻塞其他窗口
        dlg.finished.connect(lambda result: self._on_check_dialog_finished(result, dlg, step))
        dlg.show()

    def _on_check_dialog_finished(self, result, dlg, step):
        """CHECK对话框关闭后的回调"""
        if result == QDialog.DialogCode.Accepted:
            step.set_step_content(dlg.to_dsl())
            self.steps_changed.emit()

    def remove_step(self, step: StepWidget) -> None:
        if step not in self.steps:
            return
        if len(self.steps) <= 1:
            step.set_step_content("")
            return

        self.steps.remove(step)
        self.steps_layout.removeWidget(step)
        step.setParent(None)
        step.deleteLater()

        self._rebuild_layout()
        mapping = self._renumber_steps_and_get_mapping()
        self.steps_changed.emit()
        if mapping:
            self.check_id_mapping_emitted.emit(mapping)

    def _rebuild_layout(self) -> None:
        while self.steps_layout.count():
            item = self.steps_layout.takeAt(0)
            w = item.widget()
            if w is not None:
                self.steps_layout.removeWidget(w)
        for s in self.steps:
            self.steps_layout.addWidget(s)

    def _renumber_steps_and_get_mapping(self) -> Dict[str, str]:
        mapping: Dict[str, str] = {}
        for i, s in enumerate(self.steps, start=1):
            new_id = f"C{i}"
            if s.step_id != new_id and s.step_id not in ("C?", ""):
                mapping[s.step_id] = new_id
            s.set_step_id(new_id)
        return {k: v for k, v in mapping.items() if k != v}

    def get_steps(self) -> List[Dict[str, str]]:
        return [{"id": s.step_id, "content": s.get_step_content()} for s in self.steps]

    def set_steps(self, steps: List[Dict[str, Any]]) -> None:
        for s in self.steps:
            self.steps_layout.removeWidget(s)
            s.setParent(None)
            s.deleteLater()
        self.steps.clear()

        for step_data in steps:
            self.add_step(content=step_data.get("content", ""), after_step=self.steps[-1] if self.steps else None)

        if not self.steps:
            self.add_step(after_step=None)

    def clear(self) -> None:
        for s in self.steps[1:]:
            self.steps_layout.removeWidget(s)
            s.setParent(None)
            s.deleteLater()

        if self.steps:
            self.steps[0].set_step_content("")
            self.steps = [self.steps[0]]
            self._rebuild_layout()
            self._renumber_steps_and_get_mapping()
            self.steps_changed.emit()
        else:
            self.add_step(after_step=None)

    def refresh_templates(self) -> None:
        """刷新所有步骤的模板列表"""
        templates = self._get_check_templates()
        for step in self.steps:
            step.update_templates(templates)


# ----------------------------
# 顶层编辑器：组合模块 + 导入导出 DSL
# ----------------------------
class ModularCaseEditor(QWidget):
    content_changed = pyqtSignal()
    save_to_file_requested = pyqtSignal()  # 新增：请求保存到文件的信号
    status_updated = pyqtSignal(str)  # 状态更新信号，用于通知主窗口更新状态栏

    def __init__(self, parent=None, dbc_parser=None, project_manager=None):
        super().__init__(parent)
        self.completions: List[str] = []
        self._dbc_parser = dbc_parser
        self.project_manager = project_manager
        self._build_ui()
        self._wire_signals()
        
    def refresh_scene_mappings(self) -> None:
        """供主窗口调用，刷新 META 模块中的场景映射数据"""
        self.meta_info_widget.refresh_scene_mappings()

    def refresh_all_templates(self) -> None:
        """刷新所有模块的模板列表"""
        self.set_module_widget.refresh_templates()
        self.check_module_widget.refresh_templates()

    @staticmethod
    def _encode_meta_value(value: Any) -> str:
        # 保持纯 key=value，不自动加引号
        s = str(value if value is not None else "").strip()
        if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
            s = s[1:-1].strip()
        return s

    def _emit_content_changed(self, *_args) -> None:
        sig = getattr(self, "content_changed", None)
        if sig is not None:
            sig.emit()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea(self)
        _setup_scroll(scroll)

        container = QWidget(scroll)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(8)

        self.case_info_widget = CaseInfoWidget(container)
        container_layout.addWidget(self.case_info_widget)

        self.meta_info_widget = MetaInfoWidget(container, project_manager=self.project_manager)
        self.meta_info_widget.status_updated.connect(self.status_updated.emit)  # 转发状态更新信号
        container_layout.addWidget(self.meta_info_widget)

        self.set_module_widget = SetModuleWidget(container, dbc_parser=self._dbc_parser, project_manager=self.project_manager)
        container_layout.addWidget(self.set_module_widget)

        self.check_module_widget = CheckModuleWidget(container, dbc_parser=self._dbc_parser, project_manager=self.project_manager)
        container_layout.addWidget(self.check_module_widget)

        # 添加"查看流程配置"按钮
        button_row = QHBoxLayout()
        button_row.setContentsMargins(6, 6, 6, 6)
        
        self.view_flowchart_btn = QPushButton("查看流程配置")
        self.view_flowchart_btn.setStyleSheet(
            """
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            """
        )
        self.view_flowchart_btn.clicked.connect(self._on_view_flowchart)
        button_row.addWidget(self.view_flowchart_btn)
        button_row.addStretch()
        
        container_layout.addLayout(button_row)

        container_layout.addStretch()
        scroll.setWidget(container)
        layout.addWidget(scroll)

    def _wire_signals(self) -> None:
        self.check_module_widget.check_id_mapping_emitted.connect(self.set_module_widget.apply_check_id_mapping)
        self.set_module_widget.get_check_ids_provider = lambda: [s.step_id for s in self.check_module_widget.steps]
 
        # 任意变化 -> 标记内容已变更
        self.case_info_widget.case_name_edit.textChanged.connect(self._emit_content_changed)
        self.case_info_widget.case_id_edit.textChanged.connect(self._emit_content_changed)
 
        self.meta_info_widget.test_point_edit.textChanged.connect(self._emit_content_changed)
        self.meta_info_widget.priority_combo.currentTextChanged.connect(self._emit_content_changed)
        self.meta_info_widget.owner_edit.textChanged.connect(self._emit_content_changed)
        self.meta_info_widget.scenario_id_edit.textChanged.connect(self._emit_content_changed)
        self.meta_info_widget.scene_name_combo.currentTextChanged.connect(self._emit_content_changed)
        self.meta_info_widget.scene_mapping_combo.currentTextChanged.connect(self._emit_content_changed)
        self.meta_info_widget.ai_analysis_checkbox.toggled.connect(self._emit_content_changed)
        self.meta_info_widget.use_preset_checkbox.toggled.connect(self._emit_content_changed)
        self.meta_info_widget.preset_selection_changed.connect(self._emit_content_changed)
        self.meta_info_widget.record_checkbox.toggled.connect(self._emit_content_changed)
 
        self.set_module_widget.steps_changed.connect(self._emit_content_changed)
        self.check_module_widget.steps_changed.connect(self._emit_content_changed)
        
        # 更新流程图按钮状态
        self.set_module_widget.steps_changed.connect(self._update_flowchart_button_state)
        self.check_module_widget.steps_changed.connect(self._update_flowchart_button_state)
        self._update_flowchart_button_state()

    def set_completions(self, completions: List[str]) -> None:
        self.completions = completions
        self.set_module_widget.set_completions(completions)
        self.check_module_widget.set_completions(completions)
 
    # 给 MainWindow 的编辑菜单用：把操作派发给当前焦点控件
    def _dispatch_to_focused(self, method_name: str) -> None:
        w = QApplication.focusWidget()
        if w is None:
            return
        if not self.isAncestorOf(w) and w is not self:
            return
        fn = getattr(w, method_name, None)
        if callable(fn):
            fn()
 
    def undo(self) -> None:
        self._dispatch_to_focused("undo")
 
    def redo(self) -> None:
        self._dispatch_to_focused("redo")
 
    def cut(self) -> None:
        self._dispatch_to_focused("cut")
 
    def copy(self) -> None:
        self._dispatch_to_focused("copy")
 
    def paste(self) -> None:
        """禁用粘贴功能"""
        # 不执行任何操作，禁用 Ctrl+V 粘贴功能
        pass
 
    def selectAll(self) -> None:
        self._dispatch_to_focused("selectAll")

    def get_case_data(self) -> Dict[str, Any]:
        return {
            "case_info": self.case_info_widget.get_case_info(),
            "meta_info": self.meta_info_widget.get_meta_info(),
            "set_steps": self.set_module_widget.get_steps(),
            "check_steps": self.check_module_widget.get_steps(),
        }

    def set_case_data(self, case_data: Dict[str, Any]) -> None:
        self.case_info_widget.set_case_info(
            case_data.get("case_info", {}).get("name", ""),
            case_data.get("case_info", {}).get("id", ""),
        )
        self.meta_info_widget.set_meta_info(case_data.get("meta_info", {}))
        self.set_module_widget.set_steps(case_data.get("set_steps", []))
        self.check_module_widget.set_steps(case_data.get("check_steps", []))

    def to_dsl(self) -> str:
        data = self.get_case_data()
        dsl_lines: List[str] = []
        # print(data)
 
        case_info = data["case_info"]
        if case_info.get("name"):
            dsl_lines.append(f"CASE: {case_info['name']}")
 
        meta_info = data["meta_info"]
        meta_parts: List[str] = []
        
        if case_info.get("id"):
            meta_parts.append(f"case_id={self._encode_meta_value(case_info['id'])}")
 
        if meta_info.get("test_point"):
            meta_parts.append(f"test_point={self._encode_meta_value(meta_info['test_point'])}")
        if meta_info.get("priority"):
            meta_parts.append(f"priority={self._encode_meta_value(meta_info['priority'])}")
        if meta_info.get("owner"):
            meta_parts.append(f"owner={self._encode_meta_value(meta_info['owner'])}")
        if meta_info.get("scene_mapping"):
            meta_parts.append(f"scene_mapping={self._encode_meta_value(meta_info['scene_mapping'])}")
        if meta_info.get("scenario_id"):
            meta_parts.append(f"scenario_id={self._encode_meta_value(meta_info['scenario_id'])}")
        if meta_info.get("scenario_name"):
            meta_parts.append(f"scenario_name={self._encode_meta_value(meta_info['scenario_name'])}")
        if meta_info.get("ai_analysis"):
            meta_parts.append("ai_analysis=true")
        if meta_info.get("use_preset"):
            meta_parts.append("use_preset=true")
        if meta_info.get("preset_signals"):
            meta_parts.append(f"preset_signals={meta_info['preset_signals']}")
        if meta_info.get("preset_scene"):
            meta_parts.append(f"preset_scene={meta_info['preset_scene']}")
        if meta_info.get("preset_scene_runtime"):
            meta_parts.append(f"preset_scene_runtime={meta_info['preset_scene_runtime']}")
        if meta_info.get("record"):
            meta_parts.append("record=true")
 
        if meta_parts:
            dsl_lines.append("META: " + " ".join(meta_parts))
 
        dsl_lines.append("")
        dsl_lines.append("[SET]")
        for step in data["set_steps"]:
            dsl_lines.append(f"{step['id']}: {step['content']}")
 
        dsl_lines.append("")
        dsl_lines.append("[CHECK]")
        for step in data["check_steps"]:
            dsl_lines.append(f"{step['id']}: {step['content']}")
 
        return "\n".join(dsl_lines)

    def from_dsl(self, dsl_content: str) -> None:
        data = {"case_info": {"name": "", "id": ""}, "meta_info": {}, "set_steps": [], "check_steps": []}
        lines = (dsl_content or "").split("\n")
        current_section: Optional[str] = None

        for raw in lines:
            line = raw.strip()
            if not line:
                continue

            if line.startswith("CASE:"):
                data["case_info"]["name"] = line[5:].strip()
                continue

            if line.startswith("META:"):
                meta_str = line[5:].strip()
            
                # 按 key= 边界提取，支持 value 中含空格/逗号
                key_matches = list(re.finditer(r'(\w+)=', meta_str))
                for i, m in enumerate(key_matches):
                    k = m.group(1)
                    v_start = m.end()
                    v_end = key_matches[i + 1].start() if i + 1 < len(key_matches) else len(meta_str)
                    v = meta_str[v_start:v_end].strip()
            
                    # 兼容历史脏格式：去尾逗号、去外层引号
                    v = v.rstrip(",").strip()
                    if len(v) >= 2 and v[0] == v[-1] and v[0] in ('"', "'"):
                        v = v[1:-1].strip()
            
                    if k in ("ai_analysis", "use_preset", "record"):
                        data["meta_info"][k] = v.lower() in ("true", "1", "yes")
                    elif k in ("preset_signals", "preset_scene", "preset_scene_runtime"):
                        data["meta_info"][k] = v
                    elif k == "case_id":
                        # case_id 属于 case_info，不属于 meta_info
                        data["case_info"]["id"] = v
                    else:
                        data["meta_info"][k] = v
                continue

            if line == "[SET]":
                current_section = "SET"
                continue

            if line == "[CHECK]":
                current_section = "CHECK"
                continue

            m = re.match(r"^([SC]\d+):\s*(.*)$", line)
            if not m:
                continue
            step_id = m.group(1)
            step_content = m.group(2).strip()

            if current_section == "SET":
                data["set_steps"].append({"id": step_id, "content": step_content})
            elif current_section == "CHECK":
                data["check_steps"].append({"id": step_id, "content": step_content})

        self.set_case_data(data)

    def validate(self) -> List[str]:
        errors: List[str] = []
        data = self.get_case_data()

        # CASE信息验证
        if not data["case_info"]["name"]:
            errors.append("CASE名称不能为空")
        if not data["case_info"].get("id"):
            errors.append("CASE ID不能为空")

        # META信息验证
        if not data["meta_info"].get("test_point"):
            errors.append("测试点不能为空")
        if not data["meta_info"].get("scenario_id"):
            errors.append("场景ID不能为空")
        if not data["meta_info"].get("scenario_name"):
            errors.append("场景名称不能为空")
        # 负责人不填的话默认为Auto（这里只是验证，不修改数据）
        # 如果需要设置默认值，应该在保存时处理

        # SET和CHECK步骤数量验证
        if not data["set_steps"]:
            errors.append("至少需要一个SET步骤")
        if not data["check_steps"]:
            errors.append("至少需要一个CHECK步骤")

        # SET步骤内容验证
        for step in data["set_steps"]:
            step_id = step.get("id", "")
            content = step.get("content", "").strip()
            if not content:
                errors.append(f"SET步骤 {step_id} 的内容不能为空")
                continue

            # 解析SET步骤，验证信号名和信号值
            model, ok, msg = parse_set_step(content)
            if not ok:
                errors.append(f"SET步骤 {step_id} 格式错误: {msg}")
                continue

            # 验证每个signal的信号名和信号值
            for i, signal in enumerate(model.signals):
                signal_name = signal.name.strip()
                signal_value = signal.value.strip()
                if not signal_name:
                    errors.append(f"SET步骤 {step_id} 的第{i+1}个信号名不能为空")
                if not signal_value:
                    errors.append(f"SET步骤 {step_id} 的第{i+1}个信号值不能为空")

        # CHECK步骤内容验证
        for step in data["check_steps"]:
            step_id = step.get("id", "")
            content = step.get("content", "").strip()
            if not content:
                errors.append(f"CHECK步骤 {step_id} 的内容不能为空")
                continue

            # 解析CHECK步骤，验证信号名和信号值
            model, ok, msg = parse_check_step(content)
            if not ok:
                errors.append(f"CHECK步骤 {step_id} 格式错误: {msg}")
                continue

            # 验证每个检查项的信号名和信号值
            for i, item in enumerate(model.items):
                signal_name = item.name.strip()
                if not signal_name:
                    errors.append(f"CHECK步骤 {step_id} 的第{i+1}个信号名不能为空")

                # 根据值模式验证信号值
                if item.mode == "single":
                    if not item.single_value.strip():
                        errors.append(f"CHECK步骤 {step_id} 的第{i+1}个信号值（单值模式）不能为空")
                elif item.mode == "list":
                    if not item.list_values or all(not v.strip() for v in item.list_values):
                        errors.append(f"CHECK步骤 {step_id} 的第{i+1}个信号值（列表模式）不能为空")
                elif item.mode == "range":
                    if not item.range_a.strip() or not item.range_b.strip():
                        errors.append(f"CHECK步骤 {step_id} 的第{i+1}个信号值（范围模式）不能为空，需要填写范围a和b")

        # then check 引用校验（支持 C1 或 C1,C2）
        check_step_ids = {step["id"] for step in data["check_steps"]}
        pat = re.compile(r"then\s+check\s+([C]\d+(?:\s*,\s*[C]\d+)*)", re.IGNORECASE)

        for step in data["set_steps"]:
            m = pat.search(step.get("content") or "")
            if not m:
                continue
            ids = [x.strip() for x in m.group(1).split(",") if x.strip()]
            for cid in ids:
                if cid not in check_step_ids:
                    errors.append(f"SET步骤 {step['id']} 引用的检查步骤 {cid} 不存在")

        return errors
    
    def _update_flowchart_button_state(self):
        """更新流程图按钮状态"""
        set_steps = self.set_module_widget.get_steps()
        check_steps = self.check_module_widget.get_steps()
        
        # 检查是否有有效的SET和CHECK步骤
        has_valid_set = any(step.get("content", "").strip() for step in set_steps)
        has_valid_check = any(step.get("content", "").strip() for step in check_steps)
        
        # 只有当SET和CHECK都有有效内容时才启用按钮
        self.view_flowchart_btn.setEnabled(has_valid_set and has_valid_check)
    
    def _on_view_flowchart(self):
        """查看流程配置按钮点击事件"""
        set_steps = self.set_module_widget.get_steps()
        check_steps = self.check_module_widget.get_steps()
        
        # 过滤掉空步骤
        valid_set_steps = [step for step in set_steps if step.get("content", "").strip()]
        valid_check_steps = [step for step in check_steps if step.get("content", "").strip()]
        
        # 显示流程图对话框
        dlg = FlowchartViewDialog(valid_set_steps, valid_check_steps, self, editor=self, save_callback=self._save_to_file_callback)
        # 连接保存请求信号到内容变更信号
        dlg.save_requested.connect(self._emit_content_changed)
        dlg.exec()
    
    def _save_to_file_callback(self):
        """保存到文件的回调函数"""
        # 触发保存到文件请求信号
        self.save_to_file_requested.emit()


# ----------------------------
# 流程图节点类
# ----------------------------
class FlowchartNode(QGraphicsItem):
    """流程图节点基类"""
    
    def __init__(self, node_type: str, node_id: str, data: Dict[str, Any], parent=None):
        super().__init__(parent)
        self.node_type = node_type  # "start", "end", "set", "check"
        self.node_id = node_id
        self.data = data
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
        
        # 节点尺寸
        self.width = 120
        self.height = 60
        if node_type in ("start", "end"):
            self.width = 80
            self.height = 80
        
        # 颜色配置
        self.colors = {
            "start": QColor("#4CAF50"),      # 绿色
            "end": QColor("#f44336"),        # 红色
            "set": QColor("#2196F3"),        # 蓝色
            "check": QColor("#FF9800"),      # 橙色
        }
        
        # 文本项
        self.text_item = QGraphicsTextItem(self)
        self._update_text()
    
    def _update_text(self):
        """更新节点文本"""
        if self.node_type == "start":
            text = "开始"
        elif self.node_type == "end":
            text = "结束"
        elif self.node_type == "set":
            text = f"SET Signal"
        elif self.node_type == "check":
            text = f"CHECK Signal"
        else:
            text = self.node_id
        
        self.text_item.setPlainText(text)
        
        # 设置文本样式
        font = QFont("Arial", 10)
        font.setBold(True)
        self.text_item.setFont(font)
        
        # 居中文本
        text_rect = self.text_item.boundingRect()
        self.text_item.setPos(
            (self.width - text_rect.width()) / 2,
            (self.height - text_rect.height()) / 2
        )
    
    def boundingRect(self) -> QRectF:
        """返回节点的边界矩形"""
        return QRectF(0, 0, self.width, self.height)
    
    def paint(self, painter: QPainter, option, widget=None):
        """绘制节点"""
        # 设置抗锯齿
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # 设置画笔和画刷
        color = self.colors.get(self.node_type, QColor("#999999"))
        
        if self.isSelected():
            # 选中状态：加粗边框
            pen = QPen(QColor("#FFD700"), 3)  # 金色边框
        else:
            pen = QPen(QColor("#333333"), 2)
        
        painter.setPen(pen)
        painter.setBrush(QBrush(color))
        
        # 根据节点类型绘制不同形状
        if self.node_type in ("start", "end"):
            # 圆形/椭圆
            painter.drawEllipse(0, 0, self.width, self.height)
        else:
            # 圆角矩形
            painter.drawRoundedRect(0, 0, self.width, self.height, 10, 10)
    
    def get_properties(self) -> Dict[str, str]:
        """获取节点属性"""
        props = {
            "节点ID": self.node_id,
            "节点类型": self.node_type.upper(),
        }
        
        if self.node_type == "set":
            content = self.data.get("content", "")
            props["内容"] = content
            # 解析wait和then check
            if "wait" in content.lower():
                import re
                m = re.search(r"wait\s+(\d+)\s*(ms|s)", content, re.IGNORECASE)
                if m:
                    props["等待时间"] = f"{m.group(1)}{m.group(2)}"
            if "then check" in content.lower():
                import re
                m = re.search(r"then\s+check\s+([^\s]+(?:\s*,\s*[^\s]+)*)", content, re.IGNORECASE)
                if m:
                    props["后续检查"] = m.group(1)
            # 解析comment
            m = re.search(r'comment\s+"([^"]*)"', content, re.IGNORECASE)
            if m:
                props["注释"] = m.group(1)
        
        elif self.node_type == "check":
            content = self.data.get("content", "")
            props["内容"] = content
            # 解析timeout和async
            import re
            m = re.search(r"timeout\s+(\d+)\s*(ms|s)", content, re.IGNORECASE)
            if m:
                props["超时时间"] = f"{m.group(1)}{m.group(2)}"
            m = re.search(r"async\s+(true|false)", content, re.IGNORECASE)
            if m:
                props["异步模式"] = "是" if m.group(1).lower() == "true" else "否"
            # 解析comment
            m = re.search(r'comment\s+"([^"]*)"', content, re.IGNORECASE)
            if m:
                props["注释"] = m.group(1)
        
        return props


class FlowchartEdge(QGraphicsPathItem):
    """流程图边（连接线）"""
    
    def __init__(self, source_node: FlowchartNode, target_node: FlowchartNode, parent=None):
        super().__init__(parent)
        self.source_node = source_node
        self.target_node = target_node
        self.setZValue(-1)  # 放在节点下方
        
        # 设置画笔
        pen = QPen(QColor("#666666"), 2)
        pen.setStyle(Qt.PenStyle.SolidLine)
        self.setPen(pen)
        
        self._update_path()
    
    def _update_path(self):
        """更新连接线路径：同列直连，跨列回折"""
        if not self.source_node or not self.target_node:
            return
    
        sp = self.source_node.pos()
        tp = self.target_node.pos()
    
        source_bottom = QPointF(sp.x() + self.source_node.width / 2, sp.y() + self.source_node.height)
        target_top = QPointF(tp.x() + self.target_node.width / 2, tp.y())
    
        source_right_mid = QPointF(sp.x() + self.source_node.width, sp.y() + self.source_node.height / 2)
        target_left_mid = QPointF(tp.x(), tp.y() + self.target_node.height / 2)
    
        source_center_x = sp.x() + self.source_node.width / 2
        target_center_x = tp.x() + self.target_node.width / 2
        same_col = abs(source_center_x - target_center_x) < 1e-3
    
        path = QPainterPath()
    
        if same_col:
            # 同列：从下到上直连（中间微弯，视觉更平滑）
            mid_y = (source_bottom.y() + target_top.y()) / 2
            path.moveTo(source_bottom)
            path.lineTo(source_bottom.x(), mid_y)
            path.lineTo(target_top.x(), mid_y)
            path.lineTo(target_top)
        else:
            # 跨列：回折线（右移 -> 竖移 -> 接入目标左侧）
            lane_x = (source_right_mid.x() + target_left_mid.x()) / 2.0
            path.moveTo(source_bottom)
            path.lineTo(source_right_mid)
            path.lineTo(lane_x, source_right_mid.y())
            path.lineTo(lane_x, target_left_mid.y())
            path.lineTo(target_left_mid)
    
        self.setPath(path)

class _AsyncBoolDelegate(QStyledItemDelegate):
    def __init__(self, async_row_getter, parent=None):
        super().__init__(parent)
        self._async_row_getter = async_row_getter
 
    def _is_async_cell(self, index) -> bool:
        r = self._async_row_getter()
        return r is not None and index.row() == r and index.column() == 1
 
    def createEditor(self, parent, option, index):
        if self._is_async_cell(index):
            combo = QComboBox(parent)
            combo.addItems(["false", "true"])
            # 只在用户最终选择后提交，避免 currentTextChanged 引发重绘重影
            combo.activated.connect(lambda _i, c=combo: self._commit_and_close(c))
            return combo
        return super().createEditor(parent, option, index)
 
    def _commit_and_close(self, editor: QComboBox):
        self.commitData.emit(editor)
        self.closeEditor.emit(editor, QAbstractItemDelegate.EndEditHint.NoHint)
 
    def setEditorData(self, editor, index):
        if self._is_async_cell(index) and isinstance(editor, QComboBox):
            v = str(index.data(Qt.ItemDataRole.EditRole) or "false").strip().lower()
            editor.setCurrentText("true" if v == "true" else "false")
            return
        super().setEditorData(editor, index)
 
    def setModelData(self, editor, model, index):
        if self._is_async_cell(index) and isinstance(editor, QComboBox):
            model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)
            return
        super().setModelData(editor, model, index)
 
    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)
 
 
class _SetSignalInlineRow(QWidget):
    add_requested = pyqtSignal(object)
    delete_requested = pyqtSignal(object)
    changed = pyqtSignal()

    def __init__(
        self,
        completions_by_kind: Dict[str, List[str]],
        hier_index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        dbc_parser=None,
        parent=None,
    ):
        super().__init__(parent)
        self._dbc_parser = dbc_parser
        self._hier_index_by_kind = hier_index_by_kind
        self._build_ui()

    def _build_ui(self) -> None:
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(4)

        self.kind_combo = QComboBox(self)
        self.kind_combo.addItems(["env", "sys"])
        self.kind_combo.setFixedWidth(60)
        lay.addWidget(self.kind_combo)  # index 0

        self.name_edit = QLineEdit(self)
        self.name_edit.setPlaceholderText("信号名称")
        self.name_edit.setMinimumWidth(260)
        lay.addWidget(self.name_edit, 1)  # index 1

        # 使用 _SignalValueHelper 替代 value_edit
        self._value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=lay,
            insert_index=2,  # 在 name_edit 之后
            dbc_parser=self._dbc_parser,
            value_placeholder="值",
            value_width=60,
            info_width=80,
        )

        self.btn_add = QPushButton("+", self)
        self.btn_add.setFixedWidth(28)
        lay.addWidget(self.btn_add)  # index 5

        self.btn_del = QPushButton("-", self)
        self.btn_del.setFixedWidth(28)
        lay.addWidget(self.btn_del)  # index 6

        self._set_uniform_height()

        self._hier = _HierLineEditCompleter(
            edit=self.name_edit,
            kind_getter=lambda: self.kind_combo.currentText(),
            index_by_kind=self._hier_index_by_kind,
            allowed_kinds=["env", "sys"],
            dbc_parser=self._dbc_parser,
            on_completed=self._value_helper.update_from_signal_name,  # 补全完成时更新值控件
        )
        self.kind_combo.currentTextChanged.connect(lambda _k: self._hier.refresh())

        self.btn_add.clicked.connect(lambda: self.add_requested.emit(self))
        self.btn_del.clicked.connect(lambda: self.delete_requested.emit(self))

        self.kind_combo.currentTextChanged.connect(lambda _: self.changed.emit())
        self.name_edit.textChanged.connect(lambda _: self.changed.emit())

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._value_helper.update_from_signal_name)
        self.kind_combo.currentTextChanged.connect(lambda _: self._value_helper.update_from_signal_name())

    def _set_uniform_height(self) -> None:
        h = 26
        for w in (self.kind_combo, self.name_edit, self.btn_add, self.btn_del):
            w.setFixedHeight(h)
        # value_helper 的控件也需要设置高度
        if self._value_helper._value_edit:
            self._value_helper._value_edit.setFixedHeight(h)
        if self._value_helper._value_combo:
            self._value_helper._value_combo.setFixedHeight(h)

    def set_data(self, s: SetSignalModel) -> None:
        kind = s.kind if s.kind in ("env", "sys") else "sys"
        self.kind_combo.setCurrentText(kind)

        name = (s.name or "").strip()
        if name.startswith("sys::"):
            name = name[5:]
        elif name.startswith("env::"):
            name = name[5:]
        self.name_edit.setText(name)
        self._value_helper.set_value(s.value or "")
        self._hier.refresh()
        # 触发更新值控件
        self._value_helper.update_from_signal_name()

    def get_data(self) -> SetSignalModel:
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()
        if name and not name.startswith("sys::") and not name.startswith("env::"):
            name = f"{kind}::{name}"
        return SetSignalModel(kind=kind, name=name, value=self._value_helper.get_value())
 
 
class _SetSignalsInlineEditor(QWidget):
    changed = pyqtSignal()
 
    def __init__(
        self,
        completions_by_kind: Dict[str, List[str]],
        hier_index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        dbc_parser=None,
        parent=None,
    ):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind
        self._hier_index_by_kind = hier_index_by_kind
        self._dbc_parser = dbc_parser
        self._rows: List[_SetSignalInlineRow] = []
        self._build_ui()
 
    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
    
        self.scroll = QScrollArea(self)
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.StyledPanel)
        self.scroll.setStyleSheet("QScrollArea { background: #ffffff; border: 1px solid #e0e0e0; }")
        self.scroll.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.scroll.setFixedHeight(112)
    
        self.container = QWidget(self.scroll)
        self.rows_layout = QVBoxLayout(self.container)
        self.rows_layout.setContentsMargins(4, 4, 4, 4)
        self.rows_layout.setSpacing(2)
        self.rows_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
    
        self.scroll.setWidget(self.container)
        root.addWidget(self.scroll)
    
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setFixedHeight(124)
        self.scroll.setFixedHeight(118)
 
    def _add_row(self, data: Optional[SetSignalModel] = None, after_row: Optional[_SetSignalInlineRow] = None) -> None:
        row = _SetSignalInlineRow(
            completions_by_kind=self._completions_by_kind,
            hier_index_by_kind=self._hier_index_by_kind,
            dbc_parser=self._dbc_parser,
            parent=self.container,
        )
        row.add_requested.connect(self._on_add_requested)
        row.delete_requested.connect(self._on_delete_requested)
        row.changed.connect(self.changed.emit)
 
        if data is not None:
            row.set_data(data)
 
        if after_row is not None and after_row in self._rows:
            idx = self._rows.index(after_row) + 1
            self._rows.insert(idx, row)
            self.rows_layout.insertWidget(idx, row)
        else:
            self._rows.append(row)
            self.rows_layout.addWidget(row)
 
    def _on_add_requested(self, row: _SetSignalInlineRow) -> None:
        self._add_row(SetSignalModel(kind="sys", name="", value=""), after_row=row)
        self.changed.emit()
 
    def _on_delete_requested(self, row: _SetSignalInlineRow) -> None:
        if row not in self._rows:
            return
        if len(self._rows) <= 1:
            row.set_data(SetSignalModel(kind="sys", name="", value=""))
            self.changed.emit()
            return
        self._rows.remove(row)
        self.rows_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()
        self.changed.emit()
 
    def set_signals(self, signals: List[SetSignalModel]) -> None:
        for r in self._rows:
            self.rows_layout.removeWidget(r)
            r.setParent(None)
            r.deleteLater()
        self._rows.clear()
 
        if not signals:
            signals = [SetSignalModel(kind="sys", name="", value="")]
 
        for s in signals:
            self._add_row(s)
 
    def get_signals(self) -> List[SetSignalModel]:
        out = [r.get_data() for r in self._rows]
        return out if out else [SetSignalModel(kind="sys", name="", value="")]
 
 
class _CheckSignalInlineRow(QWidget):
    add_requested = pyqtSignal(object)
    delete_requested = pyqtSignal(object)
    changed = pyqtSignal()

    def __init__(
        self,
        completions_by_kind: Dict[str, List[str]],
        hier_index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        dbc_parser=None,
        parent=None,
    ):
        super().__init__(parent)
        self._dbc_parser = dbc_parser
        self._hier_index_by_kind = hier_index_by_kind
        self._wait_ms_cache = 0
        self._build_ui()

    def _build_ui(self) -> None:
        self._layout = QHBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(4)

        self.kind_combo = QComboBox(self)
        self.kind_combo.addItems(["sig", "env", "sys"])
        self.kind_combo.setFixedWidth(60)
        self._layout.addWidget(self.kind_combo)  # index 0

        self.name_edit = QLineEdit(self)
        self.name_edit.setPlaceholderText("信号名称")
        self.name_edit.setMinimumWidth(260)
        self._layout.addWidget(self.name_edit, 1)  # index 1

        self.mode_combo = QComboBox(self)
        self.mode_combo.addItems(["single", "list", "range"])
        self.mode_combo.setFixedWidth(60)
        self._layout.addWidget(self.mode_combo)  # index 2

        self.op_combo = QComboBox(self)
        self.op_combo.addItems(["==", ">", "<", ">=", "<=", "!="])
        self.op_combo.setFixedWidth(60)
        self._layout.addWidget(self.op_combo)  # index 3

        # 使用 _SignalValueHelper 替代 single_edit
        self._single_value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=self._layout,
            insert_index=4,  # 在 op_combo 之后
            dbc_parser=self._dbc_parser,
            value_placeholder="值",
            value_width=60,
            info_width=80,
        )

        self.list_edit = QLineEdit(self)
        self.list_edit.setPlaceholderText("1,2,3")
        self.list_edit.setFixedWidth(80)
        self._layout.addWidget(self.list_edit)
        self.list_edit.hide()

        self.range_a_edit = QLineEdit(self)
        self.range_a_edit.setPlaceholderText("a")
        self.range_a_edit.setFixedWidth(40)
        self._layout.addWidget(self.range_a_edit)
        self.range_a_edit.hide()

        self.range_sep = QLabel("..", self)
        self._layout.addWidget(self.range_sep)
        self.range_sep.hide()

        self.range_b_edit = QLineEdit(self)
        self.range_b_edit.setPlaceholderText("b")
        self.range_b_edit.setFixedWidth(40)
        self._layout.addWidget(self.range_b_edit)
        self.range_b_edit.hide()

        # 为 list/range 模式添加范围信息标签
        self._list_range_info_label = QLabel()
        self._list_range_info_label.setStyleSheet("color: #666666; font-size: 11px;")
        self._list_range_info_label.setMinimumWidth(80)
        self._layout.addWidget(self._list_range_info_label)
        self._list_range_info_label.hide()

        self.btn_add = QPushButton("+", self)
        self.btn_add.setFixedWidth(28)
        self._layout.addWidget(self.btn_add)

        self.btn_del = QPushButton("-", self)
        self.btn_del.setFixedWidth(28)
        self._layout.addWidget(self.btn_del)

        self._set_uniform_height()

        self._hier = _HierLineEditCompleter(
            edit=self.name_edit,
            kind_getter=lambda: self.kind_combo.currentText(),
            index_by_kind=self._hier_index_by_kind,
            allowed_kinds=["sig", "env", "sys"],
            dbc_parser=self._dbc_parser,
            on_completed=self._update_value_controls,  # 补全完成时更新值控件
        )
        self.kind_combo.currentTextChanged.connect(lambda _k: self._hier.refresh())

        self.mode_combo.currentTextChanged.connect(self._on_mode_changed)
        self.btn_add.clicked.connect(lambda: self.add_requested.emit(self))
        self.btn_del.clicked.connect(lambda: self.delete_requested.emit(self))

        self.kind_combo.currentTextChanged.connect(lambda _: self.changed.emit())
        self.name_edit.textChanged.connect(lambda _: self.changed.emit())
        self.mode_combo.currentTextChanged.connect(lambda _: self.changed.emit())
        self.op_combo.currentTextChanged.connect(lambda _: self.changed.emit())
        self.list_edit.textChanged.connect(lambda _: self.changed.emit())
        self.range_a_edit.textChanged.connect(lambda _: self.changed.emit())
        self.range_b_edit.textChanged.connect(lambda _: self.changed.emit())

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._update_value_controls)
        self.kind_combo.currentTextChanged.connect(lambda _: self._update_value_controls())

        self._on_mode_changed(self.mode_combo.currentText())

    def _update_value_controls(self) -> None:
        """根据信号名和模式更新值控件"""
        mode = self.mode_combo.currentText()

        # 先更新 single 模式的 helper
        self._single_value_helper.update_from_signal_name()

        # 对于 list/range 模式，更新范围信息标签
        if mode in ("list", "range"):
            self._update_list_range_info()

    def _update_list_range_info(self) -> None:
        """为 list/range 模式更新范围信息标签"""
        kind = self.kind_combo.currentText()
        name = self.name_edit.text()

        if kind not in ("sig", "env") or not self._dbc_parser or not name:
            self._list_range_info_label.setText("")
            self._list_range_info_label.hide()
            return

        # 构建完整路径
        full_path = name
        if not name.startswith("CAN "):
            if name.startswith("sig::") or name.startswith("env::"):
                full_path = name
            else:
                full_path = f"CAN {name}"

        info = self._dbc_parser.get_signal_info(full_path, kind)
        if info is None:
            self._list_range_info_label.setText("")
            self._list_range_info_label.hide()
            return

        choices = info.get("choices")
        minimum = info.get("minimum")
        maximum = info.get("maximum")

        if choices:
            vals = list(choices.keys())[:5]
            self._list_range_info_label.setText(f"可选: {', '.join(str(v) for v in vals)}...")
            self._list_range_info_label.show()
        elif minimum is not None and maximum is not None:
            self._list_range_info_label.setText(f"[{minimum} ~ {maximum}]")
            self._list_range_info_label.show()
        else:
            self._list_range_info_label.hide()

    def _set_uniform_height(self) -> None:
        h = 26
        for w in (
            self.kind_combo, self.name_edit, self.mode_combo, self.op_combo,
            self.list_edit, self.range_a_edit, self.range_b_edit,
            self.btn_add, self.btn_del
        ):
            w.setFixedHeight(h)
        # value_helper 的控件也需要设置高度
        if self._single_value_helper._value_edit:
            self._single_value_helper._value_edit.setFixedHeight(h)
        if self._single_value_helper._value_combo:
            self._single_value_helper._value_combo.setFixedHeight(h)

    def _on_mode_changed(self, mode: str) -> None:
        mode = (mode or "single").lower()
        is_single = mode == "single"
        is_list = mode == "list"
        is_range = mode == "range"

        self.op_combo.setVisible(is_single)
        self._single_value_helper.show_widgets() if is_single else self._single_value_helper.hide_widgets()

        self.list_edit.setVisible(is_list)

        self.range_a_edit.setVisible(is_range)
        self.range_sep.setVisible(is_range)
        self.range_b_edit.setVisible(is_range)

        # 更新范围信息标签
        self._list_range_info_label.setVisible(is_list or is_range)
        if is_list or is_range:
            self._update_list_range_info()

    def set_data(self, it: CheckItemModel) -> None:
        kind = it.kind if it.kind in ("sig", "env", "sys") else "sig"
        self.kind_combo.setCurrentText(kind)

        name = (it.name or "").strip()
        for p in ("sig::", "env::", "sys::"):
            if name.startswith(p):
                name = name[len(p):]
                break
        self.name_edit.setText(name)

        self.mode_combo.setCurrentText(it.mode if it.mode in ("single", "list", "range") else "single")
        self.op_combo.setCurrentText(it.op if it.op in ("==", ">", "<", ">=", "<=", "!=") else "==")
        self._single_value_helper.set_value(it.single_value or "")
        self.list_edit.setText(",".join(it.list_values or []))
        self.range_a_edit.setText(it.range_a or "")
        self.range_b_edit.setText(it.range_b or "")
        self._wait_ms_cache = int(it.wait_ms or 0)

        self._on_mode_changed(self.mode_combo.currentText())
        self._hier.refresh()
        # 触发更新值控件
        self._update_value_controls()

    def get_data(self) -> CheckItemModel:
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()
        if name and not name.startswith("sig::") and not name.startswith("env::") and not name.startswith("sys::"):
            name = f"{kind}::{name}"

        mode = self.mode_combo.currentText()

        return CheckItemModel(
            kind=kind,
            name=name,
            mode=mode,
            op=self.op_combo.currentText(),
            single_value=self._single_value_helper.get_value(),
            list_values=[x.strip() for x in self.list_edit.text().split(",") if x.strip()],
            range_a=self.range_a_edit.text().strip(),
            range_b=self.range_b_edit.text().strip(),
            wait_ms=self._wait_ms_cache,
            timeout_ms=1000,
            duration_ms=0,
            async_=False,
        )
 
 
class _CheckSignalsInlineEditor(QWidget):
    changed = pyqtSignal()
 
    def __init__(
        self,
        completions_by_kind: Dict[str, List[str]],
        hier_index_by_kind: Dict[str, Dict[Tuple[str, ...], List[str]]],
        dbc_parser=None,
        parent=None,
    ):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind
        self._hier_index_by_kind = hier_index_by_kind
        self._dbc_parser = dbc_parser
        self._rows: List[_CheckSignalInlineRow] = []
        self._build_ui()
 
    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(4)
 
        self.scroll = QScrollArea(self)
        self.scroll.setWidgetResizable(True)
        self.scroll.setFrameShape(QFrame.Shape.StyledPanel)
        self.scroll.setStyleSheet("QScrollArea { background: #ffffff; border: 1px solid #e0e0e0; }")
 
        self.container = QWidget(self.scroll)
        self.rows_layout = QVBoxLayout(self.container)
        self.rows_layout.setContentsMargins(4, 4, 4, 4)
        self.rows_layout.setSpacing(4)
        self.rows_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
 
        self.scroll.setWidget(self.container)
        root.addWidget(self.scroll)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setFixedHeight(136)
        self.scroll.setFixedHeight(130)
 
    def _add_row(self, data: Optional[CheckItemModel] = None, after_row: Optional[_CheckSignalInlineRow] = None) -> None:
        row = _CheckSignalInlineRow(
            completions_by_kind=self._completions_by_kind,
            hier_index_by_kind=self._hier_index_by_kind,
            dbc_parser=self._dbc_parser,
            parent=self.container,
        )
        row.add_requested.connect(self._on_add_requested)
        row.delete_requested.connect(self._on_delete_requested)
        row.changed.connect(self.changed.emit)
 
        if data is not None:
            row.set_data(data)
 
        if after_row is not None and after_row in self._rows:
            idx = self._rows.index(after_row) + 1
            self._rows.insert(idx, row)
            self.rows_layout.insertWidget(idx, row)
        else:
            self._rows.append(row)
            self.rows_layout.addWidget(row)
 
    def _on_add_requested(self, row: _CheckSignalInlineRow) -> None:
        self._add_row(CheckItemModel(kind="sig", name="", mode="single"), after_row=row)
        self.changed.emit()
 
    def _on_delete_requested(self, row: _CheckSignalInlineRow) -> None:
        if row not in self._rows:
            return
        if len(self._rows) <= 1:
            row.set_data(CheckItemModel(kind="sig", name="", mode="single"))
            self.changed.emit()
            return
        self._rows.remove(row)
        self.rows_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()
        self.changed.emit()
 
    def set_items(self, items: List[CheckItemModel]) -> None:
        for r in self._rows:
            self.rows_layout.removeWidget(r)
            r.setParent(None)
            r.deleteLater()
        self._rows.clear()
 
        if not items:
            items = [CheckItemModel(kind="sig", name="", mode="single")]
 
        for it in items:
            self._add_row(it)
 
    def get_items(self) -> List[CheckItemModel]:
        out = [r.get_data() for r in self._rows]
        return out if out else [CheckItemModel(kind="sig", name="", mode="single")]
 
 
class FlowchartViewDialog(QDialog):
    """流程图查看对话框"""
    save_requested = pyqtSignal()
 
    def __init__(self, set_steps: List[Dict[str, Any]], check_steps: List[Dict[str, Any]], parent=None, editor=None, save_callback=None):
        super().__init__(parent)
        self.setWindowTitle("查看流程配置")
        self.resize(1200, 800)
 
        self.set_steps = set_steps
        self.check_steps = check_steps
        self.editor = editor
        self.save_callback = save_callback
 
        self.nodes: List[FlowchartNode] = []
        self.edges: List[FlowchartEdge] = []
        self.selected_node: Optional[FlowchartNode] = None
 
        self.property_widgets: Dict[str, Any] = {}
        self.original_values: Dict[str, str] = {}
        self._async_row: Optional[int] = None
 
        all_completions = list(getattr(self.editor, "completions", []) or []) if self.editor is not None else []
        self._completions_by_kind = {
            "sys": [c for c in all_completions if isinstance(c, str) and c.startswith("sys::")],
            "env": [c for c in all_completions if isinstance(c, str) and c.startswith("env::")],
            "sig": [c for c in all_completions if isinstance(c, str) and c.startswith("sig::")],
        }
        self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)
        self._dbc_parser = getattr(self.editor, "_dbc_parser", None)
 
        self._build_ui()
        self._build_flowchart()
 
    def _build_ui(self):
        splitter = QSplitter(Qt.Orientation.Horizontal, self)
        splitter.setHandleWidth(5)
        splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #e0e0e0;
            }
            QSplitter::handle:hover {
                background-color: #2196F3;
            }
        """)
 
        # 左侧流程图
        flowchart_group = QGroupBox("流程图")
        _setup_groupbox_style(flowchart_group)
        flowchart_layout = QVBoxLayout(flowchart_group)
        flowchart_layout.setContentsMargins(0, 0, 0, 0)
        flowchart_layout.setSpacing(0)
 
        self.scene = QGraphicsScene(self)
        self.scene.setBackgroundBrush(QBrush(QColor("#ffffff")))
        
        self.view = QGraphicsView(self.scene)
        self.view.setRenderHint(QPainter.RenderHint.Antialiasing)
        self.view.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        self.view.setFrameShape(QFrame.Shape.NoFrame)
        self.view.setFrameShadow(QFrame.Shadow.Plain)
        self.view.setLineWidth(0)
        self.view.setMidLineWidth(0)
        self.view.setStyleSheet("""
        QGraphicsView {
            border: 0px;
            background: #ffffff;
        }
        QGraphicsView::viewport {
            border: 0px;
            background: #ffffff;
        }
        """)
        self.view.setBackgroundBrush(QBrush(QColor("#ffffff")))
        flowchart_layout.addWidget(self.view)
 
        splitter.addWidget(flowchart_group)
 
        # 右侧属性
        property_group = QGroupBox("节点属性")
        _setup_groupbox_style(property_group)
        property_layout = QVBoxLayout(property_group)
        property_layout.setContentsMargins(10, 10, 10, 10)
        property_layout.setSpacing(10)
 
        self.property_scroll = QScrollArea()
        self.property_scroll.setWidgetResizable(True)
        self.property_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.property_scroll.setStyleSheet("QScrollArea { background-color: #ffffff; border: none; }")
 
        self.property_container = QWidget()
        self.property_layout = QVBoxLayout(self.property_container)
        self.property_layout.setContentsMargins(0, 0, 0, 0)
        self.property_layout.setSpacing(3)
 
        self.property_scroll.setWidget(self.property_container)
        property_layout.addWidget(self.property_scroll, 1)
 
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
 
        self.save_btn = QPushButton("保存")
        self.save_btn.setEnabled(False)
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.save_btn.clicked.connect(self._on_save)
        button_layout.addWidget(self.save_btn)
 
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #f5f5f5;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 4px;
                padding: 6px 16px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
 
        property_layout.addLayout(button_layout)
        splitter.addWidget(property_group)
 
        splitter.setStretchFactor(0, 7)
        splitter.setStretchFactor(1, 3)
 
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.addWidget(splitter)
 
        self.scene.selectionChanged.connect(self._on_selection_changed)
 
    def _build_flowchart(self):
        self.scene.clear()
        self.nodes.clear()
        self.edges.clear()
 
        if not self.set_steps and not self.check_steps:
            text_item = QGraphicsTextItem("暂无流程数据")
            text_item.setFont(QFont("Arial", 14))
            text_item.setDefaultTextColor(QColor("#999999"))
            self.scene.addItem(text_item)
            return
 
        max_steps_per_col = 6
        col_spacing = 240
        row_spacing = 95
 
        top_margin = 20
        start_gap = 45
        end_gap = 45
 
        single_col_center_x = 300
        multi_col_left_x = 130
 
        execution_nodes: List[FlowchartNode] = []
        referenced_checks = set()
        check_by_id = {s["id"]: s for s in self.check_steps}
 
        start_node = FlowchartNode("start", "START", {})
        execution_nodes.append(start_node)
 
        for set_step in self.set_steps:
            execution_nodes.append(FlowchartNode("set", set_step["id"], set_step))
 
            content = set_step.get("content", "")
            m = re.search(r"then\s+check\s+([^\s]+(?:\s*,\s*[^\s]+)*)", content, re.IGNORECASE)
            if m:
                check_ids = [x.strip() for x in m.group(1).split(",") if x.strip()]
                for cid in check_ids:
                    chk = check_by_id.get(cid)
                    if chk is not None:
                        execution_nodes.append(FlowchartNode("check", cid, chk))
                        referenced_checks.add(cid)
 
        for chk in self.check_steps:
            if chk["id"] not in referenced_checks:
                execution_nodes.append(FlowchartNode("check", chk["id"], chk))
 
        end_node = FlowchartNode("end", "END", {})
        execution_nodes.append(end_node)
 
        step_nodes = [n for n in execution_nodes if n.node_type in ("set", "check")]
        step_count = len(step_nodes)
 
        if step_count == 0:
            start_node.setPos(single_col_center_x - start_node.width / 2, top_margin)
            end_node.setPos(
                single_col_center_x - end_node.width / 2,
                top_margin + start_node.height + start_gap
            )
        else:
            col_count = (step_count + max_steps_per_col - 1) // max_steps_per_col
            col_x0 = (single_col_center_x - 60) if col_count == 1 else multi_col_left_x
 
            start_node.setPos(col_x0 + (120 - start_node.width) / 2, top_margin)
            first_step_y = top_margin + start_node.height + start_gap
 
            for idx, node in enumerate(step_nodes):
                col = idx // max_steps_per_col
                row = idx % max_steps_per_col
                x = col_x0 + col * col_spacing
                y = first_step_y + row * row_spacing
                node.setPos(x, y)
 
            last_col = (step_count - 1) // max_steps_per_col
            last_col_nodes = [n for i, n in enumerate(step_nodes) if (i // max_steps_per_col) == last_col]
            last_col_bottom = max(n.pos().y() + n.height for n in last_col_nodes)
            last_step = step_nodes[-1]
 
            end_node.setPos(
                last_step.pos().x() + (last_step.width - end_node.width) / 2,
                last_col_bottom + end_gap
            )
 
        for n in execution_nodes:
            self.scene.addItem(n)
            self.nodes.append(n)
 
        for i in range(len(execution_nodes) - 1):
            edge = FlowchartEdge(execution_nodes[i], execution_nodes[i + 1])
            self.scene.addItem(edge)
            self.edges.append(edge)
 
        bounds = self.scene.itemsBoundingRect()
        self.scene.setSceneRect(bounds.adjusted(-80, -40, 100, 100))
 
    @staticmethod
    def _to_int(text: str, default: int) -> int:
        try:
            return int(str(text).strip())
        except Exception:
            return default
 
    @staticmethod
    def _to_bool(text: str, default: bool = False) -> bool:
        t = str(text or "").strip().lower()
        if not t:
            return default
        return t in ("true", "1", "yes", "y", "on")
 
    @staticmethod
    def _serialize_set_signals(signals: List[SetSignalModel]) -> str:
        return json.dumps(
            [{"kind": s.kind, "name": s.name, "value": s.value} for s in (signals or [])],
            ensure_ascii=False,
            sort_keys=True,
        )
 
    @staticmethod
    def _serialize_check_items(items: List[CheckItemModel]) -> str:
        return json.dumps(
            [
                {
                    "kind": it.kind,
                    "name": it.name,
                    "mode": it.mode,
                    "op": it.op,
                    "single_value": it.single_value,
                    "list_values": it.list_values,
                    "range_a": it.range_a,
                    "range_b": it.range_b,
                    "wait_ms": it.wait_ms,
                }
                for it in (items or [])
            ],
            ensure_ascii=False,
            sort_keys=True,
        )
 
    def _make_prop_name_item(self, text: str) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        return item
 
    def _make_prop_value_item(self, text: str, editable: bool = True) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        if not editable:
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        return item
 
    def _clear_property_widgets(self):
        while self.property_layout.count():
            item = self.property_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.setParent(None)
                widget.deleteLater()
        self.property_widgets.clear()
 
    def _on_selection_changed(self):
        selected_items = self.scene.selectedItems()
        if not selected_items:
            self._clear_property_widgets()
            self.save_btn.setEnabled(False)
            self.selected_node = None
            return
 
        for item in selected_items:
            if isinstance(item, FlowchartNode):
                self.selected_node = item
                self._build_property_form(item)
                self.save_btn.setEnabled(item.node_type in ("set", "check"))
                return
 
    def _build_property_form(self, node: FlowchartNode):
        self._clear_property_widgets()
        self.original_values.clear()
        self._async_row = None
 
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["节点属性", "属性值"])
        table.horizontalHeader().setStretchLastSection(True)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setEditTriggers(
            QTableWidget.EditTrigger.CurrentChanged
            | QTableWidget.EditTrigger.SelectedClicked
            | QTableWidget.EditTrigger.EditKeyPressed
            | QTableWidget.EditTrigger.AnyKeyPressed
        )
        table.setItemDelegate(_AsyncBoolDelegate(lambda: self._async_row, table))
        table.setStyleSheet("""
            QTableWidget {
                border: none;
                background-color: #ffffff;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 5px;
                border: 1px solid #e0e0e0;
                font-weight: bold;
            }
        """)
 
        row = 0
 
        table.insertRow(row)
        table.setItem(row, 0, self._make_prop_name_item("节点ID"))
        table.setItem(row, 1, self._make_prop_value_item(node.node_id, editable=False))
        row += 1
 
        table.insertRow(row)
        table.setItem(row, 0, self._make_prop_name_item("节点类型"))
        table.setItem(row, 1, self._make_prop_value_item(node.node_type.upper(), editable=False))
        row += 1
 
        if node.node_type == "set":
            content = node.data.get("content", "")
            set_model, _, _ = parse_set_step(content)
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("信号*"))
            signal_panel = _SetSignalsInlineEditor(
                completions_by_kind=self._completions_by_kind,
                hier_index_by_kind=self._hier_index_by_kind,
                dbc_parser=self._dbc_parser,
                parent=table,
            )
            signal_panel.set_signals(set_model.signals if set_model.signals else [SetSignalModel(kind="sys", name="", value="")])
            signal_panel.changed.connect(self._on_property_changed)
            table.setCellWidget(row, 1, signal_panel)
            table.setRowHeight(row, 130)
            self.property_widgets["signal_rows"] = signal_panel
            self.original_values["signal_rows"] = self._serialize_set_signals(signal_panel.get_signals())
            row += 1
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("等待时间(ms)"))
            wait_text = str(int(set_model.wait_ms or 0)) if int(set_model.wait_ms or 0) > 0 else ""
            wait_item = self._make_prop_value_item(wait_text, editable=True)
            table.setItem(row, 1, wait_item)
            self.property_widgets["wait"] = wait_item
            self.original_values["wait"] = wait_text
            row += 1
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("后续检查"))
            next_check_text = ",".join(set_model.next_checks or [])
            next_check_item = self._make_prop_value_item(next_check_text, editable=True)
            table.setItem(row, 1, next_check_item)
            self.property_widgets["next_check"] = next_check_item
            self.original_values["next_check"] = next_check_text
            row += 1
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("注释"))
            comment_text = (set_model.comment or "").strip()
            comment_item = self._make_prop_value_item(comment_text, editable=True)
            table.setItem(row, 1, comment_item)
            self.property_widgets["comment"] = comment_item
            self.original_values["comment"] = comment_text
            row += 1
 
        elif node.node_type == "check":
            content = node.data.get("content", "")
            check_model, _, _ = parse_check_step(content)
            items = check_model.items if check_model.items else [CheckItemModel(kind="sig", name="", mode="single")]
            first = items[0]
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("信号*"))
            signal_panel = _CheckSignalsInlineEditor(
                completions_by_kind=self._completions_by_kind,
                hier_index_by_kind=self._hier_index_by_kind,
                dbc_parser=self._dbc_parser,
                parent=table,
            )
            signal_panel.set_items(items)
            signal_panel.changed.connect(self._on_property_changed)
            table.setCellWidget(row, 1, signal_panel)
            table.setRowHeight(row, 142)
            self.property_widgets["signal_rows"] = signal_panel
            self.original_values["signal_rows"] = self._serialize_check_items(signal_panel.get_items())
            row += 1
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("超时时间(ms)"))
            timeout_text = str(int(first.timeout_ms or 1000))
            timeout_item = self._make_prop_value_item(timeout_text, editable=True)
            table.setItem(row, 1, timeout_item)
            self.property_widgets["timeout"] = timeout_item
            self.original_values["timeout"] = timeout_text
            row += 1
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("持续时间(ms)"))
            duration_text = str(int(first.duration_ms or 0)) if int(first.duration_ms or 0) > 0 else ""
            duration_item = self._make_prop_value_item(duration_text, editable=True)
            table.setItem(row, 1, duration_item)
            self.property_widgets["duration"] = duration_item
            self.original_values["duration"] = duration_text
            row += 1
 
            # 异步模式：标准表格项 + delegate 下拉编辑
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("异步模式"))
            async_text = "true" if bool(first.async_) else "false"
            async_item = self._make_prop_value_item(async_text, editable=True)
            table.setItem(row, 1, async_item)
            self.property_widgets["async"] = async_item
            self.original_values["async"] = async_text
            self._async_row = row
            row += 1
 
            table.insertRow(row)
            table.setItem(row, 0, self._make_prop_name_item("注释"))
            comment_text = (check_model.comment or "").strip()
            comment_item = self._make_prop_value_item(comment_text, editable=True)
            table.setItem(row, 1, comment_item)
            self.property_widgets["comment"] = comment_item
            self.original_values["comment"] = comment_text
            row += 1
 
        # 左列全部只读保险
        for r in range(table.rowCount()):
            name_item = table.item(r, 0)
            if name_item is not None:
                name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
 
        table.itemChanged.connect(self._on_property_changed)
        self.property_layout.addWidget(table)
        self.save_btn.setEnabled(False)
 
    def _on_property_changed(self, *args):
        if not self.selected_node or self.selected_node.node_type not in ("set", "check"):
            return
 
        has_changes = False
        for key, widget in self.property_widgets.items():
            original_value = self.original_values.get(key, "")
 
            if isinstance(widget, QTableWidgetItem):
                current_value = widget.text().strip()
            elif isinstance(widget, _SetSignalsInlineEditor):
                current_value = self._serialize_set_signals(widget.get_signals())
            elif isinstance(widget, _CheckSignalsInlineEditor):
                current_value = self._serialize_check_items(widget.get_items())
            else:
                continue
 
            if current_value != original_value:
                has_changes = True
                break
 
        self.save_btn.setEnabled(has_changes)
 
    def _on_save(self):
        if not self.selected_node or self.selected_node.node_type not in ("set", "check"):
            return
 
        node = self.selected_node
        node_id = node.node_id
        node_type = node.node_type
 
        if node_type == "set":
            signal_panel = self.property_widgets.get("signal_rows")
            wait_widget = self.property_widgets.get("wait")
            next_check_widget = self.property_widgets.get("next_check")
            comment_widget = self.property_widgets.get("comment")
 
            if not isinstance(signal_panel, _SetSignalsInlineEditor):
                return
 
            wait_ms = 0
            if isinstance(wait_widget, QTableWidgetItem):
                wait_ms = max(0, self._to_int(wait_widget.text(), 0))
 
            next_checks: List[str] = []
            if isinstance(next_check_widget, QTableWidgetItem):
                t = next_check_widget.text().strip()
                if t:
                    next_checks = [x.strip() for x in t.split(",") if x.strip()]
 
            comment = comment_widget.text().strip() if isinstance(comment_widget, QTableWidgetItem) else ""
 
            model = SetStepModel(
                signals=signal_panel.get_signals(),
                wait_ms=wait_ms,
                next_checks=next_checks,
                comment=comment,
            )
            new_content = render_set_step(model).strip()
 
            node.data["content"] = new_content
            for step in self.set_steps:
                if step["id"] == node_id:
                    step["content"] = new_content
                    break
 
            if self.editor:
                self.editor.set_module_widget.set_steps(self.set_steps)
 
            self._build_flowchart()
 
        elif node_type == "check":
            signal_panel = self.property_widgets.get("signal_rows")
            timeout_widget = self.property_widgets.get("timeout")
            duration_widget = self.property_widgets.get("duration")
            async_widget = self.property_widgets.get("async")
            comment_widget = self.property_widgets.get("comment")
 
            if not isinstance(signal_panel, _CheckSignalsInlineEditor):
                return
 
            timeout_ms = 1000
            if isinstance(timeout_widget, QTableWidgetItem):
                timeout_ms = max(1, self._to_int(timeout_widget.text(), 1000))
 
            duration_ms = 0
            if isinstance(duration_widget, QTableWidgetItem):
                duration_ms = max(0, self._to_int(duration_widget.text(), 0))
 
            async_value = False
            if isinstance(async_widget, QTableWidgetItem):
                async_value = self._to_bool(async_widget.text(), False)
 
            comment = comment_widget.text().strip() if isinstance(comment_widget, QTableWidgetItem) else ""
 
            items = signal_panel.get_items()
            for it in items:
                it.timeout_ms = timeout_ms
                it.duration_ms = duration_ms
                it.async_ = async_value
                if async_value:
                    it.wait_ms = 0
 
            model = CheckStepModel(items=items, comment=comment)
            new_content = render_check_step(model).strip()
 
            node.data["content"] = new_content
            for step in self.check_steps:
                if step["id"] == node_id:
                    step["content"] = new_content
                    break
 
            if self.editor:
                self.editor.check_module_widget.set_steps(self.check_steps)
 
            self._build_flowchart()
 
        for new_node in self.nodes:
            if new_node.node_id == node_id and new_node.node_type == node_type:
                new_node.setSelected(True)
                self.selected_node = new_node
                break
 
        QMessageBox.information(self, "成功", "属性已更新并保存")
        self.save_requested.emit()
        if self.save_callback:
            self.save_callback()

# ----------------------------
# Demo 入口（可删除/集成到你的主程序）
# ----------------------------
if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    w = QWidget()
    w.setWindowTitle("ModularCaseEditor Demo")
    w.resize(1200, 800)

    layout = QVBoxLayout(w)
    editor = ModularCaseEditor(w)

    # 示例补全（你后续用 DBC/系统变量索引替换）
    editor.set_completions(
        [
            "sys::FunctionSwitch::CSW_Enable_S",
            "sys::DriverAction::gear",
            "env::CAN 1::IPB_0x10C::VDC_Active",
            "sig::CAN 1::ADC_0x29C::CSW_Stats_S",
            "sig::CAN 1::ADC_0x29C::DNP_warning_text_info",
        ]
    )

    layout.addWidget(editor, 1)

    btn_row = QHBoxLayout()
    btn_export = QPushButton("导出 DSL 到弹窗", w)
    btn_validate = QPushButton("校验", w)

    def _on_export():
        dsl = editor.to_dsl()
        dlg = QDialog(w)
        dlg.setWindowTitle("导出 DSL")
        dlg.resize(900, 600)
        l = QVBoxLayout(dlg)
        t = QTextEdit(dlg)
        t.setPlainText(dsl)
        l.addWidget(t, 1)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Close, dlg)
        bb.rejected.connect(dlg.reject)
        bb.accepted.connect(dlg.accept)
        bb.clicked.connect(lambda: dlg.accept())
        l.addWidget(bb)
        dlg.exec()

    def _on_validate():
        errs = editor.validate()
        if not errs:
            QMessageBox.information(w, "校验通过", "未发现问题。")
        else:
            QMessageBox.warning(w, "校验失败", "\n".join(errs))

    btn_export.clicked.connect(_on_export)
    btn_validate.clicked.connect(_on_validate)
    btn_row.addWidget(btn_export)
    btn_row.addWidget(btn_validate)
    btn_row.addStretch(1)
    layout.addLayout(btn_row)

    w.show()
    sys.exit(app.exec())
