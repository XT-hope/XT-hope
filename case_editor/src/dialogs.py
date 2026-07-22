"""
对话框模块
包含各种配置和交互对话框
"""
import sys
import os
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel,
    QLineEdit, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QComboBox, QSpinBox, QCheckBox, QTextEdit, QMessageBox,
    QGroupBox, QRadioButton, QButtonGroup, QWidget, QScrollArea,
    QListWidget, QListWidgetItem, QAbstractItemView, QSplitter, QTabWidget
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QKeyEvent, QIntValidator
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
from .dbc_2_vsysvar import convert_dbc_files_to_vsysvar, derive_namespace_from_path
import openai

# 导入信号值提示辅助类（避免循环导入，使用延迟导入或直接导入）
try:
    from .case_editor import _SignalValueHelper, _HierLineEditCompleter, _build_hier_index_by_kind
except ImportError:
    from case_editor import _SignalValueHelper, _HierLineEditCompleter, _build_hier_index_by_kind


class ReadOnlyLineEdit(QLineEdit):
    """只读文本输入框，不支持粘贴操作"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
    
    def keyPressEvent(self, event: QKeyEvent) -> None:
        """按键事件处理，阻止粘贴操作"""
        # 检查是否是 Ctrl+V 粘贴操作
        if event.key() == Qt.Key.Key_V and event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            # 阻止粘贴操作
            return
        # 其他按键正常处理
        super().keyPressEvent(event)


class ReadOnlyTextEdit(QTextEdit):
    """只读文本编辑框，不支持粘贴操作"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
    
    def keyPressEvent(self, event: QKeyEvent) -> None:
        """按键事件处理，阻止粘贴操作"""
        # 检查是否是 Ctrl+V 粘贴操作
        if event.key() == Qt.Key.Key_V and event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            # 阻止粘贴操作
            return
        # 其他按键正常处理
        super().keyPressEvent(event)


class NewProjectDialog(QDialog):
    """新建项目对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建项目")
        self.setModal(True)
        self.resize(500, 200)
        
        self.project_path = ""
        self.project_name = ""
        
        self.init_ui()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # 项目名称
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("输入项目名称")
        form_layout.addRow("项目名称(&N):", self.name_edit)
        
        # 项目路径
        path_layout = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("选择项目保存位置")
        self.path_edit.setText("./projects")
        path_layout.addWidget(self.path_edit)
        
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self.browse_path)
        path_layout.addWidget(browse_btn)
        
        form_layout.addRow("项目位置(&L):", path_layout)
        
        layout.addLayout(form_layout)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def browse_path(self) -> None:
        """浏览路径"""
        path = QFileDialog.getExistingDirectory(self, "选择项目目录")
        if path:
            self.path_edit.setText(path)
    
    def accept(self) -> None:
        """接受对话框"""
        project_name = self.name_edit.text().strip()
        project_path = self.path_edit.text().strip()
        
        if not project_name:
            QMessageBox.warning(self, "警告", "请输入项目名称")
            return
        
        if not project_path:
            QMessageBox.warning(self, "警告", "请选择项目位置")
            return
        
        self.project_name = project_name
        self.project_path = project_path
        
        super().accept()
    
    def get_project_name(self) -> str:
        """获取项目名称"""
        return self.project_name
    
    def get_project_path(self) -> str:
        """获取项目路径"""
        return self.project_path


class DBCMappingDialog(QDialog):
    """DBC文件到CAN通道映射配置对话框"""

    def __init__(self, dbc_files: List[str], mapping: Dict[str, Dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle("配置CAN通道映射")
        self.setModal(True)
        self.resize(900, 500)

        self.dbc_files = dbc_files
        # mapping 格式: {dbc_path: {"channel": int, "short_name": str}}
        self.mapping = mapping.copy()

        self.init_ui()
        self.load_data()

    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)

        # 说明标签
        info_label = QLabel("配置DBC文件到CAN通道的映射关系（CAN通道和简称为必填项）")
        info_label.setStyleSheet("font-weight: bold; color: #666;")
        layout.addWidget(info_label)

        # 映射表格
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["DBC文件", "CAN通道", "简称", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)

    def load_data(self) -> None:
        """加载数据"""
        self.table.setRowCount(len(self.dbc_files))

        for row, dbc_file in enumerate(self.dbc_files):
            # DBC文件名
            name_item = QTableWidgetItem(Path(dbc_file).name)
            name_item.setData(Qt.ItemDataRole.UserRole, dbc_file)
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 0, name_item)

            # CAN通道 - 使用 QLineEdit 代替 QSpinBox，禁止上下箭头
            mapping_info = self.mapping.get(dbc_file, {})
            channel = mapping_info.get("channel", 0)
            channel_edit = QLineEdit()
            channel_edit.setText(str(channel))
            channel_edit.setPlaceholderText("输入通道号")
            # 只允许输入数字
            channel_edit.setValidator(QIntValidator(0, 255, self))
            channel_edit.textChanged.connect(self._on_channel_changed)
            self.table.setCellWidget(row, 1, channel_edit)

            # 简称
            short_name = mapping_info.get("short_name", "")
            short_name_edit = QLineEdit()
            short_name_edit.setText(short_name)
            short_name_edit.setPlaceholderText("输入简称")
            self.table.setCellWidget(row, 2, short_name_edit)

            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.clicked.connect(lambda checked, r=row: self.remove_mapping(r))
            self.table.setCellWidget(row, 3, delete_btn)

    def _on_channel_changed(self, text: str) -> None:
        """通道号改变时检查是否有重复"""
        self._check_duplicate_channels()

    def _check_duplicate_channels(self) -> None:
        """检查是否有重复的通道号，如果有则高亮显示"""
        channel_map = {}  # {channel: [row1, row2, ...]}

        for row in range(self.table.rowCount()):
            channel_edit = self.table.cellWidget(row, 1)
            if channel_edit:
                channel_text = channel_edit.text().strip()
                if channel_text:
                    try:
                        channel = int(channel_text)
                        if channel not in channel_map:
                            channel_map[channel] = []
                        channel_map[channel].append(row)
                    except ValueError:
                        pass

        # 重置所有行的样式
        for row in range(self.table.rowCount()):
            channel_edit = self.table.cellWidget(row, 1)
            if channel_edit:
                channel_edit.setStyleSheet("")

        # 高亮显示重复的通道
        for channel, rows in channel_map.items():
            if len(rows) > 1:
                for row in rows:
                    channel_edit = self.table.cellWidget(row, 1)
                    if channel_edit:
                        channel_edit.setStyleSheet("background-color: #ffcccc;")

    def remove_mapping(self, row: int) -> None:
        """删除映射"""
        self.table.removeRow(row)

    def accept(self) -> None:
        """接受对话框"""
        # 收集映射关系
        new_mapping = {}
        channels_used = {}  # 用于检测重复通道

        for row in range(self.table.rowCount()):
            dbc_file = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
            channel_edit = self.table.cellWidget(row, 1)
            channel_text = channel_edit.text().strip()
            short_name_edit = self.table.cellWidget(row, 2)
            short_name = short_name_edit.text().strip()

            # 验证通道号
            if not channel_text:
                QMessageBox.warning(self, "警告", f"请为 {Path(dbc_file).name} 填写CAN通道")
                return

            try:
                channel = int(channel_text)
                if channel < 0 or channel > 255:
                    QMessageBox.warning(self, "警告", f"CAN通道必须在 0-255 范围内")
                    return
            except ValueError:
                QMessageBox.warning(self, "警告", f"CAN通道必须是数字")
                return

            # 检查通道是否已被使用
            if channel in channels_used:
                QMessageBox.warning(self, "警告", f"CAN通道 {channel} 已被 {Path(channels_used[channel]).name} 使用，请更换通道")
                return
            channels_used[channel] = dbc_file

            # 验证必填项
            if not short_name:
                QMessageBox.warning(self, "警告", f"请为 {Path(dbc_file).name} 填写简称")
                return

            new_mapping[dbc_file] = {
                "channel": channel,
                "short_name": short_name
            }

        if not new_mapping:
            QMessageBox.warning(self, "警告", "至少需要配置一个DBC文件的映射")
            return

        self.mapping = new_mapping
        super().accept()

    def get_mapping(self) -> Dict[str, Dict]:
        """获取映射关系"""
        return self.mapping


class SystemVariableDialog(QDialog):
    """系统变量文件对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("系统变量文件")
        self.setModal(True)
        self.resize(600, 400)
        
        self.init_ui()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 说明
        info_label = QLabel("上传CANoe系统变量文件，用于系统变量的智能提示和补全")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666;")
        layout.addWidget(info_label)
        
        # 文件选择
        file_layout = QHBoxLayout()
        self.file_edit = ReadOnlyLineEdit()
        self.file_edit.setPlaceholderText("选择系统变量文件")
        file_layout.addWidget(self.file_edit)
        
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_btn)
        
        layout.addLayout(file_layout)
        
        # 变量列表
        layout.addWidget(QLabel("系统变量列表:"))
        self.variable_list = ReadOnlyTextEdit()
        layout.addWidget(self.variable_list)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def browse_file(self) -> None:
        """浏览文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择系统变量文件", "", "所有文件 (*.*)"
        )
        
        if file_path:
            self.file_edit.setText(file_path)
            # TODO: 解析文件并显示变量列表
    
    def get_file_path(self) -> str:
        """获取文件路径"""
        return self.file_edit.text()


class AIQuestionDialog(QDialog):
    """AI问答对话框"""
    
    def __init__(self, config_manager, parent=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.setWindowTitle("AI助手")
        self.setModal(True)
        self.resize(800, 600)
        
        self.ai_client = None
        self.init_ui()
        self._init_ai_client()
        self.check_ai_config()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 说明
        info_label = QLabel("向AI询问关于信号和变量的相关问题，获取智能帮助和建议")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666;")
        layout.addWidget(info_label)
        
        # 问题输入
        layout.addWidget(QLabel("您的问题:"))
        self.question_edit = QTextEdit()
        self.question_edit.setPlaceholderText("输入您的问题，例如：\n"
                                            "- CSW_Stats_S信号的含义是什么？\n"
                                            "- 如何设置VDC_Active信号？\n"
                                            "- 检查CSW_Stats_S是否等于3的语法是什么？")
        self.question_edit.setMaximumHeight(150)
        layout.addWidget(self.question_edit)
        
        # 上下文选择
        context_group = QGroupBox("上下文")
        context_layout = QVBoxLayout()
        
        self.include_case_cb = QCheckBox("包含当前Case内容")
        self.include_case_cb.setChecked(True)
        context_layout.addWidget(self.include_case_cb)
        
        self.include_signals_cb = QCheckBox("包含相关信号信息")
        self.include_signals_cb.setChecked(True)
        context_layout.addWidget(self.include_signals_cb)
        
        context_group.setLayout(context_layout)
        layout.addWidget(context_group)
        
        # 发送按钮
        send_btn = QPushButton("发送问题")
        send_btn.clicked.connect(self.send_question)
        layout.addWidget(send_btn)
        
        # 回答显示
        layout.addWidget(QLabel("AI回答:"))
        self.answer_text = ReadOnlyTextEdit()
        layout.addWidget(self.answer_text)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        config_btn = QPushButton("配置AI")
        config_btn.clicked.connect(self.open_ai_config)
        button_layout.addWidget(config_btn)
        
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
    
    def _init_ai_client(self) -> None:
        """初始化AI客户端"""
        try:
            ai_enabled = self.config_manager.get('ai.enabled', False)
            if not ai_enabled:
                self.ai_client = None
                return
            
            api_key = self.config_manager.get('ai.api_key', '')
            base_url = self.config_manager.get('ai.base_url', 'https://api.openai.com/v1')
            timeout = self.config_manager.get('ai.timeout', 60)
            
            if not api_key:
                self.ai_client = None
                return
            
            self.ai_client = openai.OpenAI(
                api_key=api_key,
                base_url=base_url,
                timeout=timeout
            )
        except Exception as e:
            print(f"初始化AI客户端失败: {e}")
            self.ai_client = None
    
    def check_ai_config(self) -> None:
        """检查AI配置"""
        ai_enabled = self.config_manager.get('ai.enabled', False)
        api_key = self.config_manager.get('ai.api_key', '')
        
        if not ai_enabled or not api_key:
            self.answer_text.setText("AI功能未配置。\n\n"
                                   "请点击下方\"配置AI\"按钮设置API密钥。\n\n"
                                   "配置步骤：\n"
                                   "1. 获取OpenAI API密钥\n"
                                   "2. 点击\"配置AI\"按钮\n"
                                   "3. 输入API密钥并保存")
        else:
            self.answer_text.setText("AI功能已就绪，请输入您的问题。")
    
    def open_ai_config(self) -> None:
        """打开AI配置对话框"""
        dialog = AIConfigDialog(self.config_manager, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._init_ai_client()
            self.check_ai_config()
    
    def send_question(self) -> None:
        """发送问题"""
        question = self.question_edit.toPlainText().strip()
        
        if not question:
            QMessageBox.warning(self, "警告", "请输入问题")
            return
        
        # 检查AI客户端
        if self.ai_client is None:
            QMessageBox.warning(self, "警告", "AI客户端未初始化，请先配置AI")
            return
        
        # 显示加载状态
        self.answer_text.setText("正在思考中，请稍候...")
        self.answer_text.setEnabled(False)
        
        # 构建消息
        system_prompt = self._get_system_prompt()
        user_prompt = self._build_user_prompt(question)
        
        try:
            # 获取配置
            model = self.config_manager.get('ai.model', 'gpt-4')
            max_tokens = self.config_manager.get('ai.max_tokens', 2000)
            
            # 调用OpenAI API
            response = self.ai_client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                max_tokens=max_tokens
            )
            
            # 显示回答
            answer = response.choices[0].message.content
            self.answer_text.setText(answer)
            self.answer_text.setEnabled(True)
            
        except Exception as e:
            error_msg = f"AI问答出错: {str(e)}"
            print(error_msg)
            self.answer_text.setText(f"抱歉，AI回答时出现错误：\n\n{error_msg}")
            self.answer_text.setEnabled(True)
            QMessageBox.critical(self, "错误", f"AI问答失败: {str(e)}")
    
    def _get_system_prompt(self) -> str:
        """获取系统提示词"""
        return """你是一个专业的自动驾驶HIL测试工程师助手，擅长CAN总线信号分析和测试用例编写。

你的职责：
1. 解释CAN信号的含义、取值范围、单位等信息
2. 提供信号设置的语法示例
3. 帮助编写和优化测试用例
4. 解答关于DBC文件、环境变量、系统变量的问题

回答要求：
- 使用简洁、专业的语言
- 提供具体的代码示例
- 如果涉及信号路径，使用完整格式（如：sig::CAN 0::ADC_0x29C::CSW_Stats_S）
- 对于复杂问题，分步骤说明
"""
    
    def _build_user_prompt(self, question: str) -> str:
        """构建用户提示词"""
        prompt_parts = [f"用户问题：{question}\n"]
        
        # 如果需要包含上下文
        if self.include_case_cb.isChecked():
            # TODO: 获取当前Case内容
            prompt_parts.append("\n当前Case上下文：\n（暂未实现）")
        
        if self.include_signals_cb.isChecked():
            # TODO: 获取相关信号信息
            prompt_parts.append("\n相关信号信息：\n（暂未实现）")
        
        return "\n".join(prompt_parts)
    
    def get_question(self) -> str:
        """获取问题"""
        return self.question_edit.toPlainText()

class AIConfigDialog(QDialog):
    """AI配置对话框"""
    
    def __init__(self, config_manager, parent=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.setWindowTitle("AI配置")
        self.setModal(True)
        self.resize(500, 400)
        
        self.init_ui()
        self.load_config()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 启用AI
        self.enabled_cb = QCheckBox("启用AI助手")
        self.enabled_cb.toggled.connect(self.on_enabled_changed)
        layout.addWidget(self.enabled_cb)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # API Key
        self.api_key_edit = QLineEdit()
        self.api_key_edit.setPlaceholderText("输入OpenAI API密钥")
        self.api_key_edit.setEchoMode(QLineEdit.EchoMode.Password)
        form_layout.addRow("API Key:", self.api_key_edit)
        
        # Base URL
        self.base_url_edit = QLineEdit()
        self.base_url_edit.setPlaceholderText("例如: https://api.openai.com/v1 或 http://aic.byd.com/model-gateway")
        form_layout.addRow("Base URL:", self.base_url_edit)
        
        # Model
        self.model_combo = QComboBox()
        self.model_combo.setEditable(True)
        self.model_combo.addItems(["gpt-4", "gpt-4-turbo", "gpt-3.5-turbo"])
        form_layout.addRow("模型:", self.model_combo)
        
        # Max Tokens
        self.max_tokens_spin = QSpinBox()
        self.max_tokens_spin.setRange(100, 65536)
        self.max_tokens_spin.setValue(2000)
        self.max_tokens_spin.setSuffix(" tokens")
        form_layout.addRow("最大Token数:", self.max_tokens_spin)
        
        # Timeout
        self.timeout_spin = QSpinBox()
        self.timeout_spin.setRange(10, 300)
        self.timeout_spin.setValue(60)
        self.timeout_spin.setSuffix(" 秒")
        form_layout.addRow("超时时间:", self.timeout_spin)
        
        layout.addLayout(form_layout)
        
        # 说明
        info_label = QLabel("提示：\n"
                          "1. 需要OpenAI API密钥才能使用AI功能\n"
                          "2. API密钥可在 https://platform.openai.com/api-keys 获取\n"
                          "3. Base URL 默认为 https://api.openai.com/v1\n"
                          "4. 如使用内部AI服务，可配置自定义Base URL\n"
                          "5. API密钥将保存在本地配置文件中")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666; padding: 10px; background: #f5f5f5;")
        layout.addWidget(info_label)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def load_config(self) -> None:
        """加载配置"""
        self.enabled_cb.setChecked(self.config_manager.get('ai.enabled', False))
        self.api_key_edit.setText(self.config_manager.get('ai.api_key', ''))
        self.base_url_edit.setText(self.config_manager.get('ai.base_url', 'https://api.openai.com/v1'))
        self.model_combo.setCurrentText(self.config_manager.get('ai.model', 'gpt-4'))
        self.max_tokens_spin.setValue(self.config_manager.get('ai.max_tokens', 2000))
        self.timeout_spin.setValue(self.config_manager.get('ai.timeout', 60))
        
        self.on_enabled_changed(self.enabled_cb.isChecked())
    
    def on_enabled_changed(self, enabled: bool) -> None:
        """启用状态改变"""
        self.api_key_edit.setEnabled(enabled)
        self.base_url_edit.setEnabled(enabled)
        self.model_combo.setEnabled(enabled)
        self.max_tokens_spin.setEnabled(enabled)
        self.timeout_spin.setEnabled(enabled)
    
    def accept(self) -> None:
        """接受对话框"""
        self.config_manager.set('ai.enabled', self.enabled_cb.isChecked())
        self.config_manager.set('ai.api_key', self.api_key_edit.text())
        self.config_manager.set('ai.base_url', self.base_url_edit.text())
        self.config_manager.set('ai.model', self.model_combo.currentText())
        self.config_manager.set('ai.max_tokens', self.max_tokens_spin.value())
        self.config_manager.set('ai.timeout', self.timeout_spin.value())
        
        if self.config_manager.save_config():
            QMessageBox.information(self, "成功", "AI配置保存成功")
            super().accept()
        else:
            QMessageBox.critical(self, "错误", "保存AI配置失败")



class OSSConfigDialog(QDialog):
    """OSS配置对话框"""
    
    def __init__(self, config_manager, parent=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.setWindowTitle("OSS配置")
        self.setModal(True)
        self.resize(500, 350)
        
        self.init_ui()
        self.load_config()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 启用OSS
        self.enabled_cb = QCheckBox("启用OSS存储")
        self.enabled_cb.toggled.connect(self.on_enabled_changed)
        layout.addWidget(self.enabled_cb)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # Endpoint
        self.endpoint_edit = QLineEdit()
        self.endpoint_edit.setPlaceholderText("例如: oss-cn-hangzhou.aliyuncs.com")
        form_layout.addRow("Endpoint:", self.endpoint_edit)
        
        # Access Key ID
        self.access_key_id_edit = QLineEdit()
        self.access_key_id_edit.setPlaceholderText("输入Access Key ID")
        form_layout.addRow("Access Key ID:", self.access_key_id_edit)
        
        # Access Key Secret
        self.access_key_secret_edit = QLineEdit()
        self.access_key_secret_edit.setPlaceholderText("输入Access Key Secret")
        self.access_key_secret_edit.setEchoMode(QLineEdit.EchoMode.Password)
        form_layout.addRow("Access Key Secret:", self.access_key_secret_edit)
        
        # Bucket Name
        self.bucket_name_edit = QLineEdit()
        self.bucket_name_edit.setPlaceholderText("输入Bucket名称")
        form_layout.addRow("Bucket Name:", self.bucket_name_edit)
        
        layout.addLayout(form_layout)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def load_config(self) -> None:
        """加载配置"""
        self.enabled_cb.setChecked(self.config_manager.get('oss.enabled', False))
        self.endpoint_edit.setText(self.config_manager.get('oss.endpoint', ''))
        self.access_key_id_edit.setText(self.config_manager.get('oss.access_key_id', ''))
        self.access_key_secret_edit.setText(self.config_manager.get('oss.access_key_secret', ''))
        self.bucket_name_edit.setText(self.config_manager.get('oss.bucket_name', ''))
        
        self.on_enabled_changed(self.enabled_cb.isChecked())
    
    def on_enabled_changed(self, enabled: bool) -> None:
        """启用状态改变"""
        self.endpoint_edit.setEnabled(enabled)
        self.access_key_id_edit.setEnabled(enabled)
        self.access_key_secret_edit.setEnabled(enabled)
        self.bucket_name_edit.setEnabled(enabled)
    
    def accept(self) -> None:
        """接受对话框"""
        self.config_manager.set('oss.enabled', self.enabled_cb.isChecked())
        self.config_manager.set('oss.endpoint', self.endpoint_edit.text())
        self.config_manager.set('oss.access_key_id', self.access_key_id_edit.text())
        self.config_manager.set('oss.access_key_secret', self.access_key_secret_edit.text())
        self.config_manager.set('oss.bucket_name', self.bucket_name_edit.text())
        
        if self.config_manager.save_config():
            QMessageBox.information(self, "成功", "OSS配置保存成功")
            super().accept()
        else:
            QMessageBox.critical(self, "错误", "保存OSS配置失败")


class DBCConverterDialog(QDialog):
    """DBC转换器对话框"""

    def __init__(self, parent=None, project_config=None, project_path=None):
        super().__init__(parent)
        self.setWindowTitle("DBC转换器")
        self.setModal(True)
        self.resize(600, 450)

        self.project_config = project_config or {}
        self.project_path = Path(project_path) if project_path else None

        self.init_ui()

    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)

        # 说明
        info_label = QLabel("此功能将根据车载DBC文件生成对应的系统变量文件")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666; padding: 10px;")
        layout.addWidget(info_label)

        # 源DBC文件区域
        source_group = QGroupBox("源DBC文件")
        source_layout = QVBoxLayout(source_group)

        # 项目DBC列表
        project_dbc_label = QLabel("项目DBC文件（勾选需要转换的）:")
        source_layout.addWidget(project_dbc_label)

        self.project_dbc_list = QListWidget()
        self.project_dbc_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.load_project_dbc_files()
        source_layout.addWidget(self.project_dbc_list)

        # 外部DBC区域
        external_layout = QHBoxLayout()
        external_label = QLabel("外部DBC文件:")
        external_layout.addWidget(external_label)

        add_external_btn = QPushButton("添加外部DBC")
        add_external_btn.clicked.connect(self.add_external_dbc)
        external_layout.addWidget(add_external_btn)

        remove_external_btn = QPushButton("删除选中")
        remove_external_btn.clicked.connect(self.remove_external_dbc)
        external_layout.addWidget(remove_external_btn)

        source_layout.addLayout(external_layout)

        self.external_dbc_list = QListWidget()
        self.external_dbc_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        source_layout.addWidget(self.external_dbc_list)

        layout.addWidget(source_group)

        # 目标文件区域
        target_group = QGroupBox("目标文件")
        target_layout = QFormLayout(target_group)

        # 保存位置
        if self.project_path:
            save_dir = str(self.project_path / "CANoe" / "system_variable")
        else:
            save_dir = "请先打开项目"

        self.save_dir_label = QLabel(save_dir)
        self.save_dir_label.setStyleSheet("color: #333; background: #f5f5f5; padding: 5px;")
        target_layout.addRow("保存位置:", self.save_dir_label)

        # 文件名输入
        self.target_filename_edit = QLineEdit()
        self.target_filename_edit.setPlaceholderText("请输入文件名（不含扩展名）")
        target_layout.addRow("文件名:", self.target_filename_edit)

        # 扩展名提示
        ext_label = QLabel(".vsysvar")
        ext_label.setStyleSheet("color: #666;")
        target_layout.addRow("扩展名:", ext_label)

        layout.addWidget(target_group)

        # 按钮区域
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        self.convert_btn = QPushButton("开始转换")
        self.convert_btn.clicked.connect(self.on_convert_clicked)
        button_layout.addWidget(self.convert_btn)

        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.setEnabled(False)
        self.cancel_btn.clicked.connect(self.on_cancel_clicked)
        button_layout.addWidget(self.cancel_btn)

        self.close_btn = QPushButton("关闭")
        self.close_btn.clicked.connect(self.accept)
        button_layout.addWidget(self.close_btn)

        layout.addLayout(button_layout)

    def load_project_dbc_files(self) -> None:
        """加载项目DBC文件列表"""
        canoe_config = self.project_config.get("canoe", {})
        dbc_files = canoe_config.get("dbc_files", {})

        for can_channel, dbc_info in dbc_files.items():
            path = dbc_info.get("path", "")
            short_name = dbc_info.get("short_name", "")
            channel = dbc_info.get("channel", 0)

            # 显示格式: CAN 1 - Control (filename.dbc)
            dbc_path = self.project_path / path if self.project_path and path else Path(path)
            filename = dbc_path.name if dbc_path else "未知"

            item = QListWidgetItem()
            checkbox = QCheckBox(f"{can_channel} - {short_name} ({filename})")
            checkbox.setProperty("dbc_path", str(dbc_path))
            checkbox.setProperty("can_channel", can_channel)
            checkbox.setProperty("short_name", short_name)
            checkbox.setProperty("channel", channel)

            self.project_dbc_list.addItem(item)
            self.project_dbc_list.setItemWidget(item, checkbox)

    def add_external_dbc(self) -> None:
        """添加外部DBC文件"""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "选择DBC文件", "", "DBC文件 (*.dbc);;所有文件 (*.*)"
        )

        for file_path in file_paths:
            if file_path:
                item = QListWidgetItem(Path(file_path).name)
                item.setData(Qt.ItemDataRole.UserRole, file_path)
                self.external_dbc_list.addItem(item)

    def remove_external_dbc(self) -> None:
        """删除选中的外部DBC文件"""
        current_item = self.external_dbc_list.currentItem()
        if current_item:
            self.external_dbc_list.takeItem(self.external_dbc_list.row(current_item))

    def get_selected_dbc_files(self) -> list:
        """获取选中的DBC文件列表"""
        selected_files = []

        # 获取项目DBC中勾选的文件
        for i in range(self.project_dbc_list.count()):
            item = self.project_dbc_list.item(i)
            checkbox = self.project_dbc_list.itemWidget(item)
            if checkbox and checkbox.isChecked():
                dbc_info = {
                    "path": checkbox.property("dbc_path"),
                    "can_channel": checkbox.property("can_channel"),
                    "short_name": checkbox.property("short_name"),
                    "channel": checkbox.property("channel"),
                    "type": "project"
                }
                selected_files.append(dbc_info)

        # 获取外部DBC文件
        for i in range(self.external_dbc_list.count()):
            item = self.external_dbc_list.item(i)
            file_path = item.data(Qt.ItemDataRole.UserRole)
            dbc_info = {
                "path": file_path,
                "type": "external"
            }
            selected_files.append(dbc_info)

        return selected_files

    def on_convert_clicked(self) -> None:
        """开始转换按钮点击"""
        selected_files = self.get_selected_dbc_files()
        target_filename = self.target_filename_edit.text().strip()

        if not selected_files:
            QMessageBox.warning(self, "警告", "请选择至少一个DBC文件进行转换")
            return

        if not target_filename:
            QMessageBox.warning(self, "警告", "请输入目标文件名")
            return

        # 禁用关闭按钮，启用取消按钮
        self.close_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)

        # 构建输出路径
        output_path = self.save_dir_label.text()
        if output_path == "请先打开项目":
            output_path = ""
        output_file = str(Path(output_path) / f"{target_filename}.vsysvar")

        # 构建DBC文件列表 [(namespace, path), ...]
        dbc_specs = []
        for dbc_info in selected_files:
            dbc_path = dbc_info.get("path", "")
            # 使用short_name作为namespace，如果没有则从文件名推导
            namespace = dbc_info.get("short_name", "")
            if not namespace:
                namespace = derive_namespace_from_path(dbc_path)
            dbc_specs.append((namespace, dbc_path))

        try:
            # 执行转换（使用gb18030编码处理中文DBC）
            convert_dbc_files_to_vsysvar(dbc_specs, output_file, encoding="gb18030")

            QMessageBox.information(self, "成功",
                f"DBC转换完成！\n\n"
                f"转换了 {len(dbc_specs)} 个DBC文件\n"
                f"输出文件: {output_file}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"转换失败：\n{str(e)}")

        # 恢复按钮状态
        self.close_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)

    def on_cancel_clicked(self) -> None:
        """取消按钮点击"""
        # 恢复关闭按钮，禁用取消按钮
        self.close_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)


class CANoeProjectDialog(QDialog):
    """CANoe工程文件地址配置对话框"""
    
    def __init__(self, project_path: str = "", parent=None):
        super().__init__(parent)
        self.setWindowTitle("CANoe工程文件地址")
        self.setModal(True)
        self.resize(600, 200)
        
        self.project_path = project_path
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 说明
        info_label = QLabel("配置CANoe工程文件的路径，用于后续测试执行")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666;")
        layout.addWidget(info_label)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # 工程文件路径
        path_layout = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("选择CANoe工程文件 (.cfg)")
        path_layout.addWidget(self.path_edit)
        
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self.browse_file)
        path_layout.addWidget(browse_btn)
        
        form_layout.addRow("工程文件路径:", path_layout)
        
        layout.addLayout(form_layout)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def load_data(self) -> None:
        """加载数据"""
        self.path_edit.setText(self.project_path)
    
    def browse_file(self) -> None:
        """浏览文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择CANoe工程文件", "", "CANoe工程文件 (*.cfg);;所有文件 (*.*)"
        )
        
        if file_path:
            self.path_edit.setText(file_path)
    
    def accept(self) -> None:
        """接受对话框"""
        self.project_path = self.path_edit.text().strip()
        super().accept()
    
    def get_project_path(self) -> str:
        """获取工程文件路径"""
        return self.project_path


class SimulinkFileDialog(QDialog):
    """Simulink文件管理对话框"""
    
    def __init__(self, files: List[Dict[str, Any]], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Simulink文件管理")
        self.setModal(True)
        self.resize(800, 500)
        
        self.files = files.copy()
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 说明
        info_label = QLabel("管理Simulink相关文件（m脚本、mat文件、simulink模型）")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666;")
        layout.addWidget(info_label)
        
        # 添加文件按钮
        add_layout = QHBoxLayout()
        add_layout.addWidget(QLabel("添加文件:"))
        
        add_m_btn = QPushButton("添加M脚本")
        add_m_btn.clicked.connect(lambda: self.add_file("m_script"))
        add_layout.addWidget(add_m_btn)
        
        add_mat_btn = QPushButton("添加MAT文件")
        add_mat_btn.clicked.connect(lambda: self.add_file("mat_file"))
        add_layout.addWidget(add_mat_btn)
        
        add_model_btn = QPushButton("添加Simulink模型")
        add_model_btn.clicked.connect(lambda: self.add_file("simulink_model"))
        add_layout.addWidget(add_model_btn)
        
        add_layout.addStretch()
        layout.addLayout(add_layout)
        
        # 文件列表表格
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["文件名", "类型", "路径", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def load_data(self) -> None:
        """加载数据"""
        self.table.setRowCount(len(self.files))
        
        type_names = {
            "m_script": ".m",
            "mat_file": ".mat",
            "simulink_model": ".slx"
        }
        
        for row, file_info in enumerate(self.files):
            # 文件名
            name_item = QTableWidgetItem(file_info["name"])
            name_item.setData(Qt.ItemDataRole.UserRole, file_info["name"])
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 0, name_item)
            
            # 类型
            type_name = type_names.get(file_info["type"], file_info["type"])
            type_item = QTableWidgetItem(type_name)
            type_item.setFlags(type_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 1, type_item)
            
            # 路径
            path_item = QTableWidgetItem(file_info["path"])
            path_item.setFlags(path_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 2, path_item)
            
            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.clicked.connect(lambda checked, r=row: self.remove_file(r))
            self.table.setCellWidget(row, 3, delete_btn)
    
    def add_file(self, file_type: str) -> None:
        """添加文件"""
        # 根据文件类型设置过滤器
        if file_type == "m_script":
            file_filter = "M脚本 (*.m);;所有文件 (*.*)"
        elif file_type == "mat_file":
            file_filter = "MAT文件 (*.mat);;所有文件 (*.*)"
        elif file_type == "simulink_model":
            file_filter = "Simulink模型 (*.slx *.mdl);;所有文件 (*.*)"
        else:
            file_filter = "所有文件 (*.*)"
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, f"选择{file_type}", "", file_filter
        )
        
        if file_path:
            file = Path(file_path)
            file_info = {
                "name": file.name,
                "path": file_path,
                "type": file_type,
                "created_time": ""
            }
            
            # 检查是否已存在
            existing_files = [f for f in self.files if f["name"] == file.name]
            if existing_files:
                # 更新现有文件
                idx = self.files.index(existing_files[0])
                self.files[idx] = file_info
            else:
                # 添加新文件
                self.files.append(file_info)
            
            self.load_data()
    
    def remove_file(self, row: int) -> None:
        """删除文件"""
        file_name = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        self.files = [f for f in self.files if f["name"] != file_name]
        self.load_data()
    
    def get_files(self) -> List[Dict[str, Any]]:
        """获取文件列表"""
        return self.files


class SceneMappingDialog(QDialog):
    """场景映射表对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("场景映射表")
        self.setModal(True)
        self.resize(600, 250)
        
        self.mapping_name = ""
        self.mapping_file_path = ""
        
        self.init_ui()
    
    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        
        # 说明
        info_label = QLabel("配置场景映射表，用于场景ID与场景名称的映射关系\n"
                          "Excel文件必须包含以下列：场景描述、场景文件名、场景编号")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666;")
        layout.addWidget(info_label)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # 映射表名称
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("输入映射表名称")
        form_layout.addRow("映射表名称:", self.name_edit)
        
        # Excel文件路径
        file_layout = QHBoxLayout()
        self.file_edit = ReadOnlyLineEdit()
        self.file_edit.setPlaceholderText("选择Excel文件 (.xlsx, .xls)")
        file_layout.addWidget(self.file_edit)
        
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_btn)
        
        form_layout.addRow("Excel文件:", file_layout)
        
        layout.addLayout(form_layout)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def browse_file(self) -> None:
        """浏览文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        
        if file_path:
            self.file_edit.setText(file_path)
            # 如果名称为空，自动使用文件名
            if not self.name_edit.text():
                file_name = Path(file_path).stem
                self.name_edit.setText(file_name)
    
    def validate_excel_file(self, file_path: str) -> tuple:
        """
        验证Excel文件格式和内容
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            (是否验证通过, 错误信息)
        """
        try:
            import openpyxl
            
            # 打开Excel文件
            if file_path.endswith('.xls'):
                # 对于.xls文件，需要使用xlrd
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(file_path)
                    sheet = workbook.sheet_by_index(0)
                    
                    # 获取表头
                    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                    
                    # 检查必需的列
                    required_columns = ['场景描述', '场景文件名', '场景编号']
                    for col in required_columns:
                        if col not in headers:
                            return False, f"Excel文件缺少必需的列：{col}"
                    
                    # 检查是否有数据行
                    if sheet.nrows <= 1:
                        return False, "Excel文件为空，至少需要一条数据"
                    
                    # 获取列索引
                    desc_col = headers.index('场景描述')
                    filename_col = headers.index('场景文件名')
                    number_col = headers.index('场景编号')
                    
                    # 验证每一行数据
                    for row_idx in range(1, sheet.nrows):
                        # 场景文件名必须以.mat结尾
                        filename = str(sheet.cell_value(row_idx, filename_col)).strip()
                        if not filename.endswith('.mat'):
                            return False, f"第{row_idx + 1}行：场景文件名必须以.mat结尾"
                        
                        # 场景编号必须是大于等于0的整数
                        number = sheet.cell_value(row_idx, number_col)
                        try:
                            number_int = int(number)
                            if number_int < 0:
                                return False, f"第{row_idx + 1}行：场景编号必须大于等于0"
                        except (ValueError, TypeError):
                            return False, f"第{row_idx + 1}行：场景编号必须是整数"
                    
                    return True, ""
                    
                except ImportError:
                    return False, "处理.xls文件需要安装xlrd库，请运行：pip install xlrd"
            else:
                # 对于.xlsx文件，使用openpyxl
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                
                # 获取表头
                headers = [cell.value for cell in sheet[1]]
                
                # 检查必需的列
                required_columns = ['场景描述', '场景文件名', '场景编号']
                for col in required_columns:
                    if col not in headers:
                        return False, f"Excel文件缺少必需的列：{col}"
                
                # 检查是否有数据行
                if sheet.max_row <= 1:
                    return False, "Excel文件为空，至少需要一条数据"
                
                # 获取列索引
                desc_col = headers.index('场景描述')
                filename_col = headers.index('场景文件名')
                number_col = headers.index('场景编号')
                
                # 验证每一行数据
                for row_idx in range(2, sheet.max_row + 1):
                    # 场景文件名必须以.mat结尾
                    filename = str(sheet.cell(row_idx, filename_col + 1).value).strip()
                    if not filename.endswith('.mat'):
                        return False, f"第{row_idx}行：场景文件名必须以.mat结尾"
                    
                    # 场景编号必须是大于等于0的整数
                    number = sheet.cell(row_idx, number_col + 1).value
                    try:
                        number_int = int(number)
                        if number_int < 0:
                            return False, f"第{row_idx}行：场景编号必须大于等于0"
                    except (ValueError, TypeError):
                        return False, f"第{row_idx}行：场景编号必须是整数"
                
                return True, ""
                
        except Exception as e:
            return False, f"读取Excel文件失败: {str(e)}"
    
    def accept(self) -> None:
        """接受对话框"""
        mapping_name = self.name_edit.text().strip()
        file_path = self.file_edit.text().strip()
        
        if not mapping_name:
            QMessageBox.warning(self, "警告", "请输入映射表名称")
            return
        
        if not file_path:
            QMessageBox.warning(self, "警告", "请选择Excel文件")
            return
        
        # 验证文件是否存在
        if not Path(file_path).exists():
            QMessageBox.warning(self, "警告", "文件不存在")
            return
        
        # 验证文件扩展名
        file_ext = Path(file_path).suffix.lower()
        if file_ext not in ['.xlsx', '.xls']:
            QMessageBox.warning(self, "警告", "只支持.xlsx和.xls格式的Excel文件")
            return
        
        # 验证Excel文件内容
        is_valid, error_msg = self.validate_excel_file(file_path)
        if not is_valid:
            QMessageBox.warning(self, "验证失败", error_msg)
            return
        
        self.mapping_name = mapping_name
        self.mapping_file_path = file_path
        
        super().accept()
    
    def get_mapping_name(self) -> str:
        """获取映射表名称"""
        return self.mapping_name

    def get_mapping_file_path(self) -> str:
        """获取映射表文件路径"""
        return self.mapping_file_path


class PanelFileInfoDialog(QDialog):
    """面板文件详情对话框"""

    def __init__(self, panel_info: Dict[str, Any], parent=None):
        super().__init__(parent)
        self.setWindowTitle("面板文件详情")
        self.setModal(True)
        self.resize(400, 200)

        self._panel_info = panel_info
        self.init_ui()

    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 信息区域（直接使用FormLayout，不加GroupBox）
        info_layout = QFormLayout()
        info_layout.setSpacing(10)

        # 文件名
        file_name = self._panel_info.get("name", "-")
        name_label = QLabel(file_name)
        info_layout.addRow("文件名:", name_label)

        # 所属CAN (namespace)
        namespace = self._panel_info.get("namespace", "-")
        namespace_label = QLabel(namespace if namespace else "-")
        info_layout.addRow("所属CAN:", namespace_label)

        # 所属报文 (message_name)
        message_name = self._panel_info.get("message_name", "-")
        message_label = QLabel(message_name if message_name else "-")
        info_layout.addRow("所属报文:", message_label)

        # 创建时间
        created_time = self._panel_info.get("created_time", "-")
        if created_time:
            try:
                dt = datetime.fromisoformat(created_time)
                formatted_time = dt.strftime("%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                formatted_time = created_time
        else:
            formatted_time = "-"
        time_label = QLabel(formatted_time)
        info_layout.addRow("创建时间:", time_label)

        # 文件路径
        file_path = self._panel_info.get("path", "-")
        path_label = QLabel(file_path if file_path else "-")
        path_label.setWordWrap(True)
        info_layout.addRow("路径:", path_label)

        layout.addLayout(info_layout)
        layout.addStretch()

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        close_btn.setMinimumWidth(80)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)


class CANoePanelDialog(QDialog):
    """CANoe面板生成对话框 - 使用CheckBox勾选方式"""

    def __init__(self, project_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("CANoe面板生成")
        self.setModal(True)
        self.setMinimumSize(900, 650)

        self._project_manager = project_manager
        self._selected_sysvar_path = ""
        self._namespace_data = {}
        self._selected_variables = {}
        self._var_widgets = {}  # 存储每个namespace下的checkbox

        self._apply_styles()
        self.init_ui()
        self.load_system_variable_files()

    def _apply_styles(self) -> None:
        """应用现代化样式"""
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
                background-color: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 4px 12px;
                background-color: #e9ecef;
                border-radius: 4px;
            }
            QListWidget {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: #ffffff;
                padding: 2px;
                outline: none;
            }
            QListWidget::item {
                padding: 2px 4px;
                margin: 0px;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: 500;
            }
            QPushButton#generateBtn {
                background-color: #27ae60;
                color: white;
                border: none;
            }
            QPushButton#generateBtn:hover {
                background-color: #2ecc71;
            }
            QPushButton#generateBtn:pressed {
                background-color: #1e8449;
            }
            QPushButton#cancelBtn {
                background-color: #95a5a6;
                color: white;
                border: none;
            }
            QPushButton#cancelBtn:hover {
                background-color: #7f8c8d;
            }
            QPushButton#selectAllBtn, QPushButton#clearBtn, QPushButton#invertBtn {
                background-color: #e9ecef;
                color: #2c3e50;
                border: 1px solid #dee2e6;
            }
            QPushButton#selectAllBtn:hover, QPushButton#clearBtn:hover, QPushButton#invertBtn:hover {
                background-color: #d5dbdb;
            }
            QLabel#headerLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
            }
            QLabel#countLabel {
                color: #7f8c8d;
                font-size: 12px;
            }
            QTextEdit {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: #ffffff;
                padding: 8px;
                font-family: 'Consolas', 'Monaco', monospace;
            }
        """)

    def init_ui(self) -> None:
        """初始化UI - 使用CheckBox勾选方式"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(16)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 使用水平分割器分割左侧和右侧
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ===== 左侧面板：系统变量文件选择 =====
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        # 文件选择区（使用CheckBox）
        file_group = QGroupBox("系统变量文件（勾选需要解析的）")
        file_group.setFixedWidth(280)
        file_layout = QVBoxLayout(file_group)
        file_layout.setSpacing(8)

        self.sysvar_list = QListWidget()
        self.sysvar_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.sysvar_list.setMinimumHeight(120)
        file_layout.addWidget(self.sysvar_list)

        self.file_count_label = QLabel()
        self.file_count_label.setObjectName("countLabel")
        file_layout.addWidget(self.file_count_label)

        left_layout.addWidget(file_group)
        left_layout.addStretch()

        splitter.addWidget(left_panel)

        # ===== 右侧面板：报文选择 =====
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)

        # Namespace选择区（使用CheckBox）
        ns_group = QGroupBox("Namespace（勾选需要的）")
        ns_layout = QVBoxLayout(ns_group)
        ns_layout.setSpacing(8)

        # Namespace列表和快捷操作按钮
        ns_list_layout = QHBoxLayout()
        self.namespace_list = QListWidget()
        self.namespace_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        ns_list_layout.addWidget(self.namespace_list, 1)

        # 右侧快捷按钮
        ns_btn_layout = QVBoxLayout()
        ns_btn_layout.setSpacing(6)

        select_all_ns_btn = QPushButton("全选")
        select_all_ns_btn.setObjectName("selectAllBtn")
        select_all_ns_btn.setFixedWidth(70)
        select_all_ns_btn.clicked.connect(self._select_all_namespace)
        ns_btn_layout.addWidget(select_all_ns_btn)

        clear_ns_btn = QPushButton("清空")
        clear_ns_btn.setObjectName("clearBtn")
        clear_ns_btn.setFixedWidth(70)
        clear_ns_btn.clicked.connect(self._clear_namespace)
        ns_btn_layout.addWidget(clear_ns_btn)

        ns_btn_layout.addStretch()
        ns_list_layout.addLayout(ns_btn_layout)

        ns_layout.addLayout(ns_list_layout)

        self.ns_count_label = QLabel()
        self.ns_count_label.setObjectName("countLabel")
        ns_layout.addWidget(self.ns_count_label)

        right_layout.addWidget(ns_group)

        # Variable选择区（多列显示，每列一个namespace）
        var_group = QGroupBox("Message Variables（勾选需要的）")
        var_layout = QVBoxLayout(var_group)
        var_layout.setSpacing(8)

        # 使用ScrollArea容纳多列布局
        self.var_scroll = QScrollArea()
        self.var_scroll.setWidgetResizable(True)
        self.var_scroll.setMinimumHeight(150)
        self.var_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.var_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        var_layout.addWidget(self.var_scroll)

        # 快捷操作按钮行
        var_btn_row = QHBoxLayout()
        select_all_var_btn = QPushButton("全选")
        select_all_var_btn.setObjectName("selectAllBtn")
        select_all_var_btn.clicked.connect(self._select_all_variable)
        var_btn_row.addWidget(select_all_var_btn)

        clear_var_btn = QPushButton("清空")
        clear_var_btn.setObjectName("clearBtn")
        clear_var_btn.clicked.connect(self._clear_variable)
        var_btn_row.addWidget(clear_var_btn)

        invert_var_btn = QPushButton("反选")
        invert_var_btn.setObjectName("invertBtn")
        invert_var_btn.clicked.connect(self._invert_variable)
        var_btn_row.addWidget(invert_var_btn)

        var_btn_row.addStretch()
        self.var_count_label = QLabel()
        self.var_count_label.setObjectName("countLabel")
        var_btn_row.addWidget(self.var_count_label)
        var_layout.addLayout(var_btn_row)

        right_layout.addWidget(var_group)

        splitter.addWidget(right_panel)

        # 设置分割器比例
        splitter.setSizes([280, 620])

        main_layout.addWidget(splitter, 1)

        # ===== 底部：已选择报文预览 =====
        preview_group = QGroupBox("已选择的报文预览")
        preview_layout = QVBoxLayout(preview_group)
        preview_layout.setSpacing(8)

        preview_header = QHBoxLayout()
        self.preview_count_label = QLabel()
        self.preview_count_label.setObjectName("countLabel")
        preview_header.addWidget(self.preview_count_label)
        preview_header.addStretch()
        preview_layout.addLayout(preview_header)

        self.selected_text = ReadOnlyTextEdit()
        self.selected_text.setPlaceholderText("请在上方勾选 Namespace 和 Variable...")
        self.selected_text.setMinimumHeight(100)
        self.selected_text.setMaximumHeight(150)
        preview_layout.addWidget(self.selected_text)

        main_layout.addWidget(preview_group)

        # ===== 底部按钮栏 =====
        button_layout = QHBoxLayout()
        button_layout.setSpacing(12)
        button_layout.addStretch()

        generate_btn = QPushButton("生成面板")
        generate_btn.setObjectName("generateBtn")
        generate_btn.setMinimumWidth(120)
        generate_btn.clicked.connect(self._on_generate)
        button_layout.addWidget(generate_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setObjectName("cancelBtn")
        cancel_btn.setMinimumWidth(100)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        main_layout.addLayout(button_layout)

    def load_system_variable_files(self) -> None:
        """加载项目中的系统变量文件列表（使用CheckBox）"""
        self.sysvar_list.clear()

        if not self._project_manager or not self._project_manager.is_project_open():
            self.file_count_label.setText("项目未打开")
            return

        sysvar_files = self._project_manager.get_system_variable_files()
        count = 0
        for sysvar_path in sysvar_files:
            full_path = self._project_manager.get_full_path(sysvar_path)
            if full_path and full_path.exists():
                item = QListWidgetItem()
                checkbox = QCheckBox(Path(sysvar_path).name)
                checkbox.setProperty("sysvar_path", str(full_path))
                checkbox.stateChanged.connect(self._on_sysvar_checkbox_changed)
                self.sysvar_list.addItem(item)
                self.sysvar_list.setItemWidget(item, checkbox)
                count += 1

        self.file_count_label.setText(f"共 {count} 个文件")

    def _on_sysvar_checkbox_changed(self, state: int) -> None:
        """系统变量文件CheckBox状态改变"""
        checkbox = self.sender()
        if checkbox and state == Qt.CheckState.Checked.value:
            self._selected_sysvar_path = checkbox.property("sysvar_path")
            # 清除其他CheckBox的选中状态
            for i in range(self.sysvar_list.count()):
                item = self.sysvar_list.item(i)
                other_checkbox = self.sysvar_list.itemWidget(item)
                if other_checkbox and other_checkbox != checkbox:
                    other_checkbox.setChecked(False)
            # 解析并填充namespace列表
            self._parse_sysvar_file()
            self._populate_namespace_list()
        elif checkbox and state == Qt.CheckState.Unchecked.value:
            if checkbox.property("sysvar_path") == self._selected_sysvar_path:
                self._selected_sysvar_path = ""
                self.namespace_list.clear()
                self.variable_list.clear()
                self._namespace_data = {}
                self._selected_variables = {}
                self.ns_count_label.setText("")
                self.var_count_label.setText("")
                self._update_selected_display()

    def _select_all_namespace(self) -> None:
        """全选所有namespace"""
        for i in range(self.namespace_list.count()):
            item = self.namespace_list.item(i)
            checkbox = self.namespace_list.itemWidget(item)
            if checkbox:
                checkbox.setChecked(True)

    def _clear_namespace(self) -> None:
        """清空namespace选择"""
        for i in range(self.namespace_list.count()):
            item = self.namespace_list.item(i)
            checkbox = self.namespace_list.itemWidget(item)
            if checkbox:
                checkbox.setChecked(False)

    def _select_all_variable(self) -> None:
        """全选所有variable"""
        if hasattr(self, '_var_widgets'):
            for ns_name, var_widgets in self._var_widgets.items():
                for var_name, checkbox in var_widgets.items():
                    checkbox.setChecked(True)

    def _clear_variable(self) -> None:
        """清空variable选择"""
        if hasattr(self, '_var_widgets'):
            for ns_name, var_widgets in self._var_widgets.items():
                for var_name, checkbox in var_widgets.items():
                    checkbox.setChecked(False)

    def _invert_variable(self) -> None:
        """反选variable"""
        if hasattr(self, '_var_widgets'):
            for ns_name, var_widgets in self._var_widgets.items():
                for var_name, checkbox in var_widgets.items():
                    checkbox.setChecked(not checkbox.isChecked())

    def _parse_sysvar_file(self) -> None:
        """解析系统变量文件，提取namespace和variable结构"""
        self._namespace_data = {}

        if not self._selected_sysvar_path or not Path(self._selected_sysvar_path).exists():
            return

        try:
            import xml.etree.ElementTree as ET
            tree = ET.parse(self._selected_sysvar_path)
            root = tree.getroot()

            for elem in root.iter():
                if elem.tag == 'systemvariables':
                    for top_ns in elem.findall('./namespace'):
                        self._parse_namespace_recursive(top_ns, "")
                    break
                elif elem.tag == 'namespace':
                    self._parse_namespace_recursive(elem, "")
                    break

        except Exception as e:
            print(f"解析系统变量文件失败: {e}")
            QMessageBox.warning(self, "警告", f"解析系统变量文件失败: {str(e)}")

    def _parse_namespace_recursive(self, namespace_elem: 'ET.Element', parent_path: str) -> None:
        """递归解析namespace"""
        ns_name = namespace_elem.get('name', '')

        if parent_path:
            current_path = f"{parent_path}::{ns_name}"
        else:
            current_path = ns_name

        variables = []
        for var_elem in namespace_elem.findall('./variable'):
            var_name = var_elem.get('name', '')
            if var_name and '_Info' not in var_name:
                variables.append(var_name)

        if variables:
            if current_path not in self._namespace_data:
                self._namespace_data[current_path] = []
            self._namespace_data[current_path].extend(variables)

        for child_ns in namespace_elem.findall('./namespace'):
            self._parse_namespace_recursive(child_ns, current_path)

    def _populate_namespace_list(self) -> None:
        """填充namespace列表（使用CheckBox）"""
        self.namespace_list.clear()
        self._var_widgets = {}
        self.var_scroll.setWidget(QWidget())  # 清空variable显示区域

        for ns_name in sorted(self._namespace_data.keys()):
            item = QListWidgetItem()
            checkbox = QCheckBox(ns_name)
            checkbox.setProperty("ns_name", ns_name)
            checkbox.stateChanged.connect(self._on_namespace_checkbox_changed)
            self.namespace_list.addItem(item)
            self.namespace_list.setItemWidget(item, checkbox)

        self.ns_count_label.setText(f"共 {self.namespace_list.count()} 个 Namespace")
        self.var_count_label.setText("")

    def _on_namespace_checkbox_changed(self, state: int) -> None:
        """Namespace CheckBox状态改变"""
        self._update_variable_list()
        self._update_selected_display()

    def _update_variable_list(self) -> None:
        """根据选中的namespace更新variable列表（多列布局）"""
        # 获取所有选中的namespace
        selected_ns = []
        for i in range(self.namespace_list.count()):
            item = self.namespace_list.item(i)
            checkbox = self.namespace_list.itemWidget(item)
            if checkbox and checkbox.isChecked():
                selected_ns.append(checkbox.property("ns_name"))

        # 在重建UI之前，保存已勾选的状态
        self._selected_vars_cache = {}
        if hasattr(self, '_var_widgets') and self._var_widgets:
            for ns_name, var_widgets in self._var_widgets.items():
                checked_vars = []
                for var_name, checkbox in var_widgets.items():
                    if checkbox.isChecked():
                        checked_vars.append(var_name)
                if checked_vars:
                    self._selected_vars_cache[ns_name] = checked_vars

        # 清除现有内容
        container = QWidget()
        container_layout = QHBoxLayout(container)
        container_layout.setSpacing(10)
        container_layout.setContentsMargins(5, 5, 5, 5)

        self._var_widgets = {}  # 存储每个namespace下的checkbox: {ns_name: {var_name: checkbox}}
        total_count = 0

        if not selected_ns:
            self.var_scroll.setWidget(container)
            self.var_count_label.setText("")
            return

        # 为每个namespace创建一列
        for ns_name in sorted(selected_ns):
            column_widget = QWidget()
            column_layout = QVBoxLayout(column_widget)
            column_layout.setSpacing(4)
            column_layout.setContentsMargins(5, 5, 5, 5)

            # 列标题
            title_label = QLabel(ns_name)
            title_label.setStyleSheet("font-weight: bold; color: #2c3e50; background-color: #e9ecef; padding: 4px; border-radius: 3px;")
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            column_layout.addWidget(title_label)

            # 该namespace下的variables
            variables = self._namespace_data.get(ns_name, [])
            self._var_widgets[ns_name] = {}

            # 获取之前缓存的勾选状态
            cached_checked = self._selected_vars_cache.get(ns_name, [])

            for var_name in sorted(variables):
                checkbox = QCheckBox(var_name)
                checkbox.setProperty("var_name", var_name)
                checkbox.setProperty("ns_name", ns_name)
                checkbox.stateChanged.connect(self._on_variable_checkbox_changed)
                # 恢复勾选状态
                if var_name in cached_checked:
                    checkbox.setChecked(True)
                column_layout.addWidget(checkbox)
                self._var_widgets[ns_name][var_name] = checkbox
                total_count += 1

            column_layout.addStretch()
            column_widget.setMinimumWidth(120)
            container_layout.addWidget(column_widget)

        container_layout.addStretch()
        self.var_scroll.setWidget(container)
        self.var_count_label.setText(f"共 {total_count} 个 Variable")

    def _on_variable_checkbox_changed(self, state: int) -> None:
        """Variable CheckBox状态改变"""
        self._update_selected_display()

    def _update_selected_display(self) -> None:
        """更新已选择报文的显示"""
        self._selected_variables = {}

        if not hasattr(self, '_var_widgets'):
            self.selected_text.setText("")
            self.preview_count_label.setText("")
            return

        total_count = 0
        for ns_name, var_widgets in self._var_widgets.items():
            selected_vars = []
            for var_name, checkbox in var_widgets.items():
                if checkbox.isChecked():
                    selected_vars.append(var_name)
            if selected_vars:
                self._selected_variables[ns_name] = selected_vars
                total_count += len(selected_vars)

        if not self._selected_variables:
            self.selected_text.setText("")
            self.preview_count_label.setText("")
            return

        # 格式化显示
        display_text = ""
        for ns_name, vars_list in sorted(self._selected_variables.items()):
            display_text += f"📂 {ns_name}\n"
            for var_name in sorted(vars_list):
                display_text += f"   └─ {var_name}\n"
            display_text += "\n"

        self.selected_text.setText(display_text.strip())
        self.preview_count_label.setText(f"已选择 {len(self._selected_variables)} 个 Namespace，共 {total_count} 个 Variable")

    def _on_generate(self) -> None:
        """生成面板"""
        if not self._selected_sysvar_path:
            QMessageBox.warning(self, "提示", "请先勾选一个系统变量文件")
            return

        if not self._selected_variables:
            QMessageBox.warning(self, "提示", "请勾选需要生成面板的 Namespace 和 Variable")
            return

        if not self._project_manager or not self._project_manager.is_project_open():
            QMessageBox.warning(self, "提示", "项目未打开")
            return

        self.accept()

    def get_selected_sysvar_path(self) -> str:
        """获取选中的系统变量文件完整路径"""
        return self._selected_sysvar_path

    def get_selected_variables(self) -> Dict[str, List[str]]:
        """获取选中的变量字典"""
        return self._selected_variables


class CANoeCAPLDialog(QDialog):
    """CANoe仿真节点CAPL生成对话框"""

    def __init__(self, project_manager, dbc_parser, parent=None):
        super().__init__(parent)
        self.setWindowTitle("CANoe仿真节点CAPL生成")
        self.setModal(True)
        self.setMinimumSize(1000, 750)

        self._project_manager = project_manager
        self._dbc_parser = dbc_parser
        self._selected_sysvar_path = ""

        # 存储DBC数据
        self._dbc_data = {}  # {dbc_path: {"db": Database, "nodes": [], "name": str}}
        self._selected_dbc_paths = []  # 选中的DBC路径列表

        # 存储节点选择
        self._node_selections = {}  # {dbc_path: selected_node}

        # 存储报文选择和校验配置（包含发送节点信息）
        self._message_configs = {}  # {dbc_path: {message_name: {"sender_node": str, "check_signal": str, "counter_signal": str, "check_method": str, "check_parameters": dict}}}

        self._apply_styles()
        self.init_ui()
        self.load_dbc_files()
        self.load_system_variable_files()

    def _apply_styles(self) -> None:
        """应用现代化样式"""
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
                background-color: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 4px 12px;
                background-color: #e9ecef;
                border-radius: 4px;
            }
            QListWidget {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: #ffffff;
                padding: 2px;
                outline: none;
            }
            QListWidget::item {
                padding: 2px 4px;
                margin: 0px;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: 500;
            }
            QPushButton#generateBtn {
                background-color: #27ae60;
                color: white;
                border: none;
            }
            QPushButton#generateBtn:hover {
                background-color: #2ecc71;
            }
            QPushButton#generateBtn:pressed {
                background-color: #1e8449;
            }
            QPushButton#cancelBtn {
                background-color: #95a5a6;
                color: white;
                border: none;
            }
            QPushButton#cancelBtn:hover {
                background-color: #7f8c8d;
            }
            QPushButton#funcBtn {
                background-color: #3498db;
                color: white;
                border: none;
            }
            QPushButton#funcBtn:hover {
                background-color: #5dade2;
            }
            QTextEdit {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: #ffffff;
                padding: 8px;
                font-family: 'Consolas', 'Monaco', monospace;
            }
            QLabel#countLabel {
                color: #7f8c8d;
                font-size: 12px;
            }
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: #ffffff;
            }
        """)

    def init_ui(self) -> None:
        """初始化UI"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(16)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 使用水平分割器分割左侧和右侧
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ===== 左侧面板：DBC文件选择 =====
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        dbc_group = QGroupBox("DBC文件（勾选需要的）")
        dbc_group.setFixedWidth(280)
        dbc_layout = QVBoxLayout(dbc_group)
        dbc_layout.setSpacing(8)

        self.dbc_list = QListWidget()
        self.dbc_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.dbc_list.setMinimumHeight(120)
        dbc_layout.addWidget(self.dbc_list)

        self.dbc_count_label = QLabel()
        self.dbc_count_label.setObjectName("countLabel")
        dbc_layout.addWidget(self.dbc_count_label)

        left_layout.addWidget(dbc_group)
        left_layout.addStretch()

        splitter.addWidget(left_panel)

        # ===== 右侧面板 =====
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)

        # === 第二块：中心节点选择 ===
        node_group = QGroupBox("中心节点选择")
        node_layout = QVBoxLayout(node_group)
        node_layout.setSpacing(8)

        self.node_scroll = QScrollArea()
        self.node_scroll.setWidgetResizable(True)
        self.node_scroll.setMinimumHeight(100)
        self.node_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.node_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        node_layout.addWidget(self.node_scroll)

        self.node_count_label = QLabel()
        self.node_count_label.setObjectName("countLabel")
        node_layout.addWidget(self.node_count_label)

        right_layout.addWidget(node_group)

        # === 第三块：校验&Counting信号选择 ===
        msg_group = QGroupBox("校验&Counting信号选择（报文勾选）")
        msg_layout = QVBoxLayout(msg_group)
        msg_layout.setSpacing(8)

        self.msg_scroll = QScrollArea()
        self.msg_scroll.setWidgetResizable(True)
        self.msg_scroll.setMinimumHeight(150)
        self.msg_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.msg_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        msg_layout.addWidget(self.msg_scroll)

        # 功能按钮行
        func_btn_row = QHBoxLayout()

        config_btn = QPushButton("配置校验信号")
        config_btn.setObjectName("funcBtn")
        config_btn.clicked.connect(self._on_select_check_signal)
        func_btn_row.addWidget(config_btn)

        func_btn_row.addStretch()
        self.msg_count_label = QLabel()
        self.msg_count_label.setObjectName("countLabel")
        func_btn_row.addWidget(self.msg_count_label)
        msg_layout.addLayout(func_btn_row)

        right_layout.addWidget(msg_group)

        splitter.addWidget(right_panel)
        splitter.setSizes([280, 720])

        main_layout.addWidget(splitter, 1)

        # ===== 第四块：预览 + 系统变量文件选择 =====
        preview_group = QGroupBox("信号校验方式/校验参数预览")
        preview_layout = QVBoxLayout(preview_group)
        preview_layout.setSpacing(8)

        preview_header = QHBoxLayout()
        self.preview_count_label = QLabel()
        self.preview_count_label.setObjectName("countLabel")
        preview_header.addWidget(self.preview_count_label)
        preview_header.addStretch()
        preview_layout.addLayout(preview_header)

        self.preview_text = ReadOnlyTextEdit()
        self.preview_text.setPlaceholderText("请先勾选DBC文件，选择中心节点和报文...")
        self.preview_text.setMinimumHeight(120)
        self.preview_text.setMaximumHeight(200)
        preview_layout.addWidget(self.preview_text)

        # 系统变量文件选择
        sysvar_row = QHBoxLayout()
        sysvar_label = QLabel("系统变量文件：")
        sysvar_row.addWidget(sysvar_label)

        self.sysvar_combo = QComboBox()
        self.sysvar_combo.setMinimumWidth(250)
        sysvar_row.addWidget(self.sysvar_combo)

        sysvar_row.addStretch()
        preview_layout.addLayout(sysvar_row)

        main_layout.addWidget(preview_group)

        # ===== 底部按钮栏 =====
        button_layout = QHBoxLayout()
        button_layout.setSpacing(12)
        button_layout.addStretch()

        generate_btn = QPushButton("生成CAPL")
        generate_btn.setObjectName("generateBtn")
        generate_btn.setMinimumWidth(120)
        generate_btn.clicked.connect(self._on_generate)
        button_layout.addWidget(generate_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setObjectName("cancelBtn")
        cancel_btn.setMinimumWidth(100)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        main_layout.addLayout(button_layout)

    def load_dbc_files(self) -> None:
        """加载项目中的DBC文件列表"""
        self.dbc_list.clear()
        self._dbc_data.clear()

        if not self._project_manager or not self._project_manager.is_project_open():
            self.dbc_count_label.setText("项目未打开")
            return

        # 从项目配置中获取DBC文件信息
        dbc_files_config = self._project_manager.project_config.get("canoe", {}).get("dbc_files", {})
        count = 0

        # 构建DBC信息列表（统一处理新旧格式）
        dbc_info_list = []
        if isinstance(dbc_files_config, dict):
            for can_key, dbc_info in dbc_files_config.items():
                if isinstance(dbc_info, dict) and dbc_info.get("path"):
                    dbc_info_list.append({
                        "path": dbc_info.get("path"),
                        "short_name": dbc_info.get("short_name", Path(dbc_info.get("path")).stem)
                    })
        else:
            # 兼容旧格式（列表）
            for dbc_path in dbc_files_config:
                dbc_info_list.append({
                    "path": dbc_path,
                    "short_name": Path(dbc_path).stem
                })

        # 加载DBC文件
        for dbc_info in dbc_info_list:
            dbc_path = dbc_info["path"]
            short_name = dbc_info["short_name"]
            full_path = self._project_manager.get_full_path(dbc_path)

            if full_path and full_path.exists():
                try:
                    import cantools
                    db = cantools.database.load_file(str(full_path))
                    nodes = [n.name for n in db.nodes]

                    self._dbc_data[str(full_path)] = {
                        "db": db,
                        "nodes": nodes,
                        "name": short_name,
                        "relative_path": dbc_path
                    }

                    # 添加CheckBox到列表（显示简称）
                    item = QListWidgetItem()
                    checkbox = QCheckBox(short_name)
                    checkbox.setProperty("dbc_path", str(full_path))
                    checkbox.stateChanged.connect(self._on_dbc_checkbox_changed)
                    self.dbc_list.addItem(item)
                    self.dbc_list.setItemWidget(item, checkbox)
                    count += 1
                except Exception as e:
                    print(f"加载DBC文件失败 {full_path}: {e}")

        self.dbc_count_label.setText(f"共 {count} 个 DBC文件")

    def load_system_variable_files(self) -> None:
        """加载系统变量文件列表"""
        self.sysvar_combo.clear()
        self.sysvar_combo.addItem("请选择系统变量文件")

        if not self._project_manager or not self._project_manager.is_project_open():
            return

        sysvar_files = self._project_manager.get_system_variable_files()
        for sysvar_path in sysvar_files:
            full_path = self._project_manager.get_full_path(sysvar_path)
            if full_path and full_path.exists():
                self.sysvar_combo.addItem(Path(sysvar_path).name, str(full_path))

    def _on_dbc_checkbox_changed(self, state: int) -> None:
        """DBC文件CheckBox状态改变"""
        checkbox = self.sender()
        if checkbox:
            dbc_path = checkbox.property("dbc_path")
            if state == Qt.CheckState.Checked.value:
                if dbc_path not in self._selected_dbc_paths:
                    self._selected_dbc_paths.append(dbc_path)
            else:
                if dbc_path in self._selected_dbc_paths:
                    self._selected_dbc_paths.remove(dbc_path)
                    # 清除该DBC的节点选择和报文配置
                    if dbc_path in self._node_selections:
                        del self._node_selections[dbc_path]
                    if dbc_path in self._message_configs:
                        del self._message_configs[dbc_path]

        self._update_node_selection_panel()
        self._update_message_panel()
        self._update_preview()

    def _update_node_selection_panel(self) -> None:
        """更新中心节点选择面板（每个DBC一列，单选）"""
        container = QWidget()
        container_layout = QHBoxLayout(container)
        container_layout.setSpacing(16)
        container_layout.setContentsMargins(5, 5, 5, 5)

        if not self._selected_dbc_paths:
            self.node_scroll.setWidget(container)
            self.node_count_label.setText("")
            return

        total_nodes = 0
        for dbc_path in sorted(self._selected_dbc_paths):
            dbc_info = self._dbc_data.get(dbc_path)
            if not dbc_info:
                continue

            column_widget = QWidget()
            column_layout = QVBoxLayout(column_widget)
            column_layout.setSpacing(4)
            column_layout.setContentsMargins(5, 5, 5, 5)

            # 列标题（DBC简称）
            title_label = QLabel(dbc_info["name"])
            title_label.setStyleSheet("font-weight: bold; color: #2c3e50; background-color: #e9ecef; padding: 4px; border-radius: 3px;")
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            column_layout.addWidget(title_label)

            # 使用ButtonGroup实现单选
            button_group = QButtonGroup(column_widget)
            button_group.setProperty("dbc_path", dbc_path)
            button_group.buttonClicked.connect(self._on_node_radio_clicked)

            for node_name in sorted(dbc_info["nodes"]):
                radio = QRadioButton(node_name)
                radio.setProperty("node_name", node_name)
                button_group.addButton(radio)
                # 恢复之前的选择
                if dbc_path in self._node_selections and self._node_selections[dbc_path] == node_name:
                    radio.setChecked(True)
                column_layout.addWidget(radio)
                total_nodes += 1

            column_layout.addStretch()
            column_widget.setMinimumWidth(150)
            container_layout.addWidget(column_widget)

        container_layout.addStretch()
        self.node_scroll.setWidget(container)
        self.node_count_label.setText(f"共 {len(self._selected_dbc_paths)} 个DBC，总计 {total_nodes} 个节点")

    def _on_node_radio_clicked(self, button: QRadioButton) -> None:
        """节点单选按钮点击"""
        button_group = button.group()
        dbc_path = button_group.property("dbc_path")
        node_name = button.property("node_name")

        self._node_selections[dbc_path] = node_name
        self._update_message_panel()
        self._update_preview()

    def _get_messages_received_by_node(self, db, node_name: str) -> Dict[str, List[str]]:
        """获取节点接收的报文列表，按发送节点分组

        Returns:
            Dict[str, List[str]]: {sender_node: [msg_name1, msg_name2, ...]}
        """
        result = {}
        for msg in db.messages:
            recv = set()
            for sig in msg.signals:
                recv.update(sig.receivers)
            if node_name in recv:
                # 获取发送节点（可能有多个发送者，通常只有一个）
                senders = getattr(msg, 'senders', [])
                if senders:
                    sender = senders[0]  # 通常报文只有一个发送节点
                else:
                    sender = "Unknown"  # 如果没有发送节点信息

                if sender not in result:
                    result[sender] = []
                result[sender].append(msg.name)

        # 对每个发送节点的报文列表排序
        for sender in result:
            result[sender] = sorted(result[sender])

        return result

    def _update_message_panel(self) -> None:
        """更新报文选择面板（按发送节点分组显示）"""
        container = QWidget()
        container_layout = QHBoxLayout(container)
        container_layout.setSpacing(16)
        container_layout.setContentsMargins(5, 5, 5, 5)

        self._msg_widgets = {}  # {dbc_path: {msg_name: checkbox}}
        self._msg_senders = {}  # {dbc_path: {msg_name: sender_node}} 记录每个报文的发送节点
        total_msgs = 0

        if not self._selected_dbc_paths:
            self.msg_scroll.setWidget(container)
            self.msg_count_label.setText("")
            return

        for dbc_path in sorted(self._selected_dbc_paths):
            dbc_info = self._dbc_data.get(dbc_path)
            selected_node = self._node_selections.get(dbc_path)

            if not dbc_info or not selected_node:
                continue

            # 获取该节点接收的报文（按发送节点分组）
            received_msgs_by_sender = self._get_messages_received_by_node(dbc_info["db"], selected_node)

            if not received_msgs_by_sender:
                continue

            column_widget = QWidget()
            column_layout = QVBoxLayout(column_widget)
            column_layout.setSpacing(4)
            column_layout.setContentsMargins(5, 5, 5, 5)

            # 列标题（DBC简称 + 中心节点）
            title_label = QLabel(f"{dbc_info['name']} [{selected_node}]")
            title_label.setStyleSheet("font-weight: bold; color: #2c3e50; background-color: #e9ecef; padding: 4px; border-radius: 3px;")
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            column_layout.addWidget(title_label)

            self._msg_widgets[dbc_path] = {}
            self._msg_senders[dbc_path] = {}

            # 初始化该DBC的报文配置字典
            if dbc_path not in self._message_configs:
                self._message_configs[dbc_path] = {}

            # 获取该DBC已勾选的报文列表（用于恢复勾选状态）
            checked_msgs = set(self._message_configs[dbc_path].keys())

            # 按发送节点分组显示
            for sender_node in sorted(received_msgs_by_sender.keys()):
                msg_names = received_msgs_by_sender[sender_node]

                # 发送节点分组标题
                sender_label = QLabel(f"  ↳ 发送节点: {sender_node}")
                sender_label.setStyleSheet("color: #e67e22; font-weight: bold; font-size: 11px; margin-left: 5px;")
                column_layout.addWidget(sender_label)

                for msg_name in msg_names:
                    checkbox = QCheckBox(f"    {msg_name}")
                    checkbox.setProperty("dbc_path", dbc_path)
                    checkbox.setProperty("msg_name", msg_name)
                    checkbox.setProperty("sender_node", sender_node)  # 记录发送节点
                    checkbox.stateChanged.connect(self._on_msg_checkbox_changed)
                    # 恢复勾选状态
                    if msg_name in checked_msgs:
                        checkbox.setChecked(True)
                    column_layout.addWidget(checkbox)
                    self._msg_widgets[dbc_path][msg_name] = checkbox
                    self._msg_senders[dbc_path][msg_name] = sender_node  # 记录发送节点
                    total_msgs += 1

            column_layout.addStretch()
            column_widget.setMinimumWidth(250)
            container_layout.addWidget(column_widget)

        container_layout.addStretch()
        self.msg_scroll.setWidget(container)
        self.msg_count_label.setText(f"共 {total_msgs} 个接收报文")

    def _on_msg_checkbox_changed(self, state: int) -> None:
        """报文CheckBox状态改变"""
        checkbox = self.sender()
        if checkbox:
            dbc_path = checkbox.property("dbc_path")
            msg_name = checkbox.property("msg_name")
            sender_node = checkbox.property("sender_node")  # 获取发送节点

            if dbc_path not in self._message_configs:
                self._message_configs[dbc_path] = {}

            if state == Qt.CheckState.Checked.value:
                # 添加报文配置（初始为空，包含发送节点信息）
                if msg_name not in self._message_configs[dbc_path]:
                    self._message_configs[dbc_path][msg_name] = {
                        "sender_node": sender_node,  # 记录发送节点
                        "check_signal": "",
                        "counter_signal": "",
                        "check_method": "crc16",
                        "check_parameters": {
                            "poly": "0x1021",
                            "init": "0xFFFF",
                            "refIn": "false",
                            "refOut": "false",
                            "xorOut": "0x0000"
                        }
                    }
            else:
                # 移除报文配置
                if msg_name in self._message_configs[dbc_path]:
                    del self._message_configs[dbc_path][msg_name]

        self._update_preview()

    def _get_selected_messages(self) -> List[Tuple[str, str]]:
        """获取所有选中的报文，返回 [(dbc_path, msg_name)]"""
        selected = []
        for dbc_path, msg_widgets in self._msg_widgets.items():
            for msg_name, checkbox in msg_widgets.items():
                if checkbox.isChecked():
                    selected.append((dbc_path, msg_name))
        return selected

    def _on_select_check_signal(self) -> None:
        """校验信号选择 - 打开统一的配置对话框"""
        selected_msgs = self._get_selected_messages()
        if not selected_msgs:
            QMessageBox.warning(self, "提示", "请先勾选报文")
            return

        # 弹出统一的表格式配置对话框
        dialog = MessageCheckConfigDialog(selected_msgs, self._dbc_data, self._message_configs, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._message_configs = dialog.get_configs()
            self._update_preview()

    def _update_preview(self) -> None:
        """更新预览显示（带颜色高亮，显示发送节点）"""
        preview_html = ""
        total_signals = 0

        for dbc_path in sorted(self._selected_dbc_paths):
            dbc_info = self._dbc_data.get(dbc_path)
            if not dbc_info:
                continue

            dbc_name = dbc_info["name"]
            center_node = self._node_selections.get(dbc_path, "")
            msg_configs = self._message_configs.get(dbc_path, {})

            if msg_configs:
                # DBC名称用蓝色高亮，显示中心节点
                preview_html += f'<span style="color: #2980b9; font-weight: bold;">{dbc_name}</span> '
                preview_html += f'<span style="color: #7f8c8d;">[中心节点: {center_node}]</span><br>'

                # 按发送节点分组显示
                sender_msgs = {}
                for msg_name in msg_configs.keys():
                    config = msg_configs[msg_name]
                    sender_node = config.get("sender_node", "Unknown")
                    if sender_node not in sender_msgs:
                        sender_msgs[sender_node] = []
                    sender_msgs[sender_node].append((msg_name, config))

                for sender_node in sorted(sender_msgs.keys()):
                    # 发送节点用橙色高亮
                    preview_html += f'    <span style="color: #e67e22; font-weight: bold;">↳ 发送节点: {sender_node}</span><br>'
                    for msg_name, config in sender_msgs[sender_node]:
                        check_signal = config.get("check_signal", "")
                        counter_signal = config.get("counter_signal", "")
                        check_method = config.get("check_method", "crc16")
                        check_parameters = config.get("check_parameters", {})

                        if check_signal:
                            # 报文名称用深灰色
                            preview_html += f'        <span style="color: #34495e;">{msg_name}</span> '
                            # check_signal=xxx 用绿色
                            preview_html += f'<span style="color: #27ae60;">check_signal</span>=<span style="color: #e74c3c;">{check_signal}</span> '
                            # check_method 用紫色
                            preview_html += f'<span style="color: #8e44ad;">check_method</span>=<span style="color: #e74c3c;">{check_method}</span> '
                            if check_method == "crc16":
                                # 参数用橙色
                                params_str = " ".join([f'<span style="color: #d35400;">{k}</span>=<span style="color: #e74c3c;">{v}</span>' for k, v in check_parameters.items()])
                                preview_html += params_str
                            preview_html += '<br>'
                            total_signals += 1

                        if counter_signal:
                            preview_html += f'        <span style="color: #34495e;">{msg_name}</span> '
                            preview_html += f'<span style="color: #f39c12;">counter_signal</span>=<span style="color: #e74c3c;">{counter_signal}</span><br>'
                            total_signals += 1

                preview_html += '<br>'

        self.preview_text.setHtml(preview_html.strip())
        self.preview_count_label.setText(f"共 {total_signals} 个信号配置")

    def _on_generate(self) -> None:
        """生成CAPL"""
        # 验证选择
        if not self._selected_dbc_paths:
            QMessageBox.warning(self, "提示", "请先勾选DBC文件")
            return

        if not self._node_selections:
            QMessageBox.warning(self, "提示", "请选择中心节点")
            return

        # 检查每个选中的DBC是否都有节点选择
        for dbc_path in self._selected_dbc_paths:
            if dbc_path not in self._node_selections:
                dbc_info = self._dbc_data.get(dbc_path)
                dbc_name = dbc_info.get("name", dbc_path) if dbc_info else dbc_path
                QMessageBox.warning(self, "提示", f"请为 {dbc_name} 选择中心节点")
                return

        # 检查是否有报文配置
        has_config = False
        for dbc_path, msg_configs in self._message_configs.items():
            if msg_configs:
                has_config = True
                break

        if not has_config:
            QMessageBox.warning(self, "提示", "请勾选报文并配置校验信号")
            return

        # 检查系统变量文件
        sysvar_index = self.sysvar_combo.currentIndex()
        if sysvar_index <= 0:
            QMessageBox.warning(self, "提示", "请选择系统变量文件")
            return

        self._selected_sysvar_path = self.sysvar_combo.currentData()

        # 验证勾选报文中“声明有校验”的报文配置完整性
        # 说明：
        # - 未勾选 / 未配置 check_signal 的报文视为 has_validation=false，直接放行，不拦截。
        # - 允许“只有 checksum 没有 counter”的报文通过（has_counting_signal=false）。
        #   即有 check_signal 即视为有校验，counter_signal 可为空。
        for dbc_path, msg_configs in self._message_configs.items():
            for msg_name, config in msg_configs.items():
                # 仅当用户为该报文选择了 check_signal 时，才视为“有校验”
                if not config.get("check_signal"):
                    continue  # 无 check_signal → has_validation=false，放行
                # counter_signal 可为空：只有 checksum 没有 counting 的报文也合法

        self.accept()

    def get_capl_config(self) -> Dict[str, Any]:
        """获取CAPL生成配置数据（按发送节点组织，返回中心节点接收的全部报文）。

        每条报文用 has_validation 标记是否含校验配置：
        - True: 用户已勾选并配置了 check_signal/counter_signal，带完整 check 字段
        - False: 未勾选或未配置，仅含 message_name
        """
        dbc_configs = []

        for dbc_path in sorted(self._selected_dbc_paths):
            dbc_info = self._dbc_data.get(dbc_path)
            if not dbc_info:
                continue

            node = self._node_selections.get(dbc_path, "")  # 中心节点（接收节点）
            if not node:
                continue

            msg_configs = self._message_configs.get(dbc_path, {})

            # 重新计算该中心节点接收的全部报文，按发送节点分组
            # {sender_node: [msg_name1, msg_name2, ...]}
            received_by_sender = self._get_messages_received_by_node(dbc_info["db"], node)

            # 按发送节点组装报文条目（包含全部报文，无 check 的标 has_validation=false）
            sender_messages: Dict[str, List[Dict[str, Any]]] = {}
            for sender_node in sorted(received_by_sender.keys()):
                msgs = []
                for msg_name in received_by_sender[sender_node]:
                    config = msg_configs.get(msg_name)
                    if config and config.get("check_signal"):
                        # 有校验：带完整 check 配置
                        counter_signal = config.get("counter_signal", "")
                        msg_entry: Dict[str, Any] = {
                            "message_name": msg_name,
                            "has_validation": True,
                            "has_counting_signal": bool(counter_signal),  # 仅 checksum 无 counting 时为 False
                            "check_signal": config.get("check_signal", ""),
                            "check_method": config.get("check_method", "crc16"),
                        }
                        # 有 counting 才带 counter_signal；无 counting 不输出该字段
                        if counter_signal:
                            msg_entry["counter_signal"] = counter_signal
                        if config.get("check_method", "crc16") == "crc16":
                            msg_entry["check_parameters"] = config.get("check_parameters", {})
                    else:
                        # 无校验：仅报文名
                        msg_entry = {
                            "message_name": msg_name,
                            "has_validation": False,
                        }
                    msgs.append(msg_entry)
                sender_messages[sender_node] = msgs

            senders = [
                {"sender_node": sn, "messages": sender_messages[sn]}
                for sn in sorted(sender_messages.keys())
            ]

            if senders:
                dbc_configs.append({
                    "dbc_path": dbc_path,
                    "dbc_name": dbc_info["name"],
                    "center_node": node,  # 中心节点（接收节点）
                    "senders": senders,  # 发送节点及其全部报文（含 has_validation 标记）
                })

        return {
            "dbc_configs": dbc_configs,
            "selected_system_variable_file": self._selected_sysvar_path
        }


class CRC16ParamConfigDialog(QDialog):
    """CRC16参数配置对话框"""

    def __init__(self, current_params: Dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("CRC16参数配置")
        self.setModal(True)
        self.setMinimumSize(400, 300)

        self._current_params = current_params.copy()
        self._param_inputs = {}

        self._apply_styles()
        self.init_ui()

    def _apply_styles(self) -> None:
        """应用样式"""
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
                background-color: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 4px 12px;
                background-color: #e9ecef;
                border-radius: 4px;
            }
            QLineEdit {
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 6px;
                background-color: white;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
            QComboBox {
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 6px;
                background-color: white;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: 500;
            }
            QPushButton#okBtn {
                background-color: #27ae60;
                color: white;
                border: none;
            }
            QPushButton#okBtn:hover {
                background-color: #2ecc71;
            }
            QPushButton#cancelBtn {
                background-color: #95a5a6;
                color: white;
                border: none;
            }
            QPushButton#cancelBtn:hover {
                background-color: #7f8c8d;
            }
            QLabel#descLabel {
                color: #666;
                font-size: 11px;
            }
        """)

    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)

        # 说明标签
        hint_label = QLabel("配置CRC16校验算法的各项参数")
        hint_label.setObjectName("descLabel")
        layout.addWidget(hint_label)

        # 参数配置组
        param_group = QGroupBox("CRC16参数")
        param_layout = QFormLayout(param_group)
        param_layout.setSpacing(12)

        # poly参数
        poly_edit = QLineEdit()
        poly_edit.setPlaceholderText("例如: 0x1021")
        poly_edit.setText(self._current_params.get("poly", "0x1021"))
        poly_edit.setProperty("param_name", "poly")
        param_layout.addRow("Poly:", poly_edit)
        self._param_inputs["poly"] = poly_edit

        # init参数
        init_edit = QLineEdit()
        init_edit.setPlaceholderText("例如: 0xFFFF")
        init_edit.setText(self._current_params.get("init", "0xFFFF"))
        init_edit.setProperty("param_name", "init")
        param_layout.addRow("Init:", init_edit)
        self._param_inputs["init"] = init_edit

        # xorOut参数
        xor_edit = QLineEdit()
        xor_edit.setPlaceholderText("例如: 0x0000")
        xor_edit.setText(self._current_params.get("xorOut", "0x0000"))
        xor_edit.setProperty("param_name", "xorOut")
        param_layout.addRow("XorOut:", xor_edit)
        self._param_inputs["xorOut"] = xor_edit

        # refIn参数（下拉框）
        refIn_combo = QComboBox()
        refIn_combo.addItems(["false", "true"])
        refIn_combo.setCurrentText(self._current_params.get("refIn", "false"))
        refIn_combo.setProperty("param_name", "refIn")
        param_layout.addRow("RefIn:", refIn_combo)
        self._param_inputs["refIn"] = refIn_combo

        # refOut参数（下拉框）
        refOut_combo = QComboBox()
        refOut_combo.addItems(["false", "true"])
        refOut_combo.setCurrentText(self._current_params.get("refOut", "false"))
        refOut_combo.setProperty("param_name", "refOut")
        param_layout.addRow("RefOut:", refOut_combo)
        self._param_inputs["refOut"] = refOut_combo

        layout.addWidget(param_group)

        # 参数说明
        desc_group = QGroupBox("参数说明")
        desc_layout = QVBoxLayout(desc_group)
        desc_text = QLabel(
            "<b>Poly</b>: 多项式（16位，十六进制）<br>"
            "<b>Init</b>: 初始值（16位，十六进制）<br>"
            "<b>XorOut</b>: 输出异或值（16位，十六进制）<br>"
            "<b>RefIn</b>: 输入反转（true/false）<br>"
            "<b>RefOut</b>: 输出反转（true/false）"
        )
        desc_text.setObjectName("descLabel")
        desc_text.setWordWrap(True)
        desc_layout.addWidget(desc_text)
        layout.addWidget(desc_group)

        layout.addStretch()

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.setObjectName("okBtn")
        ok_btn.setMinimumWidth(100)
        ok_btn.clicked.connect(self._on_ok)
        btn_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setObjectName("cancelBtn")
        cancel_btn.setMinimumWidth(100)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def _on_ok(self) -> None:
        """确认保存参数"""
        # 验证参数
        params = {}
        for param_name, input_widget in self._param_inputs.items():
            if isinstance(input_widget, QLineEdit):
                value = input_widget.text().strip()
                # 验证十六进制格式（可选）
                if value and not (value.startswith("0x") or value.startswith("0X")):
                    # 允许十进制输入，但建议十六进制
                    pass
                params[param_name] = value
            elif isinstance(input_widget, QComboBox):
                params[param_name] = input_widget.currentText()

        self._current_params = params
        self.accept()

    def get_params(self) -> Dict:
        """获取配置的参数"""
        return self._current_params


class MessageCheckConfigDialog(QDialog):
    """报文校验配置对话框 - 使用QTabWidget实现Excel Sheet风格"""

    def __init__(self, selected_msgs, dbc_data, message_configs, parent=None):
        super().__init__(parent)
        self.setWindowTitle("报文校验配置")
        self.setModal(True)
        self.setMinimumSize(900, 600)

        self._selected_msgs = selected_msgs  # [(dbc_path, msg_name)]
        self._dbc_data = dbc_data
        self._message_configs = message_configs.copy()

        # 按DBC分组报文
        self._msgs_by_dbc = {}  # {dbc_path: [msg_names]}
        self._tab_tables = {}  # {dbc_path: QTableWidget}

        self._apply_styles()
        self.init_ui()

    def _apply_styles(self) -> None:
        """应用现代化样式"""
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QTabWidget::pane {
                border: 1px solid #dee2e6;
                background-color: #ffffff;
            }
            QTabBar::tab {
                background-color: #e9ecef;
                color: #2c3e50;
                padding: 8px 20px;
                margin-right: 2px;
                border: 1px solid #dee2e6;
                border-bottom: none;
                border-radius: 4px 4px 0 0;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover:!selected {
                background-color: #d5dbdb;
            }
            QTableWidget {
                border: 1px solid #dee2e6;
                background-color: #ffffff;
                gridline-color: #dee2e6;
            }
            QTableWidget::item {
                padding: 4px;
            }
            QHeaderView::section {
                background-color: #e9ecef;
                color: #2c3e50;
                font-weight: bold;
                border: 1px solid #dee2e6;
                padding: 6px;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: 500;
            }
            QPushButton#okBtn {
                background-color: #27ae60;
                color: white;
                border: none;
            }
            QPushButton#okBtn:hover {
                background-color: #2ecc71;
            }
            QPushButton#cancelBtn {
                background-color: #95a5a6;
                color: white;
                border: none;
            }
            QPushButton#cancelBtn:hover {
                background-color: #7f8c8d;
            }
            QComboBox {
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 4px 8px;
                background-color: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #dee2e6;
                background-color: white;
                selection-background-color: #3498db;
                selection-color: white;
            }
            QLineEdit {
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 4px 8px;
                background-color: white;
            }
            QLabel#hintLabel {
                color: #666;
                font-size: 12px;
            }
        """)

    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)

        # 提示标签
        hint_label = QLabel("为每个报文配置校验信号、Counter信号、校验方式和校验参数")
        hint_label.setObjectName("hintLabel")
        layout.addWidget(hint_label)

        # 按DBC分组报文
        for dbc_path, msg_name in self._selected_msgs:
            if dbc_path not in self._msgs_by_dbc:
                self._msgs_by_dbc[dbc_path] = []
            self._msgs_by_dbc[dbc_path].append(msg_name)

        # 创建TabWidget
        self.tab_widget = QTabWidget()

        # 为每个DBC创建一个Tab
        for dbc_path in sorted(self._msgs_by_dbc.keys()):
            dbc_info = self._dbc_data.get(dbc_path)
            dbc_name = dbc_info.get("name", Path(dbc_path).stem) if dbc_info else Path(dbc_path).stem

            # 创建表格
            table_widget = self._create_table_for_dbc(dbc_path)
            self._tab_tables[dbc_path] = table_widget

            # 添加Tab
            self.tab_widget.addTab(table_widget, dbc_name)

        layout.addWidget(self.tab_widget, 1)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.setObjectName("okBtn")
        ok_btn.setMinimumWidth(100)
        ok_btn.clicked.connect(self._on_ok)
        btn_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setObjectName("cancelBtn")
        cancel_btn.setMinimumWidth(100)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def _create_table_for_dbc(self, dbc_path: str) -> QTableWidget:
        """为指定DBC创建表格"""
        msg_names = self._msgs_by_dbc.get(dbc_path, [])
        dbc_info = self._dbc_data.get(dbc_path)

        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels([
            "报文名称", "校验信号", "Counter信号", "校验方式", "校验参数"
        ])

        # 设置列宽
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)

        table.setColumnWidth(0, 150)
        table.setColumnWidth(3, 100)

        table.setRowCount(len(msg_names))
        table.verticalHeader().setVisible(False)

        # 填充数据
        for row, msg_name in enumerate(sorted(msg_names)):
            # 第1列：报文名称（只读）
            msg_item = QTableWidgetItem(msg_name)
            msg_item.setFlags(msg_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            msg_item.setData(Qt.ItemDataRole.UserRole, (dbc_path, msg_name))
            table.setItem(row, 0, msg_item)

            # 获取报文的信号列表
            signals = []
            if dbc_info:
                try:
                    msg = dbc_info["db"].get_message_by_name(msg_name)
                    signals = [sig.name for sig in msg.signals]
                except KeyError:
                    pass

            # 第2列：校验信号（下拉框）
            check_signal_combo = QComboBox()
            check_signal_combo.addItem("请选择")
            for sig_name in sorted(signals):
                check_signal_combo.addItem(sig_name)

            # 恢复之前的选择
            current_config = self._message_configs.get(dbc_path, {}).get(msg_name, {})
            current_signal = current_config.get("check_signal", "")
            if current_signal:
                index = check_signal_combo.findText(current_signal)
                if index > 0:
                    check_signal_combo.setCurrentIndex(index)

            check_signal_combo.setProperty("col_type", "check_signal")
            table.setCellWidget(row, 1, check_signal_combo)

            # 第3列：Counter信号（下拉框）
            counter_signal_combo = QComboBox()
            counter_signal_combo.addItem("请选择")
            for sig_name in sorted(signals):
                counter_signal_combo.addItem(sig_name)

            # 恢复之前的选择
            current_counter = current_config.get("counter_signal", "")
            if current_counter:
                index = counter_signal_combo.findText(current_counter)
                if index > 0:
                    counter_signal_combo.setCurrentIndex(index)

            counter_signal_combo.setProperty("col_type", "counter_signal")
            table.setCellWidget(row, 2, counter_signal_combo)

            # 第4列：校验方式（下拉框）
            method_combo = QComboBox()
            method_combo.addItems(["crc16", "checkSum"])

            current_method = current_config.get("check_method", "crc16")
            method_combo.setCurrentText(current_method)

            method_combo.setProperty("col_type", "check_method")
            method_combo.currentTextChanged.connect(self._on_method_changed)
            table.setCellWidget(row, 3, method_combo)

            # 第5列：校验参数（根据校验方式动态显示）
            param_widget = self._create_param_widget(row, dbc_path, msg_name, current_method, current_config)
            table.setCellWidget(row, 4, param_widget)

        return table

    def _create_param_widget(self, row: int, dbc_path: str, msg_name: str,
                             method: str, current_config: Dict) -> QWidget:
        """创建校验参数配置控件 - 简化版：只显示配置按钮"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(8)
        layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)  # 确保上下居中

        if method == "crc16":
            # CRC16参数 - 只显示配置按钮
            current_params = current_config.get("check_parameters", {
                "poly": "0x1021",
                "init": "0xFFFF",
                "refIn": "false",
                "refOut": "false",
                "xorOut": "0x0000"
            })

            # 配置按钮（齿轮图标，无背景，上下居中）
            config_btn = QPushButton("⚙")
            config_btn.setFixedSize(40, 24)
            config_btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    color: #3498db;
                    border: none;
                    border-radius: 3px;
                    font-size: 18px;
                    padding: 0px;
                }
                QPushButton:hover {
                    color: #5dade2;
                    background-color: #ecf0f1;
                }
                QPushButton:pressed {
                    color: #2980b9;
                    background-color: #d5dbdb;
                }
            """)
            config_btn.setProperty("row", row)
            config_btn.setProperty("dbc_path", dbc_path)
            config_btn.setProperty("msg_name", msg_name)
            config_btn.clicked.connect(self._on_config_button_clicked)
            layout.addWidget(config_btn)
            layout.addStretch()

            widget.setProperty("method", method)
            widget.setProperty("current_params", current_params)

        else:  # checkSum
            # checkSum不需要参数
            label = QLabel("无需参数")
            label.setStyleSheet("color: #666;")
            layout.addWidget(label)
            layout.addStretch()
            widget.setProperty("method", method)

        return widget

    def _on_config_button_clicked(self) -> None:
        """配置按钮点击 - 弹出参数配置对话框"""
        config_btn = self.sender()
        if not config_btn:
            return

        row = config_btn.property("row")
        dbc_path = config_btn.property("dbc_path")
        msg_name = config_btn.property("msg_name")

        # 获取当前参数
        current_config = self._message_configs.get(dbc_path, {}).get(msg_name, {})
        current_params = current_config.get("check_parameters", {
            "poly": "0x1021",
            "init": "0xFFFF",
            "refIn": "false",
            "refOut": "false",
            "xorOut": "0x0000"
        })

        # 弹出参数配置对话框
        dialog = CRC16ParamConfigDialog(current_params, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_params = dialog.get_params()

            # 更新配置
            if dbc_path not in self._message_configs:
                self._message_configs[dbc_path] = {}
            if msg_name not in self._message_configs[dbc_path]:
                self._message_configs[dbc_path][msg_name] = {}
            self._message_configs[dbc_path][msg_name]["check_parameters"] = new_params

            # 更新widget的属性（参数存储在property中）
            table = self._tab_tables.get(dbc_path)
            if table:
                param_widget = table.cellWidget(row, 4)
                if param_widget:
                    param_widget.setProperty("current_params", new_params)

    def _on_method_changed(self, method: str) -> None:
        """校验方式改变时更新参数控件"""
        method_combo = self.sender()
        if not method_combo:
            return

        # 找到对应的表格和行
        for dbc_path, table in self._tab_tables.items():
            for row in range(table.rowCount()):
                cell_widget = table.cellWidget(row, 3)
                if cell_widget == method_combo:
                    # 更新参数列
                    msg_item = table.item(row, 0)
                    msg_name = msg_item.data(Qt.ItemDataRole.UserRole)[1]

                    current_config = self._message_configs.get(dbc_path, {}).get(msg_name, {})
                    new_param_widget = self._create_param_widget(row, dbc_path, msg_name, method, current_config)
                    table.setCellWidget(row, 4, new_param_widget)
                    break

    def _on_ok(self) -> None:
        """确认保存配置"""
        # 收集所有表格的配置
        for dbc_path, table in self._tab_tables.items():
            if dbc_path not in self._message_configs:
                self._message_configs[dbc_path] = {}

            for row in range(table.rowCount()):
                msg_item = table.item(row, 0)
                msg_name = msg_item.data(Qt.ItemDataRole.UserRole)[1]

                # 获取各列的值
                check_signal_combo = table.cellWidget(row, 1)
                counter_signal_combo = table.cellWidget(row, 2)
                method_combo = table.cellWidget(row, 3)
                param_widget = table.cellWidget(row, 4)

                # 获取原有的sender_node信息（如果有）
                existing_config = self._message_configs.get(dbc_path, {}).get(msg_name, {})
                sender_node = existing_config.get("sender_node", "Unknown")

                # 构建配置（保留sender_node信息）
                config = {
                    "sender_node": sender_node,  # 保留发送节点信息
                    "check_signal": check_signal_combo.currentText() if check_signal_combo.currentIndex() > 0 else "",
                    "counter_signal": counter_signal_combo.currentText() if counter_signal_combo.currentIndex() > 0 else "",
                    "check_method": method_combo.currentText()
                }

                # 收集参数（方案3：参数存储在widget的property中）
                if config["check_method"] == "crc16":
                    # 从widget的property获取当前参数
                    params = param_widget.property("current_params") if param_widget.property("current_params") else {
                        "poly": "0x1021",
                        "init": "0xFFFF",
                        "refIn": "false",
                        "refOut": "false",
                        "xorOut": "0x0000"
                    }
                    config["check_parameters"] = params

                self._message_configs[dbc_path][msg_name] = config

        self.accept()

    def get_configs(self) -> Dict:
        """获取配置数据"""
        return self._message_configs


class ConvertDialog(QDialog):
    """DSL 转换为 Automation Cases 的对话框"""

    def __init__(self, file_count: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle("转换为 Automation Cases")
        self.setMinimumWidth(350)
        self.file_count = file_count
        self.exist_action = "ask"  # ask, overwrite, skip, rename
        self.apply_to_all = False

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)

        # 文件数量提示
        if self.file_count == 1:
            info_label = QLabel("将转换 1 个 DSL 文件")
        else:
            info_label = QLabel(f"将批量转换 {self.file_count} 个 DSL 文件")
        layout.addWidget(info_label)

        layout.addSpacing(10)

        # 输出格式选择
        format_group = QGroupBox("输出格式")
        format_layout = QVBoxLayout(format_group)

        self.py_checkbox = QCheckBox("转换为 Python (.py)")
        self.py_checkbox.setChecked(True)
        format_layout.addWidget(self.py_checkbox)

        self.json_checkbox = QCheckBox("转换为 JSON (.json)")
        self.json_checkbox.setChecked(True)
        format_layout.addWidget(self.json_checkbox)

        layout.addWidget(format_group)

        layout.addSpacing(10)

        # 已存在文件处理
        exist_group = QGroupBox("目标文件已存在时")
        exist_layout = QVBoxLayout(exist_group)

        self.overwrite_radio = QRadioButton("覆盖")
        self.skip_radio = QRadioButton("跳过")
        self.rename_radio = QRadioButton("重命名")

        self.overwrite_radio.setChecked(True)

        exist_layout.addWidget(self.overwrite_radio)
        exist_layout.addWidget(self.skip_radio)
        exist_layout.addWidget(self.rename_radio)

        layout.addWidget(exist_group)

        # 按钮
        button_layout = QHBoxLayout()
        self.convert_btn = QPushButton("转换")
        self.convert_btn.clicked.connect(self.on_convert)
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)

        button_layout.addStretch()
        button_layout.addWidget(self.convert_btn)
        button_layout.addWidget(self.cancel_btn)

        layout.addLayout(button_layout)

    def on_convert(self):
        """转换按钮点击"""
        if not self.py_checkbox.isChecked() and not self.json_checkbox.isChecked():
            QMessageBox.warning(self, "提示", "请至少选择一种输出格式")
            return

        # 设置已存在文件的处理方式
        if self.overwrite_radio.isChecked():
            self.exist_action = "overwrite"
        elif self.skip_radio.isChecked():
            self.exist_action = "skip"
        elif self.rename_radio.isChecked():
            self.exist_action = "rename"

        self.accept()

    def get_options(self) -> Dict[str, Any]:
        """获取转换选项"""
        return {
            "convert_py": self.py_checkbox.isChecked(),
            "convert_json": self.json_checkbox.isChecked(),
            "exist_action": self.exist_action,
            "apply_to_all": self.apply_to_all
        }


class PresetSignalRow(QWidget):
    """预设信号行"""
    removed = pyqtSignal(object)

    def __init__(self, completions_by_kind: Dict[str, List[str]] = None,
                 hier_index_by_kind: Dict[str, Dict] = None,
                 dbc_parser=None, parent=None):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind or {}
        self._hier_index_by_kind = hier_index_by_kind or {}
        self._dbc_parser = dbc_parser
        self._hier = None
        self._value_helper = None
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        # 信号类型选择
        self.kind_combo = QComboBox(self)
        self.kind_combo.addItems(["sys", "env"])
        self.kind_combo.setCurrentText("sys")
        layout.addWidget(self.kind_combo)  # index 0

        # 信号名称输入
        self.name_edit = QLineEdit(self)
        self.name_edit.setPlaceholderText("信号名称（支持补全）")
        layout.addWidget(self.name_edit, 1)  # index 1

        # 使用 _SignalValueHelper 替代 value_edit
        self._value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=layout,
            insert_index=2,  # 在 name_edit 之后
            dbc_parser=self._dbc_parser,
            value_placeholder="信号值",
            value_width=80,
            info_width=80,
        )

        # 保持时间输入
        self.duration_edit = QLineEdit(self)
        self.duration_edit.setPlaceholderText("保持时间/ms")
        self.duration_edit.setMaximumWidth(100)
        layout.addWidget(self.duration_edit)

        # 注释输入（必填）
        self.comment_edit = QLineEdit(self)
        self.comment_edit.setPlaceholderText("注释（必填）")
        layout.addWidget(self.comment_edit, 1)

        # 删除按钮
        btn_del = QPushButton("删除", self)
        btn_del.clicked.connect(lambda: self.removed.emit(self))
        layout.addWidget(btn_del)

        # 设置信号补全
        self._setup_completer()

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._value_helper.update_from_signal_name)
        self.kind_combo.currentTextChanged.connect(lambda _: self._value_helper.update_from_signal_name())

    def _setup_completer(self) -> None:
        """设置信号补全"""
        if self._hier_index_by_kind and hasattr(self, 'name_edit'):
            try:
                self._hier = _HierLineEditCompleter(
                    edit=self.name_edit,
                    kind_getter=lambda: self.kind_combo.currentText(),
                    index_by_kind=self._hier_index_by_kind,
                    allowed_kinds=["env", "sys"],
                    dbc_parser=self._dbc_parser,
                    on_completed=self._value_helper.update_from_signal_name,  # 补全完成时更新值控件
                )
                self.kind_combo.currentTextChanged.connect(self._on_kind_changed)
            except Exception as e:
                print(f"设置补全失败: {e}")

    def _on_kind_changed(self, kind: str) -> None:
        if self._hier:
            self._hier.refresh()
        if kind == "sys":
            self.name_edit.setPlaceholderText("namespace::variable")
        else:
            self.name_edit.setPlaceholderText("CAN X::Message::Signal")

    def set_data(self, data: Dict[str, Any]) -> None:
        """设置数据"""
        signal_name = data.get("signal_name", "")
        kind = "sys"
        name = signal_name
        if signal_name.startswith("sys::"):
            kind = "sys"
            name = signal_name[5:]
        elif signal_name.startswith("env::"):
            kind = "env"
            name = signal_name[5:]

        self.kind_combo.setCurrentText(kind)
        self.name_edit.setText(name)
        self._value_helper.set_value(data.get("signal_value", ""))
        self.duration_edit.setText(data.get("duration", ""))
        self.comment_edit.setText(data.get("comment", ""))
        # 触发更新值控件
        self._value_helper.update_from_signal_name()

    def get_data(self) -> Dict[str, Any]:
        """获取数据"""
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()
        if name and not name.startswith("sys::") and not name.startswith("env::"):
            name = f"{kind}::{name}"
        return {
            "signal_name": name,
            "signal_value": self._value_helper.get_value(),
            "duration": self.duration_edit.text().strip(),
            "comment": self.comment_edit.text().strip()
        }

    def validate(self) -> Tuple[bool, Optional[str]]:
        """验证数据，comment必填，值需在范围内"""
        # comment 必填
        if not self.comment_edit.text().strip():
            return False, "注释为必填项"

        # 验证值是否在范围内
        value_error = self._value_helper.validate_value()
        if value_error:
            return False, value_error

        return True, None


class PresetSettingDialog(QDialog):
    """预设设置对话框"""

    def __init__(self, project_manager, dbc_parser, parent=None):
        super().__init__(parent)
        self.setWindowTitle("设置预设")
        self.resize(900, 650)

        self._project_manager = project_manager
        self._dbc_parser = dbc_parser
        self._completions_by_kind = {}
        self._hier_index_by_kind = {}
        self._scene_mapping_data = {}  # 场景映射表数据

        # 获取补全数据
        self._load_completions()
        # 加载场景映射表数据
        self._load_scene_mappings()

        self._build_ui()
        self._load_preset_data()

    def _load_completions(self) -> None:
        """加载信号补全数据"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return

        completions = []

        # 从系统变量文件获取
        sysvar_files = self._project_manager.get_system_variable_files()
        for sysvar_path in sysvar_files:
            full_path = self._project_manager.get_full_path(sysvar_path)
            if full_path and full_path.exists():
                try:
                    import xml.etree.ElementTree as ET
                    tree = ET.parse(full_path)
                    root = tree.getroot()
                    for elem in root.iter():
                        if elem.get('Name'):
                            namespace = elem.get('Namespace', '')
                            name = elem.get('Name')
                            if namespace:
                                completions.append(f"sys::{namespace}::{name}")
                            else:
                                completions.append(f"sys::{name}")
                except Exception:
                    pass

        # 使用 dbc_parser 获取信号补全
        if self._dbc_parser:
            # 获取 env 类型信号
            env_signals = self._dbc_parser.get_signal_completion("env::", "env")
            completions.extend(env_signals)

            # 获取 sys 类型变量
            sys_vars = self._dbc_parser.get_system_variables()
            for var in sys_vars:
                completions.append(f"sys::{var}")

        # 按类型分组
        self._completions_by_kind = {
            "sys": [c for c in completions if c.startswith("sys::")],
            "env": [c for c in completions if c.startswith("env::")],
        }

        # 构建分层索引
        try:
            from .case_editor import _build_hier_index_by_kind
            self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)
        except Exception as e:
            print(f"构建补全索引失败: {e}")

    def _load_scene_mappings(self) -> None:
        """加载场景映射表数据"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return

        try:
            scene_mappings = self._project_manager.get_scene_mappings()
            self._scene_mapping_data = {}

            for mapping_info in scene_mappings:
                mapping_name = mapping_info.get("name", "")
                if not mapping_name:
                    continue

                file_path = self._project_manager.load_scene_mapping(mapping_name)
                if not file_path or not file_path.exists():
                    continue

                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active

                    # 查找列索引
                    header_row = None
                    for row in ws.iter_rows(min_row=1, max_row=10):
                        row_values = [cell.value for cell in row]
                        if "场景描述" in row_values and "场景文件名" in row_values and "场景编号" in row_values:
                            header_row = row
                            break

                    if not header_row:
                        continue

                    # 获取列索引
                    desc_col = None
                    name_col = None
                    id_col = None
                    dir_col = None
                    for idx, cell in enumerate(header_row):
                        if cell.value == "场景描述":
                            desc_col = idx
                        elif cell.value == "场景文件名":
                            name_col = idx
                        elif cell.value == "场景编号":
                            id_col = idx
                        elif cell.value == "场景目录":
                            dir_col = idx

                    if name_col is None or id_col is None:
                        continue

                    # 读取数据行
                    scene_data = {}
                    common_scene_dir = ""
                    for row in ws.iter_rows(min_row=header_row[0].row + 1):
                        scene_name = row[name_col].value
                        scene_id = row[id_col].value
                        scene_desc = row[desc_col].value if desc_col is not None else ""
                        scene_dir = row[dir_col].value if dir_col is not None else ""

                        if not common_scene_dir and scene_dir:
                            common_scene_dir = str(scene_dir)

                        if scene_name and scene_id is not None:
                            try:
                                scene_id = int(scene_id)
                                final_dir = str(scene_dir) if scene_dir else common_scene_dir
                                scene_data[str(scene_name)] = {
                                    "id": scene_id,
                                    "desc": str(scene_desc) if scene_desc else "",
                                    "dir": final_dir
                                }
                            except (ValueError, TypeError):
                                continue

                    self._scene_mapping_data[mapping_name] = scene_data

                except Exception as e:
                    print(f"加载场景映射表 '{mapping_name}' 失败: {e}")

        except Exception as e:
            print(f"加载场景映射表数据失败: {e}")

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        # 使用选项卡
        from PyQt6.QtWidgets import QTabWidget, QScrollArea
        self.tab_widget = QTabWidget(self)

        # 预设信号选项卡
        signal_tab = QWidget(self.tab_widget)
        signal_tab_layout = QVBoxLayout(signal_tab)

        # 工具栏
        signal_toolbar = QWidget(signal_tab)
        signal_toolbar_layout = QHBoxLayout(signal_toolbar)
        signal_toolbar_layout.setContentsMargins(0, 0, 0, 0)

        signal_hint = QLabel("可添加多条预设信号，用于提前设置信号（适用于大规模测试中有大量相同的信号设置）", signal_tab)
        signal_hint.setStyleSheet("color: #666666;")
        signal_toolbar_layout.addWidget(signal_hint, 1)

        btn_add_signal = QPushButton("添加预设信号", signal_tab)
        btn_add_signal.clicked.connect(self._add_signal_row)
        signal_toolbar_layout.addWidget(btn_add_signal)

        signal_tab_layout.addWidget(signal_toolbar)

        # 信号列表滚动区域
        self.signal_scroll = QScrollArea(signal_tab)
        self.signal_scroll.setWidgetResizable(True)
        self.signal_scroll.setMinimumHeight(200)

        self.signal_container = QWidget(self.signal_scroll)
        self.signal_container.setContentsMargins(0, 0, 0, 0)

        self.signal_layout = QVBoxLayout(self.signal_container)
        self.signal_layout.setContentsMargins(0, 0, 0, 0)
        self.signal_layout.setSpacing(6)
        self.signal_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.signal_scroll.setWidget(self.signal_container)
        signal_tab_layout.addWidget(self.signal_scroll)

        self.tab_widget.addTab(signal_tab, "预设信号")

        # 预设场景选项卡
        scene_tab = QWidget(self.tab_widget)
        scene_tab_layout = QVBoxLayout(scene_tab)

        # 场景映射表选择
        scene_form = QWidget(scene_tab)
        scene_form_layout = QFormLayout(scene_form)

        self.scene_mapping_combo = QComboBox(scene_tab)
        self.scene_mapping_combo.addItem("请选择场景映射表")
        for mapping_name in self._scene_mapping_data.keys():
            self.scene_mapping_combo.addItem(mapping_name)
        self.scene_mapping_combo.currentTextChanged.connect(self._on_scene_mapping_changed)
        scene_form_layout.addRow("场景映射表:", self.scene_mapping_combo)

        # 场景名称选择
        self.scene_name_combo = QComboBox(scene_tab)
        self.scene_name_combo.addItem("请先选择场景映射表")
        self.scene_name_combo.setEnabled(False)
        self.scene_name_combo.currentTextChanged.connect(self._on_scene_name_changed)
        scene_form_layout.addRow("场景名称:", self.scene_name_combo)

        # 场景ID（只读）
        self.scene_id_label = QLabel("-", scene_tab)
        self.scene_id_label.setStyleSheet("color: #666666;")
        scene_form_layout.addRow("场景ID:", self.scene_id_label)

        # 运行时间
        self.runtime_edit = QLineEdit(scene_tab)
        self.runtime_edit.setPlaceholderText("运行时间（毫秒）")
        scene_form_layout.addRow("运行时间:", self.runtime_edit)

        scene_tab_layout.addWidget(scene_form)
        scene_tab_layout.addStretch(1)

        self.tab_widget.addTab(scene_tab, "预设场景")

        layout.addWidget(self.tab_widget)

        # 按钮
        from PyQt6.QtWidgets import QDialogButtonBox
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            self
        )
        buttons.accepted.connect(self._on_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _on_scene_mapping_changed(self, mapping_name: str) -> None:
        """场景映射表选择变化"""
        self.scene_name_combo.clear()
        if mapping_name and mapping_name in self._scene_mapping_data:
            self.scene_name_combo.addItem("请选择场景名称")
            for scene_name in self._scene_mapping_data[mapping_name].keys():
                self.scene_name_combo.addItem(scene_name)
            self.scene_name_combo.setEnabled(True)
        else:
            self.scene_name_combo.addItem("请先选择场景映射表")
            self.scene_name_combo.setEnabled(False)
        self.scene_id_label.setText("-")

    def _on_scene_name_changed(self, scene_name: str) -> None:
        """场景名称选择变化"""
        mapping_name = self.scene_mapping_combo.currentText()
        if mapping_name in self._scene_mapping_data and scene_name in self._scene_mapping_data[mapping_name]:
            scene_info = self._scene_mapping_data[mapping_name][scene_name]
            self.scene_id_label.setText(str(scene_info.get("id", "-")))
        else:
            self.scene_id_label.setText("-")

    def _add_signal_row(self, data: Dict[str, Any] = None) -> None:
        """添加信号行"""
        row = PresetSignalRow(
            self._completions_by_kind,
            self._hier_index_by_kind,
            self._dbc_parser,
            self.signal_container
        )
        row.removed.connect(self._remove_signal_row)
        if data:
            row.set_data(data)
        self.signal_layout.addWidget(row)

    def _remove_signal_row(self, row: PresetSignalRow) -> None:
        """移除信号行"""
        if self.signal_layout.count() <= 1:
            row.set_data({"signal_name": "", "signal_value": "", "comment": ""})
            return
        self.signal_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()

    def _load_preset_data(self) -> None:
        """加载预设数据"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return

        config = self._project_manager.project_config
        automation = config.get("automation", {})
        set_preset = automation.get("set_preset", {})

        # 加载预设信号
        preset_signals = set_preset.get("preset_signals", [])
        for signal_data in preset_signals:
            self._add_signal_row(signal_data)

        # 如果没有预设信号，添加一个空行
        if self.signal_layout.count() == 0:
            self._add_signal_row()

        # 加载预设场景
        preset_scene = set_preset.get("preset_scene", {})
        if preset_scene:
            scene_mapping = preset_scene.get("scene_mapping", "")
            scene_name = preset_scene.get("scene_name", "")
            runtime = preset_scene.get("runtime", "")

            # 设置场景映射表
            if scene_mapping:
                idx = self.scene_mapping_combo.findText(scene_mapping)
                if idx >= 0:
                    self.scene_mapping_combo.setCurrentIndex(idx)

                # 设置场景名称
                if scene_name:
                    idx = self.scene_name_combo.findText(scene_name)
                    if idx >= 0:
                        self.scene_name_combo.setCurrentIndex(idx)

            # 设置运行时间
            if runtime:
                self.runtime_edit.setText(str(runtime))

    def _on_ok(self) -> None:
        """确认保存"""
        if not self._project_manager or not self._project_manager.is_project_open():
            self.reject()
            return

        # 收集预设信号
        preset_signals = []
        for i in range(self.signal_layout.count()):
            item = self.signal_layout.itemAt(i)
            if item and item.widget():
                row = item.widget()
                if isinstance(row, PresetSignalRow):
                    data = row.get_data()
                    if data.get("signal_name"):  # 只保存有信号名的
                        valid, error_msg = row.validate()
                        if not valid and error_msg:
                            QMessageBox.warning(self, "警告", f"预设信号验证失败：{error_msg}")
                            return
                        data["id"] = f"P{len(preset_signals) + 1}"
                        preset_signals.append(data)

        # 收集预设场景
        preset_scene = {}
        scene_mapping = self.scene_mapping_combo.currentText()
        scene_name = self.scene_name_combo.currentText()
        if scene_mapping and scene_mapping != "请选择场景映射表" and scene_name and scene_name != "请选择场景名称":
            preset_scene = {
                "scene_mapping": scene_mapping,
                "scene_name": scene_name,
                "scene_id": self.scene_id_label.text() if self.scene_id_label.text() != "-" else "",
                "runtime": self.runtime_edit.text().strip()
            }

        # 保存到项目配置
        config = self._project_manager.project_config
        if "automation" not in config:
            config["automation"] = {}
        config["automation"]["set_preset"] = {
            "preset_signals": preset_signals,
            "preset_scene": preset_scene
        }
        self._project_manager.save_project()

        self.accept()


class SetTemplateRow(QWidget):
    """SET模板行"""
    removed = pyqtSignal(object)

    def __init__(self, completions_by_kind: Dict[str, List[str]] = None,
                 hier_index_by_kind: Dict[str, Dict] = None,
                 dbc_parser=None, parent=None):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind or {}
        self._hier_index_by_kind = hier_index_by_kind or {}
        self._dbc_parser = dbc_parser
        self._hier = None
        self._value_helper = None
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        # 6列布局：信号类型 | 信号名 | 信号值 | choice解释/值范围 | 注释 | 删除

        # 第1列：信号类型选择
        self.kind_combo = QComboBox(self)
        self.kind_combo.addItems(["sys", "env"])
        self.kind_combo.setCurrentText("sys")
        self.kind_combo.setFixedWidth(60)
        layout.addWidget(self.kind_combo)

        # 第2列：信号名称输入（stretch=2，信号名通常较长）
        self.name_edit = QLineEdit(self)
        self.name_edit.setPlaceholderText("信号名称（支持补全）")
        layout.addWidget(self.name_edit, 2)

        # 第3、4列：使用 _SignalValueHelper（信号值 + choice解释/值范围）
        self._value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=layout,
            insert_index=2,  # 在 name_edit 之后插入第3、4列
            dbc_parser=self._dbc_parser,
            value_placeholder="信号值（如 0x1, 1, true）",
            value_width=100,  # 固定信号值控件宽度
            info_width=120,   # choice解释/值范围标签宽度
        )

        # 第5列：注释输入（必填，stretch=1）
        self.comment_edit = QLineEdit(self)
        self.comment_edit.setPlaceholderText("注释（必填，用于选择模板）")
        layout.addWidget(self.comment_edit, 1)

        # 删除按钮
        btn_del = QPushButton("删除", self)
        btn_del.clicked.connect(lambda: self.removed.emit(self))
        layout.addWidget(btn_del)

        # 设置信号补全
        self._setup_completer()

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._value_helper.update_from_signal_name)
        self.kind_combo.currentTextChanged.connect(lambda _: self._value_helper.update_from_signal_name())

    def _setup_completer(self) -> None:
        """设置信号补全"""
        try:
            # 如果没有传入索引，尝试从completions构建
            if not self._hier_index_by_kind and self._completions_by_kind:
                self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)

            if self._hier_index_by_kind:
                self._hier = _HierLineEditCompleter(
                    edit=self.name_edit,
                    kind_getter=lambda: self.kind_combo.currentText(),
                    index_by_kind=self._hier_index_by_kind,
                    allowed_kinds=["env", "sys"],
                    dbc_parser=self._dbc_parser,
                    on_completed=self._value_helper.update_from_signal_name,  # 补全完成时更新值控件
                )
                self.kind_combo.currentTextChanged.connect(self._on_kind_changed)
        except Exception as e:
            print(f"设置补全失败: {e}")

    def _on_kind_changed(self, kind: str) -> None:
        if self._hier:
            self._hier.refresh()
        if kind == "sys":
            self.name_edit.setPlaceholderText("namespace::variable")
        else:
            self.name_edit.setPlaceholderText("CAN X::Message::Signal")

    def set_data(self, data: Dict[str, Any]) -> None:
        """设置数据"""
        signal_name = data.get("signal_name", "")
        kind = "sys"
        name = signal_name
        if signal_name.startswith("sys::"):
            kind = "sys"
            name = signal_name[5:]
        elif signal_name.startswith("env::"):
            kind = "env"
            name = signal_name[5:]

        self.kind_combo.setCurrentText(kind)
        self.name_edit.setText(name)
        self._value_helper.set_value(data.get("signal_value", ""))
        self.comment_edit.setText(data.get("comment", ""))
        # 触发更新值控件
        self._value_helper.update_from_signal_name()

    def get_data(self) -> Dict[str, Any]:
        """获取数据"""
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()
        if name and not name.startswith("sys::") and not name.startswith("env::"):
            name = f"{kind}::{name}"
        return {
            "signal_name": name,
            "signal_value": self._value_helper.get_value(),
            "comment": self.comment_edit.text().strip()
        }

    def validate(self) -> Tuple[bool, Optional[str]]:
        """验证数据，comment必填，值需在范围内"""
        # comment 必填
        if not self.comment_edit.text().strip():
            return False, "注释为必填项"

        # 验证值是否在范围内
        value_error = self._value_helper.validate_value()
        if value_error:
            return False, value_error

        return True, None


class CheckTemplateRow(QWidget):
    """CHECK模板行"""
    removed = pyqtSignal(object)

    def __init__(self, completions_by_kind: Dict[str, List[str]] = None,
                 hier_index_by_kind: Dict[str, Dict] = None,
                 dbc_parser=None, parent=None):
        super().__init__(parent)
        self._completions_by_kind = completions_by_kind or {}
        self._hier_index_by_kind = hier_index_by_kind or {}
        self._dbc_parser = dbc_parser
        self._hier = None
        self._value_helper = None
        self._list_range_info_label = None
        self._build_ui()

    def _build_ui(self) -> None:
        self._layout = QHBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(6)

        # 信号类型选择
        self.kind_combo = QComboBox(self)
        self.kind_combo.addItems(["sig", "env", "sys"])
        self.kind_combo.setCurrentText("sig")
        self._layout.addWidget(self.kind_combo)  # index 0

        # 信号名称输入
        self.name_edit = QLineEdit(self)
        self.name_edit.setPlaceholderText("信号名称（支持补全）")
        self._layout.addWidget(self.name_edit, 1)  # index 1

        # 值模式选择
        self.mode_combo = QComboBox(self)
        self.mode_combo.addItems(["single", "list", "range"])
        self.mode_combo.setCurrentText("single")
        self._layout.addWidget(self.mode_combo)  # index 2

        # 比较符选择
        self.op_combo = QComboBox(self)
        self.op_combo.addItems(["==", ">", "<", ">=", "<=", "!="])
        self._layout.addWidget(self.op_combo)  # index 3

        # 使用 _SignalValueHelper 替代 value_edit（用于 single 模式）
        self._value_helper = _SignalValueHelper(
            kind_getter=lambda: self.kind_combo.currentText(),
            name_getter=lambda: self.name_edit.text(),
            layout=self._layout,
            insert_index=4,  # 在 op_combo 之后
            dbc_parser=self._dbc_parser,
            value_placeholder="信号值",
            info_width=80,
        )

        # list/range 模式的值输入框
        self.list_edit = QLineEdit(self)
        self.list_edit.setPlaceholderText("列表值（如 1,2,3）")
        self.list_edit.hide()
        self._layout.addWidget(self.list_edit)

        self.range_a_edit = QLineEdit(self)
        self.range_a_edit.setPlaceholderText("a")
        self.range_a_edit.hide()
        self._layout.addWidget(self.range_a_edit)

        self.range_sep = QLabel("..", self)
        self.range_sep.hide()
        self._layout.addWidget(self.range_sep)

        self.range_b_edit = QLineEdit(self)
        self.range_b_edit.setPlaceholderText("b")
        self.range_b_edit.hide()
        self._layout.addWidget(self.range_b_edit)

        # 为 list/range 模式添加范围信息标签
        self._list_range_info_label = QLabel()
        self._list_range_info_label.setStyleSheet("color: #666666; font-size: 11px;")
        self._list_range_info_label.setMinimumWidth(80)
        self._list_range_info_label.hide()
        self._layout.addWidget(self._list_range_info_label)

        # 注释输入（必填）
        self.comment_edit = QLineEdit(self)
        self.comment_edit.setPlaceholderText("注释（必填）")
        self._layout.addWidget(self.comment_edit, 1)

        # 删除按钮
        btn_del = QPushButton("删除", self)
        btn_del.clicked.connect(lambda: self.removed.emit(self))
        self._layout.addWidget(btn_del)

        # 设置信号补全
        self._setup_completer()

        # 模式变化时更新界面
        self.mode_combo.currentTextChanged.connect(self._on_mode_changed)
        self._on_mode_changed(self.mode_combo.currentText())

        # 新增：监听信号名变化，更新值控件
        self.name_edit.textChanged.connect(self._update_value_controls)
        self.kind_combo.currentTextChanged.connect(lambda _: self._update_value_controls())

    def _update_value_controls(self) -> None:
        """根据信号名和模式更新值控件"""
        mode = self.mode_combo.currentText()

        # 先更新 single 模式的 helper
        self._value_helper.update_from_signal_name()

        # 对于 list/range 模式，更新范围信息标签
        if mode in ("list", "range"):
            self._update_list_range_info()

    def _update_list_range_info(self) -> None:
        """为 list/range 模式更新范围信息标签"""
        kind = self.kind_combo.currentText()
        name = self.name_edit.text()

        if kind not in ("sig", "env") or not self._dbc_parser or not name:
            self._list_range_info_label.setText("")
            self._list_range_info_label.hide()
            return

        # 构建完整路径
        full_path = name
        if not name.startswith("CAN "):
            if name.startswith("sig::") or name.startswith("env::"):
                full_path = name
            else:
                full_path = f"CAN {name}"

        info = self._dbc_parser.get_signal_info(full_path, kind)
        if info is None:
            self._list_range_info_label.setText("")
            self._list_range_info_label.hide()
            return

        choices = info.get("choices")
        minimum = info.get("minimum")
        maximum = info.get("maximum")

        if choices:
            vals = list(choices.keys())[:5]
            self._list_range_info_label.setText(f"可选: {', '.join(str(v) for v in vals)}...")
            self._list_range_info_label.show()
        elif minimum is not None and maximum is not None:
            self._list_range_info_label.setText(f"[{minimum} ~ {maximum}]")
            self._list_range_info_label.show()
        else:
            self._list_range_info_label.hide()

    def _setup_completer(self) -> None:
        """设置信号补全"""
        try:
            # 如果没有传入索引，尝试从completions构建
            if not self._hier_index_by_kind and self._completions_by_kind:
                self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)

            if self._hier_index_by_kind:
                self._hier = _HierLineEditCompleter(
                    edit=self.name_edit,
                    kind_getter=lambda: self.kind_combo.currentText(),
                    index_by_kind=self._hier_index_by_kind,
                    allowed_kinds=["sig", "env", "sys"],
                    dbc_parser=self._dbc_parser,
                    on_completed=self._update_value_controls,  # 补全完成时更新值控件
                )
                self.kind_combo.currentTextChanged.connect(self._on_kind_changed)
        except Exception as e:
            print(f"设置补全失败: {e}")

    def _on_kind_changed(self, kind: str) -> None:
        if self._hier:
            self._hier.refresh()
        if kind == "sys":
            self.name_edit.setPlaceholderText("namespace::variable")
        elif kind == "env":
            self.name_edit.setPlaceholderText("CAN X::Message::Signal")
        else:
            self.name_edit.setPlaceholderText("CAN X::Message::Signal")

    def _on_mode_changed(self, mode: str) -> None:
        """根据模式更新界面"""
        if mode == "single":
            self.op_combo.show()
            self._value_helper.show_widgets()
            self.list_edit.hide()
            self.range_a_edit.hide()
            self.range_sep.hide()
            self.range_b_edit.hide()
            self._list_range_info_label.hide()
        elif mode == "list":
            self.op_combo.hide()
            self._value_helper.hide_widgets()
            self.list_edit.show()
            self.range_a_edit.hide()
            self.range_sep.hide()
            self.range_b_edit.hide()
            self._update_list_range_info()
        elif mode == "range":
            self.op_combo.hide()
            self._value_helper.hide_widgets()
            self.list_edit.hide()
            self.range_a_edit.show()
            self.range_sep.show()
            self.range_b_edit.show()
            self._update_list_range_info()

    def set_data(self, data: Dict[str, Any]) -> None:
        """设置数据"""
        signal_name = data.get("signal_name", "")
        kind = "sig"
        name = signal_name
        if signal_name.startswith("sig::"):
            kind = "sig"
            name = signal_name[5:]
        elif signal_name.startswith("sys::"):
            kind = "sys"
            name = signal_name[5:]
        elif signal_name.startswith("env::"):
            kind = "env"
            name = signal_name[5:]

        self.kind_combo.setCurrentText(kind)
        self.name_edit.setText(name)
        self.mode_combo.setCurrentText(data.get("value_mode", "single"))
        self.op_combo.setCurrentText(data.get("operator", "=="))

        # 根据模式设置值
        mode = data.get("value_mode", "single")
        if mode == "single":
            self._value_helper.set_value(data.get("signal_value", ""))
        elif mode == "list":
            self.list_edit.setText(data.get("signal_value", ""))
        elif mode == "range":
            # range 模式可能需要特殊处理，假设值格式为 "a..b"
            val = data.get("signal_value", "")
            if ".." in val:
                parts = val.split("..")
                self.range_a_edit.setText(parts[0].strip())
                self.range_b_edit.setText(parts[1].strip() if len(parts) > 1 else "")
            else:
                self.range_a_edit.setText(val)

        self.comment_edit.setText(data.get("comment", ""))
        self._on_mode_changed(self.mode_combo.currentText())
        # 触发更新值控件
        self._update_value_controls()

    def get_data(self) -> Dict[str, Any]:
        """获取数据"""
        kind = self.kind_combo.currentText()
        name = self.name_edit.text().strip()
        if name and not name.startswith("sig::") and not name.startswith("env::") and not name.startswith("sys::"):
            name = f"{kind}::{name}"

        mode = self.mode_combo.currentText()
        if mode == "single":
            signal_value = self._value_helper.get_value()
        elif mode == "list":
            signal_value = self.list_edit.text().strip()
        elif mode == "range":
            signal_value = f"{self.range_a_edit.text().strip()}..{self.range_b_edit.text().strip()}"
        else:
            signal_value = ""

        return {
            "signal_name": name,
            "value_mode": mode,
            "operator": self.op_combo.currentText(),
            "signal_value": signal_value,
            "comment": self.comment_edit.text().strip()
        }

    def validate(self) -> Tuple[bool, Optional[str]]:
        """验证数据，comment必填，值需在范围内"""
        # comment 必填
        if not self.comment_edit.text().strip():
            return False, "注释为必填项"

        # single 模式验证值是否在范围内
        mode = self.mode_combo.currentText()
        if mode == "single":
            value_error = self._value_helper.validate_value()
            if value_error:
                return False, value_error

        return True, None


class TemplateSettingDialog(QDialog):
    """模板设置对话框"""
    templates_saved = pyqtSignal()  # 模板保存成功信号

    def __init__(self, project_manager, dbc_parser, parent=None):
        super().__init__(parent)
        self.setWindowTitle("模板设置")
        self.resize(1000, 700)

        self._project_manager = project_manager
        self._dbc_parser = dbc_parser
        self._completions_by_kind = {}
        self._hier_index_by_kind = {}

        # 获取补全数据
        self._load_completions()

        self._build_ui()
        self._load_template_data()

    def _load_completions(self) -> None:
        """加载信号补全数据"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return

        completions = []

        # 从系统变量文件获取
        sysvar_files = self._project_manager.get_system_variable_files()
        for sysvar_path in sysvar_files:
            full_path = self._project_manager.get_full_path(sysvar_path)
            if full_path and full_path.exists():
                try:
                    import xml.etree.ElementTree as ET
                    tree = ET.parse(full_path)
                    root = tree.getroot()
                    for elem in root.iter():
                        if elem.get('Name'):
                            namespace = elem.get('Namespace', '')
                            name = elem.get('Name')
                            if namespace:
                                completions.append(f"sys::{namespace}::{name}")
                            else:
                                completions.append(f"sys::{name}")
                except Exception:
                    pass

        # 使用 dbc_parser 获取信号补全
        if self._dbc_parser:
            # 获取 sig 类型信号 (使用空前缀获取所有)
            sig_signals = self._dbc_parser.get_signal_completion("sig::", "sig")
            completions.extend(sig_signals)

            # 获取 env 类型信号
            env_signals = self._dbc_parser.get_signal_completion("env::", "env")
            completions.extend(env_signals)

            # 获取 sys 类型变量
            sys_vars = self._dbc_parser.get_system_variables()
            for var in sys_vars:
                completions.append(f"sys::{var}")

        # 按类型分组
        self._completions_by_kind = {
            "sys": [c for c in completions if c.startswith("sys::")],
            "env": [c for c in completions if c.startswith("env::")],
            "sig": [c for c in completions if c.startswith("sig::")],
        }

        # 构建分层索引
        try:
            from .case_editor import _build_hier_index_by_kind
            self._hier_index_by_kind = _build_hier_index_by_kind(self._completions_by_kind)
        except Exception as e:
            print(f"构建补全索引失败: {e}")

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        # 创建选项卡
        from PyQt6.QtWidgets import QTabWidget
        self.tab_widget = QTabWidget(self)

        # SET模板选项卡
        set_tab = QWidget(self.tab_widget)
        set_layout = QVBoxLayout(set_tab)

        # 工具栏
        set_toolbar = QWidget(set_tab)
        set_toolbar_layout = QHBoxLayout(set_toolbar)
        set_toolbar_layout.setContentsMargins(0, 0, 0, 0)

        set_hint = QLabel("SET模板：定义常用的set信号配置，comment用于在编辑界面中选择模板", set_tab)
        set_hint.setStyleSheet("color: #666666;")
        set_toolbar_layout.addWidget(set_hint, 1)

        btn_add_set = QPushButton("添加SET模板", set_tab)
        btn_add_set.clicked.connect(self._add_set_template_row)
        set_toolbar_layout.addWidget(btn_add_set)

        set_layout.addWidget(set_toolbar)

        # SET模板列表滚动区域
        self.set_scroll = QScrollArea(set_tab)
        self.set_scroll.setWidgetResizable(True)
        self.set_scroll.setMinimumHeight(200)

        self.set_container = QWidget(self.set_scroll)
        self.set_container.setContentsMargins(0, 0, 0, 0)

        self.set_layout = QVBoxLayout(self.set_container)
        self.set_layout.setContentsMargins(0, 0, 0, 0)
        self.set_layout.setSpacing(6)
        self.set_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.set_scroll.setWidget(self.set_container)
        set_layout.addWidget(self.set_scroll)

        self.tab_widget.addTab(set_tab, "SET模板")

        # CHECK模板选项卡
        check_tab = QWidget(self.tab_widget)
        check_layout = QVBoxLayout(check_tab)

        # 工具栏
        check_toolbar = QWidget(check_tab)
        check_toolbar_layout = QHBoxLayout(check_toolbar)
        check_toolbar_layout.setContentsMargins(0, 0, 0, 0)

        check_hint = QLabel("CHECK模板：定义常用的check信号配置，comment用于在编辑界面中选择模板", check_tab)
        check_hint.setStyleSheet("color: #666666;")
        check_toolbar_layout.addWidget(check_hint, 1)

        btn_add_check = QPushButton("添加CHECK模板", check_tab)
        btn_add_check.clicked.connect(self._add_check_template_row)
        check_toolbar_layout.addWidget(btn_add_check)

        check_layout.addWidget(check_toolbar)

        # CHECK模板列表滚动区域
        self.check_scroll = QScrollArea(check_tab)
        self.check_scroll.setWidgetResizable(True)
        self.check_scroll.setMinimumHeight(200)

        self.check_container = QWidget(self.check_scroll)
        self.check_container.setContentsMargins(0, 0, 0, 0)

        self.check_layout = QVBoxLayout(self.check_container)
        self.check_layout.setContentsMargins(0, 0, 0, 0)
        self.check_layout.setSpacing(6)
        self.check_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.check_scroll.setWidget(self.check_container)
        check_layout.addWidget(self.check_scroll)

        self.tab_widget.addTab(check_tab, "CHECK模板")

        layout.addWidget(self.tab_widget)

        # 按钮
        from PyQt6.QtWidgets import QDialogButtonBox
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            self
        )
        buttons.accepted.connect(self._on_ok)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _add_set_template_row(self, data: Dict[str, Any] = None) -> None:
        """添加SET模板行"""
        row = SetTemplateRow(
            self._completions_by_kind,
            self._hier_index_by_kind,
            self._dbc_parser,
            self.set_container
        )
        row.removed.connect(self._remove_set_template_row)
        if data:
            row.set_data(data)
        self.set_layout.addWidget(row)

    def _remove_set_template_row(self, row: SetTemplateRow) -> None:
        """移除SET模板行"""
        if self.set_layout.count() <= 1:
            row.set_data({"signal_name": "", "signal_value": "", "comment": ""})
            return
        self.set_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()

    def _add_check_template_row(self, data: Dict[str, Any] = None) -> None:
        """添加CHECK模板行"""
        row = CheckTemplateRow(
            self._completions_by_kind,
            self._hier_index_by_kind,
            self._dbc_parser,
            self.check_container
        )
        row.removed.connect(self._remove_check_template_row)
        if data:
            row.set_data(data)
        self.check_layout.addWidget(row)

    def _remove_check_template_row(self, row: CheckTemplateRow) -> None:
        """移除CHECK模板行"""
        if self.check_layout.count() <= 1:
            row.set_data({"signal_name": "", "value_mode": "single", "operator": "==", "signal_value": "", "comment": ""})
            return
        self.check_layout.removeWidget(row)
        row.setParent(None)
        row.deleteLater()

    def _load_template_data(self) -> None:
        """加载模板数据"""
        if not self._project_manager or not self._project_manager.is_project_open():
            return

        config = self._project_manager.project_config
        automation = config.get("automation", {})

        # 加载SET模板
        set_templates = automation.get("set_template", {}).get("templates", [])
        for template_data in set_templates:
            self._add_set_template_row(template_data)

        # 如果没有模板，添加一个空行
        if self.set_layout.count() == 0:
            self._add_set_template_row()

        # 加载CHECK模板
        check_templates = automation.get("check_template", {}).get("templates", [])
        for template_data in check_templates:
            self._add_check_template_row(template_data)

        # 如果没有模板，添加一个空行
        if self.check_layout.count() == 0:
            self._add_check_template_row()

    def _on_ok(self) -> None:
        """确认保存"""
        if not self._project_manager or not self._project_manager.is_project_open():
            self.reject()
            return

        # 收集SET模板
        set_templates = []
        for i in range(self.set_layout.count()):
            item = self.set_layout.itemAt(i)
            if item and item.widget():
                row = item.widget()
                if isinstance(row, SetTemplateRow):
                    data = row.get_data()
                    valid, error_msg = row.validate()
                    if valid and data.get("signal_name"):
                        data["id"] = f"ST{len(set_templates) + 1}"
                        set_templates.append(data)
                    elif not valid and data.get("signal_name") and error_msg:
                        QMessageBox.warning(self, "警告", f"SET模板验证失败：{error_msg}")
                        return

        # 收集CHECK模板
        check_templates = []
        for i in range(self.check_layout.count()):
            item = self.check_layout.itemAt(i)
            if item and item.widget():
                row = item.widget()
                if isinstance(row, CheckTemplateRow):
                    data = row.get_data()
                    valid, error_msg = row.validate()
                    if valid and data.get("signal_name"):
                        data["id"] = f"CT{len(check_templates) + 1}"
                        check_templates.append(data)
                    elif not valid and data.get("signal_name") and error_msg:
                        QMessageBox.warning(self, "警告", f"CHECK模板验证失败：{error_msg}")
                        return

        # 保存到项目配置
        config = self._project_manager.project_config
        if "automation" not in config:
            config["automation"] = {}
        config["automation"]["set_template"] = {
            "templates": set_templates
        }
        config["automation"]["check_template"] = {
            "templates": check_templates
        }
        self._project_manager.save_project()

        # 发出信号通知模板已更新
        self.templates_saved.emit()

        QMessageBox.information(self, "成功", "模板设置已保存")
        self.accept()


class RunAutomationDialog(QDialog):
    """运行 Automation Case 配置对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("运行配置")
        self.setModal(True)
        self.resize(350, 150)

        self.output_dir = ""
        self.domain_version = ""

        self.init_ui()

    def init_ui(self) -> None:
        """初始化UI"""
        layout = QVBoxLayout(self)

        # 说明提示
        info_label = QLabel("保存目录将位于 Test Results 下的 trace data和record data 中")
        info_label.setStyleSheet("color: #666;")
        layout.addWidget(info_label)

        # 表单布局
        form_layout = QFormLayout()

        # 目录填写
        self.dir_edit = QLineEdit()
        self.dir_edit.setPlaceholderText("输入保存目录名称")
        form_layout.addRow("保存目录:", self.dir_edit)

        # 域控版本
        self.version_edit = QLineEdit()
        self.version_edit.setPlaceholderText("输入域控版本（如 J6M-V1.0.0）")
        form_layout.addRow("域控版本:", self.version_edit)

        layout.addLayout(form_layout)

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        ok_btn = QPushButton("运行")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)

    def accept(self) -> None:
        """接受对话框"""
        output_dir = self.dir_edit.text().strip()
        domain_version = self.version_edit.text().strip()
        # run_in_background = self.background_checkbox.isChecked()

        if not output_dir:
            QMessageBox.warning(self, "警告", "请填写保存目录")
            return

        if not domain_version:
            QMessageBox.warning(self, "警告", "请填写域控版本")
            return

        # # 如果勾选了后台运行，检查工程文件配置
        # if run_in_background:
        missing_files = self._check_project_files()
        if missing_files:
            QMessageBox.warning(self, "警告", f"后台运行需要配置以下文件：\n{missing_files}")
            return

        self.output_dir = output_dir
        self.domain_version = domain_version
        # self.run_in_background = run_in_background

        super().accept()

    def _check_project_files(self) -> str:
        """检查项目是否配置了CANoe和Simulink工程文件"""
        missing = []

        # 通过 parent (MainWindow) 获取项目配置
        if self.parent() and hasattr(self.parent(), 'project_manager'):
            project_manager = self.parent().project_manager
            project_config = project_manager.project_config

            if project_config:
                # 检查CANoe工程文件 (.cfg)
                canoe_config = project_config.get('canoe', {})
                canoe_project_path = canoe_config.get('project_path', '')
                if not canoe_project_path or not canoe_project_path.lower().endswith('.cfg'):
                    missing.append("CANoe工程文件")

                # 检查Simulink文件 (.slx)
                simulink_config = project_config.get('simulink', {})
                simulink_files = simulink_config.get('files', [])
                has_slx = any(
                    f.get('type') == 'simulink_model' or
                    (f.get('name', '').lower().endswith('.slx') if f.get('name') else False)
                    for f in simulink_files
                )
                if not has_slx:
                    missing.append("Simulink工程文件")

        return '\n'.join(missing)

    def get_config(self) -> Dict[str, Any]:
        """获取配置信息"""
        return {
            "output_dir": self.output_dir,
            "domain_version": self.domain_version,
            # "run_in_background": self.run_in_background
        }


class ECURecordDialog(QDialog):
    """ECU Record configuration dialog"""

    DEFAULT_CONFIG = {
        "record_config": {
            "connection": {
                "host": "192.168.195.3",
                "port": 22,
                "user": "idc",
                "user_pass": "IDC123@byd",
                "root_pass": "idc123@BYD"
            },
            "paths": {
                "work_dir": "/app",
                "init_script": "./script/.release.bash",
                "remote_glob_prefix": "2025"
            },
            "timeouts": {
                "login": 30,
                "command": 40,
                "sftp": 300
            },
            "behavior": {
                "post_start_sleep_sec": 1.0,
                "post_stop_sleep_sec": 1.0,
                "drain_log_sample": False
            },
            "logging": {
                "log_file": "ecu_recorder.log",
                "max_bytes": 2097152,
                "backup_count": 3,
                "mask_secrets": True
            }
        }
    }

    def __init__(self, project_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ECU Record Config")
        self.setModal(True)
        self.resize(600, 500)

        self._project_manager = project_manager
        self._config = {}

        self.init_ui()
        self.load_config()

    def init_ui(self) -> None:
        """Initialize UI"""
        layout = QVBoxLayout(self)

        # Use scroll area for all config sections
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(15)

        # === Connection Group ===
        connection_group = QGroupBox("connection")
        connection_layout = QFormLayout(connection_group)

        self.host_edit = QLineEdit()
        self.host_edit.setPlaceholderText("host")
        connection_layout.addRow("host:", self.host_edit)

        self.port_edit = QLineEdit()
        self.port_edit.setPlaceholderText("port")
        self.port_edit.setValidator(QIntValidator(1, 65535, self))
        connection_layout.addRow("port:", self.port_edit)

        self.user_edit = QLineEdit()
        self.user_edit.setPlaceholderText("user")
        connection_layout.addRow("user:", self.user_edit)

        self.user_pass_edit = QLineEdit()
        self.user_pass_edit.setPlaceholderText("user_pass")
        self.user_pass_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.user_pass_toggle = QPushButton("Show")
        self.user_pass_toggle.setFixedWidth(60)
        self.user_pass_toggle.clicked.connect(lambda: self._toggle_password(self.user_pass_edit, self.user_pass_toggle))
        user_pass_layout = QHBoxLayout()
        user_pass_layout.addWidget(self.user_pass_edit)
        user_pass_layout.addWidget(self.user_pass_toggle)
        connection_layout.addRow("user_pass:", user_pass_layout)

        self.root_pass_edit = QLineEdit()
        self.root_pass_edit.setPlaceholderText("root_pass")
        self.root_pass_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.root_pass_toggle = QPushButton("Show")
        self.root_pass_toggle.setFixedWidth(60)
        self.root_pass_toggle.clicked.connect(lambda: self._toggle_password(self.root_pass_edit, self.root_pass_toggle))
        root_pass_layout = QHBoxLayout()
        root_pass_layout.addWidget(self.root_pass_edit)
        root_pass_layout.addWidget(self.root_pass_toggle)
        connection_layout.addRow("root_pass:", root_pass_layout)

        scroll_layout.addWidget(connection_group)

        # === Paths Group ===
        paths_group = QGroupBox("paths")
        paths_layout = QFormLayout(paths_group)

        self.work_dir_edit = QLineEdit()
        self.work_dir_edit.setPlaceholderText("work_dir")
        paths_layout.addRow("work_dir:", self.work_dir_edit)

        self.init_script_edit = QLineEdit()
        self.init_script_edit.setPlaceholderText("init_script")
        paths_layout.addRow("init_script:", self.init_script_edit)

        self.remote_glob_prefix_edit = QLineEdit()
        self.remote_glob_prefix_edit.setPlaceholderText("remote_glob_prefix")
        paths_layout.addRow("remote_glob_prefix:", self.remote_glob_prefix_edit)

        scroll_layout.addWidget(paths_group)

        # === Timeouts Group ===
        timeouts_group = QGroupBox("timeouts")
        timeouts_layout = QFormLayout(timeouts_group)

        self.login_edit = QLineEdit()
        self.login_edit.setPlaceholderText("login")
        self.login_edit.setValidator(QIntValidator(1, 300, self))
        timeouts_layout.addRow("login:", self.login_edit)

        self.command_edit = QLineEdit()
        self.command_edit.setPlaceholderText("command")
        self.command_edit.setValidator(QIntValidator(1, 600, self))
        timeouts_layout.addRow("command:", self.command_edit)

        self.sftp_edit = QLineEdit()
        self.sftp_edit.setPlaceholderText("sftp")
        self.sftp_edit.setValidator(QIntValidator(1, 1000, self))
        timeouts_layout.addRow("sftp:", self.sftp_edit)

        scroll_layout.addWidget(timeouts_group)

        # === Behavior Group ===
        behavior_group = QGroupBox("behavior")
        behavior_layout = QFormLayout(behavior_group)

        self.post_start_sleep_edit = QLineEdit()
        self.post_start_sleep_edit.setPlaceholderText("post_start_sleep_sec")
        behavior_layout.addRow("post_start_sleep_sec:", self.post_start_sleep_edit)

        self.post_stop_sleep_edit = QLineEdit()
        self.post_stop_sleep_edit.setPlaceholderText("post_stop_sleep_sec")
        behavior_layout.addRow("post_stop_sleep_sec:", self.post_stop_sleep_edit)

        self.drain_log_sample_cb = QCheckBox()
        behavior_layout.addRow("drain_log_sample:", self.drain_log_sample_cb)

        scroll_layout.addWidget(behavior_group)

        # === Logging Group ===
        logging_group = QGroupBox("logging")
        logging_layout = QFormLayout(logging_group)

        self.log_file_edit = QLineEdit()
        self.log_file_edit.setPlaceholderText("log_file")
        logging_layout.addRow("log_file:", self.log_file_edit)

        self.max_bytes_edit = QLineEdit()
        self.max_bytes_edit.setPlaceholderText("max_bytes")
        self.max_bytes_edit.setValidator(QIntValidator(1, 100000000, self))
        logging_layout.addRow("max_bytes:", self.max_bytes_edit)

        self.backup_count_edit = QLineEdit()
        self.backup_count_edit.setPlaceholderText("backup_count")
        self.backup_count_edit.setValidator(QIntValidator(0, 100, self))
        logging_layout.addRow("backup_count:", self.backup_count_edit)

        self.mask_secrets_cb = QCheckBox()
        logging_layout.addRow("mask_secrets:", self.mask_secrets_cb)

        scroll_layout.addWidget(logging_group)

        scroll_layout.addStretch(1)
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        ok_btn = QPushButton("OK")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)

    def _toggle_password(self, line_edit: QLineEdit, button: QPushButton) -> None:
        """Toggle password visibility"""
        if line_edit.echoMode() == QLineEdit.EchoMode.Password:
            line_edit.setEchoMode(QLineEdit.EchoMode.Normal)
            button.setText("Hide")
        else:
            line_edit.setEchoMode(QLineEdit.EchoMode.Password)
            button.setText("Show")

    def load_config(self) -> None:
        """Load config from project"""
        if not self._project_manager or not self._project_manager.is_project_open():
            self._config = self.DEFAULT_CONFIG["record_config"].copy()
        else:
            project_config = self._project_manager.project_config
            self._config = project_config.get("automation", {}).get("record_config", self.DEFAULT_CONFIG["record_config"])

        # Load connection
        connection = self._config.get("connection", self.DEFAULT_CONFIG["record_config"]["connection"])
        self.host_edit.setText(connection.get("host", "192.168.195.3"))
        self.port_edit.setText(str(connection.get("port", 22)))
        self.user_edit.setText(connection.get("user", "idc"))
        self.user_pass_edit.setText(connection.get("user_pass", "IDC123@byd"))
        self.root_pass_edit.setText(connection.get("root_pass", "idc123@BYD"))

        # Load paths
        paths = self._config.get("paths", self.DEFAULT_CONFIG["record_config"]["paths"])
        self.work_dir_edit.setText(paths.get("work_dir", "/app"))
        self.init_script_edit.setText(paths.get("init_script", "./script/.release.bash"))
        self.remote_glob_prefix_edit.setText(paths.get("remote_glob_prefix", "2025"))

        # Load timeouts
        timeouts = self._config.get("timeouts", self.DEFAULT_CONFIG["record_config"]["timeouts"])
        self.login_edit.setText(str(timeouts.get("login", 30)))
        self.command_edit.setText(str(timeouts.get("command", 40)))
        self.sftp_edit.setText(str(timeouts.get("sftp", 300)))

        # Load behavior
        behavior = self._config.get("behavior", self.DEFAULT_CONFIG["record_config"]["behavior"])
        self.post_start_sleep_edit.setText(str(behavior.get("post_start_sleep_sec", 1.0)))
        self.post_stop_sleep_edit.setText(str(behavior.get("post_stop_sleep_sec", 1.0)))
        self.drain_log_sample_cb.setChecked(behavior.get("drain_log_sample", False))

        # Load logging
        logging = self._config.get("logging", self.DEFAULT_CONFIG["record_config"]["logging"])
        self.log_file_edit.setText(logging.get("log_file", "ecu_recorder.log"))
        self.max_bytes_edit.setText(str(logging.get("max_bytes", 2097152)))
        self.backup_count_edit.setText(str(logging.get("backup_count", 3)))
        self.mask_secrets_cb.setChecked(logging.get("mask_secrets", True))

    def accept(self) -> None:
        """Save config and accept"""
        if not self._project_manager or not self._project_manager.is_project_open():
            QMessageBox.warning(self, "Warning", "No project open")
            return

        # Parse numeric values
        try:
            port = int(self.port_edit.text().strip()) if self.port_edit.text().strip() else 22
            login = int(self.login_edit.text().strip()) if self.login_edit.text().strip() else 30
            command = int(self.command_edit.text().strip()) if self.command_edit.text().strip() else 40
            sftp = int(self.sftp_edit.text().strip()) if self.sftp_edit.text().strip() else 300
            post_start_sleep = float(self.post_start_sleep_edit.text().strip()) if self.post_start_sleep_edit.text().strip() else 1.0
            post_stop_sleep = float(self.post_stop_sleep_edit.text().strip()) if self.post_stop_sleep_edit.text().strip() else 1.0
            max_bytes = int(self.max_bytes_edit.text().strip()) if self.max_bytes_edit.text().strip() else 2097152
            backup_count = int(self.backup_count_edit.text().strip()) if self.backup_count_edit.text().strip() else 3
        except ValueError as e:
            QMessageBox.warning(self, "Warning", f"Invalid numeric value: {e}")
            return

        # Build config dict
        self._config = {
            "connection": {
                "host": self.host_edit.text().strip(),
                "port": port,
                "user": self.user_edit.text().strip(),
                "user_pass": self.user_pass_edit.text().strip(),
                "root_pass": self.root_pass_edit.text().strip()
            },
            "paths": {
                "work_dir": self.work_dir_edit.text().strip(),
                "init_script": self.init_script_edit.text().strip(),
                "remote_glob_prefix": self.remote_glob_prefix_edit.text().strip()
            },
            "timeouts": {
                "login": login,
                "command": command,
                "sftp": sftp
            },
            "behavior": {
                "post_start_sleep_sec": post_start_sleep,
                "post_stop_sleep_sec": post_stop_sleep,
                "drain_log_sample": self.drain_log_sample_cb.isChecked()
            },
            "logging": {
                "log_file": self.log_file_edit.text().strip(),
                "max_bytes": max_bytes,
                "backup_count": backup_count,
                "mask_secrets": self.mask_secrets_cb.isChecked()
            }
        }

        # Save to project config
        project_config = self._project_manager.project_config
        if "automation" not in project_config:
            project_config["automation"] = {}
        project_config["automation"]["record_config"] = self._config
        self._project_manager.save_project()

        QMessageBox.information(self, "Success", "ECU Record config saved successfully")
        super().accept()

    def get_config(self) -> Dict[str, Any]:
        """Get config dict"""
        return {"record_config": self._config}


class OutputDialog(QDialog):
    """Output dialog for displaying run logs in real-time"""

    output_signal = pyqtSignal(str)  # Signal for thread-safe output

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Run Output")
        self.setModal(False)  # Non-modal dialog
        self.resize(900, 600)

        # Save original stdout/stderr
        self._original_stdout = sys.stdout
        self._original_stderr = sys.stderr

        # Log file path (will be set when running)
        self._log_file = None
        self._log_handle = None

        # Task state flags
        self._task_running = False  # Whether task is running
        self._stopping = False      # Whether user requested stop

        self.init_ui()
        self._setup_signal()

    def init_ui(self) -> None:
        """Initialize UI"""
        layout = QVBoxLayout(self)

        # Output text area
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setStyleSheet("font-family: Consolas, 'Courier New', monospace; font-size: 12px;")
        layout.addWidget(self.text_edit)

        # Button layout
        button_layout = QHBoxLayout()

        self.stop_btn = QPushButton("停止")
        self.stop_btn.clicked.connect(self._on_stop_clicked)
        button_layout.addWidget(self.stop_btn)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.clicked.connect(self.clear_output)
        button_layout.addWidget(self.clear_btn)

        button_layout.addStretch()

        self.close_btn = QPushButton("关闭")
        self.close_btn.clicked.connect(self.close)
        button_layout.addWidget(self.close_btn)

        layout.addLayout(button_layout)

    def _setup_signal(self) -> None:
        """Setup signal connection for thread-safe output"""
        self.output_signal.connect(self._append_output)

    def _on_stop_clicked(self) -> None:
        """Handle stop button click"""
        if self._task_running and not self._stopping:
            QMessageBox.information(
                self, "提示",
                "自动化测试正在停止中，数据保存后将自动关闭，请耐心等待..."
            )
            self._stopping = True
            self.stop_btn.setEnabled(False)
            self.stop_btn.setText("停止中...")

    def set_task_running(self, running: bool) -> None:
        """Set task running state"""
        self._task_running = running
        if running:
            self.stop_btn.setEnabled(True)
            self.stop_btn.setText("停止")
        else:
            # Task completed, mark as closing to prevent double cleanup in closeEvent
            self._is_closing = True
            self.close()

    def is_stop_requested(self) -> bool:
        """Check if stop is requested (called by worker)"""
        return self._stopping

    def setup_log_file(self, log_dir: str, log_name: str) -> str:
        """
        Setup log file for persistent logging

        Args:
            log_dir: Directory to save log file
            log_name: Name of the log file (without extension)

        Returns:
            Full path to the log file
        """
        Path(log_dir).mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._log_file = os.path.join(log_dir, f"{log_name}_{timestamp}.log")
        self._log_handle = open(self._log_file, "w", encoding="utf-8")
        return self._log_file

    def close_log_file(self) -> None:
        """Close log file handle"""
        if self._log_handle:
            self._log_handle.close()
            self._log_handle = None

    def redirect_output(self) -> None:
        """Redirect stdout and stderr to this dialog"""
        sys.stdout = self
        sys.stderr = self

    def restore_output(self) -> None:
        """Restore original stdout and stderr"""
        sys.stdout = self._original_stdout
        sys.stderr = self._original_stderr

    def write(self, text: str) -> None:
        """Write output (called by print)"""
        if text:
            # Write to original stdout for VSCode compatibility
            self._original_stdout.write(text)
            self._original_stdout.flush()

            # Write to log file
            if self._log_handle:
                self._log_handle.write(text)
                self._log_handle.flush()

            # Emit signal for thread-safe UI update
            self.output_signal.emit(text)

    def flush(self) -> None:
        """Flush (required for stdout interface)"""
        self._original_stdout.flush()
        if self._log_handle:
            self._log_handle.flush()

    def _append_output(self, text: str) -> None:
        """Append text to output area (called via signal)"""
        self.text_edit.insertPlainText(text)
        # Auto scroll to bottom
        scrollbar = self.text_edit.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def clear_output(self) -> None:
        """Clear output area"""
        self.text_edit.clear()

    def closeEvent(self, event) -> None:
        """Handle close event"""
        if self._task_running:
            # Task is still running
            if not self._stopping:
                # First time user requests stop
                QMessageBox.information(
                    self, "提示",
                    "自动化测试正在停止中，数据保存后将自动关闭，请耐心等待..."
                )
                self._stopping = True
                self.stop_btn.setEnabled(False)
                self.stop_btn.setText("停止中...")
            event.ignore()  # Prevent close, wait for task to complete
        else:
            # Task completed, allow close
            # Add protection: check if already closing to avoid double cleanup
            if not hasattr(self, '_is_closing') or not self._is_closing:
                self._is_closing = True
                try:
                    self.restore_output()
                    self.close_log_file()
                except Exception as e:
                    self._original_stdout.write(f"Warning: cleanup error: {e}\n")
            super().closeEvent(event)

    def get_log_file_path(self) -> Optional[str]:
        """Get log file path"""
        return self._log_file